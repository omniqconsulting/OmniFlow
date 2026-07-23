"""
Phase 2 — FMS Core  (§10, §11, §12, §19.3)
Full ticket lifecycle: flow builder, stage transitions, swimlane dashboard,
reassignment, help requests, flagging, manager override, and analytics.
"""
import csv, io, json as _json, logging
from datetime import datetime, timedelta
from typing import List, Optional

logger = logging.getLogger(__name__)

from fastapi import APIRouter, Depends, Form, HTTPException, Query, Request, UploadFile, File
from fastapi.responses import HTMLResponse, RedirectResponse, StreamingResponse, JSONResponse
from fastapi.templating import Jinja2Templates
from sqlalchemy import func
from sqlalchemy.orm import Session

from .database import (
    get_db, new_id,
    Tenant, User, Department, Branch,
    FMSFlow, FMSStage, FMSTicket, FMSStageHistory, FMSEvent, FMSTicketHelper, FMSFlowGroup,
    FMSTicketSplit, FMSFieldEditLog, FMSSplitEvidence,
    LibrarySubmoduleDefinition, TenantDeployedItem,
    Notification, MediaUpload,
    PMSDailyLog, DispatchRecord, InvoiceRecord,
    Customer, Vendor, RawMaterial,
    CustomReferenceList, CustomReferenceItem,
    KnowledgeItem, FMSTicketKnowledgeLink,
)
from .auth import (
    get_current_user, get_current_user_or_redirect,
    require_admin, require_manager,
    require_admin_or_redirect, require_manager_or_redirect,
    get_nav_flags,
)
from .labels import get_labels, DEFAULT_L
from .constants import has_feature, PLAN_LIMITS, BULK_IMPORT_MAX_ROWS, FMS_INACTIVE_STATUSES
from .bulk_common import check_required_headers
from .notifications import (
    create_notification,
    notify_fms_stage_transition,
    notify_fms_flagged,
    send_whatsapp_for_fms_ticket_created,
    send_whatsapp_for_fms_ticket_closed,
    notify_fms_ticket_opened,
    notify_fms_split_created,
)
from .notification_rules import channel_enabled
from .ws_manager import broadcast_sync, FMS_STAGE_TRANSITION
import json as _json_module

# Terminal statuses for FMSTicket/FMSTicketSplit — "not one of these" means still
# active/open. Alias of the shared constant (also used by ai_router.py, analytics.py,
# main.py, scheduler.py, setup_routes.py, superadmin.py, superadmin_library.py).
_INACTIVE_STATUSES = FMS_INACTIVE_STATUSES


def _build_ref_lists_json(tenant_id: str, db) -> str:
    """Combined system entity tables + custom reference lists for field-builder dropdowns."""
    result = []
    _sys = [
        ("__system_customer__",    "Customers",     Customer,    "name"),
        ("__system_vendor__",      "Vendors",       Vendor,      "name"),
        ("__system_rawmaterial__", "Raw Materials", RawMaterial, "name"),
        ("__system_endproduct__",  "EndProduct",    None,        "name"),
        ("__system_department__",  "Departments",   Department,  "name"),
        ("__system_branch__",      "Branches",      Branch,      "name"),
        ("__system_employee__",    "Employees",     User,        "name"),
    ]
    # import EndProduct locally to avoid re-importing at module level
    from .database import EndProduct as _EP
    _sys[3] = ("__system_endproduct__", "End Products", _EP, "name")

    for sys_id, sys_name, model, name_col in _sys:
        rows = db.query(model).filter(
            model.tenant_id == tenant_id,
            model.is_deleted == False,
        ).order_by(getattr(model, name_col)).all()
        items = [getattr(r, name_col) for r in rows if getattr(r, name_col, None)]
        if items:
            result.append({"id": sys_id, "name": sys_name, "items": items, "system": True})

    custom = db.query(CustomReferenceList).filter(
        CustomReferenceList.tenant_id == tenant_id,
        CustomReferenceList.is_deleted == False,
        CustomReferenceList.is_active != False,
    ).order_by(CustomReferenceList.list_name).all()
    for l in custom:
        items = [i.value for i in l.items if i.is_active and not i.is_deleted]
        result.append({"id": l.id, "name": l.list_name, "items": items, "system": False})

    return _json_module.dumps(result)


def _next_fms_display_id(db: Session, tenant: Tenant) -> str:
    """
    Generate the next FMS-only sequential display ID, e.g. F-0042.
    Uses MAX over existing FMS display IDs so the sequence is independent
    from the regular ticket T- counter.
    """
    max_id = db.query(func.max(FMSTicket.display_id)).filter(
        FMSTicket.tenant_id == tenant.id,
        FMSTicket.display_id.isnot(None),
    ).scalar()
    if max_id and max_id.startswith("F-"):
        try:
            next_num = int(max_id[2:]) + 1
        except ValueError:
            next_num = 1
    else:
        next_num = 1
    return f"F-{next_num:04d}"


def _check_fms_flow_limit(db: Session, tenant: Tenant) -> tuple[bool, int, object]:
    """
    Check if tenant can deploy another FMS flow given their plan.
    Returns (allowed, current_count, limit)
    """
    plan = tenant.plan or "STARTER"
    limits = PLAN_LIMITS.get(plan, {})
    max_flows = limits.get("max_fms_flows")
    current = db.query(FMSFlow).filter(
        FMSFlow.tenant_id == tenant.id,
        FMSFlow.is_deleted == False,
    ).count()
    if max_flows is None:
        return True, current, None  # unlimited
    return current < max_flows, current, max_flows

import os
from .templates_env import templates  # shared instance — has all filters





router = APIRouter(prefix="/fms", tags=["FMS"])

# ── Constants ────────────────────────────────────────────────────────────────
MANAGER_OVERRIDE_HOURS = 2   # default configurable override window (§11.3)
FMS_STATUSES = ["ACTIVE", "STAGE_COMPLETE", "IN_TRANSITION",
                "HELP_REQUESTED", "FLAGGED", "ON_HOLD", "COMPLETED", "CLOSED"]
PRIORITIES   = ["LOW", "MEDIUM", "HIGH", "CRITICAL"]


# ── Helpers ──────────────────────────────────────────────────────────────────

def _redirect(path: str):
    return RedirectResponse(path, status_code=302)

def _L(db, user):
    if user is None: return DEFAULT_L
    return get_labels(db, user.tenant_id)

def _unread(db: Session, user: User) -> int:
    return db.query(Notification).filter(
        Notification.user_id == user.id, Notification.is_read == False).count()

def _ctx(request, user, db, **kw):
    tenant = db.query(Tenant).filter(Tenant.id == user.tenant_id).first() if user else None
    return {"request": request, "user": user,
            "L": _L(db, user), "unread": _unread(db, user),
            **get_nav_flags(db, user, tenant),
            **kw}

_LOG_META_SEP = "\x1f"  # unit separator — never appears in normal text/JSON values

def _log(db: Session, ticket_id: str, actor_id: str, event_type: str, detail: str = "", meta: dict | None = None):
    """meta (optional) carries structured data — stage name, assignee, TAT window,
    custom-field values keyed by label, etc. — appended after a control-char
    separator so the Log view can render real columns instead of parsing free
    text, with zero schema change and full backward-compat for old rows."""
    stored = detail
    if meta:
        import json as _json_meta
        stored = f"{detail}{_LOG_META_SEP}{_json_meta.dumps(meta)}"
    db.add(FMSEvent(ticket_id=ticket_id, actor_id=actor_id,
                    event_type=event_type, detail=stored))

def _split_log_meta(detail: str) -> tuple[str, dict]:
    """Inverse of _log's meta encoding. Returns (display_text, meta_dict)."""
    if not detail or _LOG_META_SEP not in detail:
        return detail or "", {}
    text, _, raw_meta = detail.partition(_LOG_META_SEP)
    import json as _json_meta
    try:
        return text, _json_meta.loads(raw_meta)
    except Exception:
        return text, {}

def _fmt_cf(values: dict) -> str:
    """Render a custom-field value dict (ticket or stage columns) for log detail text."""
    if not values:
        return "—"
    return ", ".join(f"{k}={v}" for k, v in values.items())

def _cf_by_label(values: dict, field_defs: list) -> dict:
    """Translate a fid-keyed custom-field value dict to label-keyed, using the
    field definitions (each {'id':..., 'label':...}) — for structured log meta."""
    if not values:
        return {}
    label_by_id = {fd.get("id", ""): fd.get("label") or fd.get("id", "") for fd in (field_defs or [])}
    return {label_by_id.get(fid, fid): val for fid, val in values.items()}

def _fmt_window(start, end) -> str:
    if not start and not end:
        return "—"
    s = start.strftime("%d %b %H:%M") if start else "?"
    e = end.strftime("%d %b %H:%M") if end else "?"
    return f"{s} → {e}"

def _admin_ids(db, tenant_id):
    return [u.id for u in db.query(User).filter(
        User.tenant_id == tenant_id, User.role == "ADMIN",
        User.is_deleted == False).all()]

def _manager_ids_for(db, assignee_id):
    if not assignee_id: return []
    u = db.query(User).get(assignee_id)
    return [u.manager_id] if u and u.manager_id else []

def _planned_dates(ticket, stages, tenant=None) -> dict:
    """Calculate (planned_start, planned_end) for each stage from ticket.created_at + TaT,
    counted in the tenant's office hours (a ticket opened in the evening starts
    its TaT clock at the next working hour, not immediately).
    Returns dict mapping stage_id → (planned_start, planned_end).
    If any stage has no target_tat_hours, that stage and all subsequent get (None, None)."""
    from .notifications import add_business_hours
    sorted_stages = sorted([s for s in stages if not getattr(s, "is_deleted", False)], key=lambda s: s.order)
    result = {}
    cursor = ticket.created_at
    for s in sorted_stages:
        ps = cursor
        if s.target_tat_hours:
            pe = add_business_hours(tenant, cursor, s.target_tat_hours) if tenant else cursor + timedelta(hours=s.target_tat_hours)
        else:
            # No TAT defined — give a 1-minute placeholder so plan dates are always present
            pe = cursor + timedelta(minutes=1)
        result[s.id] = (ps, pe)
        cursor = pe
    return result


def _cross_stage_cf(db: Session, ticket_id: str, stages: list, split_id: str = None, exclude_history_id: str = None) -> dict:
    """Aggregate custom field values from every stage this split has already
    passed through, keyed by both field id (UUID) and field label, so that
    formula columns and 'already captured' field dedup can look up values
    captured in earlier stages — not just the current stage's own fields.

    When split_id is given (a single id, or an iterable of ids — pass a
    split's full lineage from _split_lineage_ids() to also see values
    captured under its ancestors' ids before it existed as its own entity,
    e.g. an auto-split moved split), scoped strictly to those splits' own
    history rows (plus any legacy rows with no split_id, for safety).
    Inheritance across a MANUAL split point is handled NOT here but by
    fms_split_ticket(), which seeds
    a new split's opening history row with a snapshot of everything its
    source split had accumulated at the moment of the split (see that
    function's docstring). A live time-cutoff approach was tried first and
    rejected: a source split's *currently open* row can keep being edited
    after a sibling is carved off it (via the "Enter Data" stage-data
    endpoint), sharing the same entered_at — a cutoff on entered_at can't
    tell "captured before the split" from "captured after, same row" apart.
    Snapshotting at the moment of the split sidesteps that entirely: each
    split's own history is then a fully self-contained, independent copy
    from that point forward, and nested splits (S2 split again into S3)
    compose correctly for free, since S2's own history already carries
    S1's baked-in snapshot."""
    import json as _json
    hist = (
        db.query(FMSStageHistory)
        .filter(FMSStageHistory.ticket_id == ticket_id)
        .order_by(FMSStageHistory.entered_at)
        .all()
    )
    if split_id:
        _split_ids = {split_id} if isinstance(split_id, str) else set(split_id)
        hist = [h for h in hist if h.split_id is None or h.split_id in _split_ids]
    cf_all: dict = {}
    for h in hist:
        if exclude_history_id and h.id == exclude_history_id:
            continue
        if not h.custom_fields_data_json:
            continue
        try:
            cf_data = _json.loads(h.custom_fields_data_json)
        except Exception:
            continue
        cf_all.update(cf_data)  # id-keyed
        src_stage = next((s for s in stages if s.id == h.stage_id), None)
        if src_stage and src_stage.custom_fields_json:
            try:
                for fdef in _json.loads(src_stage.custom_fields_json):
                    fid = fdef.get("id", "")
                    lbl = fdef.get("label", "")
                    if fid and lbl and fid in cf_data:
                        cf_all[lbl] = cf_data[fid]  # label-keyed
            except Exception:
                pass
    return cf_all


def _live_eval_formulas(cf_all: dict, stages: list) -> None:
    """Live-evaluate formula-type custom columns for display, using already-
    captured cross-stage values. Mutates cf_all in place.
    Formula columns are normally only computed and persisted when the stage
    that defines them is transitioned out (see the transition handler below).
    That leaves a display gap: a later stage's formula column that references
    an earlier stage's value shows blank until the later stage itself closes,
    even though the referenced value is already known. Recomputing here from
    the same aggregated cf_all closes that gap for table/stage-view display."""
    for s in stages:
        if not s.custom_fields_json:
            continue
        try:
            fdefs = _json.loads(s.custom_fields_json)
        except Exception:
            continue
        for fdef in fdefs:
            fid = fdef.get("id", "")
            if fdef.get("field_type") != "formula" or not fid or fid in cf_all:
                continue
            steps = fdef.get("formula_steps") or []
            result = None
            for i, step in enumerate(steps):
                raw = cf_all.get(step.get("col_id", ""), "")
                try:
                    val = float(raw)
                except (ValueError, TypeError):
                    result = None
                    break
                if i == 0:
                    result = val
                    continue
                op = step.get("op", "+")
                if op == "+":   result += val
                elif op == "-": result -= val
                elif op == "*": result *= val
                elif op == "/":
                    if val == 0:
                        result = None
                        break
                    result /= val
            if result is not None:
                computed = str(int(result)) if result == int(result) else f"{result:.4f}".rstrip("0")
                cf_all[fid] = computed
                lbl = fdef.get("label", "")
                if lbl:
                    cf_all[lbl] = computed


def _tat_pct(history_row: FMSStageHistory, stage: FMSStage = None) -> Optional[int]:
    """Return 0-100+ percentage of TaT used.
    Prefers ticket-specific planned_end/planned_start from history row;
    falls back to stage.target_tat_hours for legacy rows without a schedule."""
    until = history_row.exited_at or datetime.utcnow()
    elapsed_h = (until - history_row.entered_at).total_seconds() / 3600
    # Use ticket-specific planned window if available
    if history_row.planned_start and history_row.planned_end:
        planned_h = (history_row.planned_end - history_row.planned_start).total_seconds() / 3600
        if planned_h > 0:
            return int(elapsed_h / planned_h * 100)
    # Fall back to flow-level stage target
    if stage and stage.target_tat_hours:
        return int(elapsed_h / stage.target_tat_hours * 100)
    return None

def _can_transition(user: User, ticket: FMSTicket, split=None) -> bool:
    """Admins and managers can always transition; employees only their own stage.
    When split is given, checked against that split's assignee (for multi-split
    tickets); otherwise falls back to the ticket-level cache, unchanged."""
    if user.role in ("ADMIN", "MANAGER"):
        return True
    if split is not None:
        return split.current_assignee_id == user.id
    return ticket.current_assignee_id == user.id

def _can_act_on_ticket(user: User, ticket: FMSTicket, split=None) -> bool:
    """Flow setup option: 'only specific whitelisted employees may open/act
    on tickets in this flow' (FMSFlow.restrict_to_assignee +
    FMSFlow.allowed_opener_ids_json). Admins/managers are always exempt, and
    flows without the flag keep today's looser per-action checks (this is an
    additional gate, not a replacement for them)."""
    if user.role in ("ADMIN", "MANAGER"):
        return True
    if not (ticket.flow and ticket.flow.restrict_to_assignee):
        return True
    import json as _json_gate
    try:
        allowed_ids = set(_json_gate.loads(ticket.flow.allowed_opener_ids_json or "[]"))
    except Exception:
        allowed_ids = set()
    return user.id in allowed_ids

def _stage_default_assignee_ids(stage) -> list:
    """A stage can now have several eligible default assignees (setup page:
    multi-select). default_assignee_ids_json is the source of truth when
    set; falls back to the legacy single default_assignee_id column for
    stages configured before this field existed."""
    if stage is None:
        return []
    if stage.default_assignee_ids_json:
        try:
            ids = _json.loads(stage.default_assignee_ids_json)
            if ids:
                return ids
        except Exception:
            pass
    return [stage.default_assignee_id] if stage.default_assignee_id else []

def _stage_default_assignee(stage) -> Optional[str]:
    """Single-value pick for call sites that pre-fill one assignee (e.g. a
    ticket's current_assignee_id) — the first configured default. Whoever
    creates/transitions the ticket can still change it before saving."""
    ids = _stage_default_assignee_ids(stage)
    return ids[0] if ids else None

def _mark_completed_by(ticket, user_id: str) -> None:
    """Record who actually performed the completing action, the first time
    the ticket reaches COMPLETED — distinct from completed_at (when)."""
    if ticket.status == "COMPLETED" and not ticket.completed_by_id:
        ticket.completed_by_id = user_id

def _can_create_in_flow(user: User, flow) -> bool:
    """Ticket creation is normally manager/admin-only, but a flow's
    'Allowed Employees' whitelist (restrict_to_assignee +
    allowed_opener_ids_json — set up to let specific employees open/act on
    that flow's tickets) is meant to fully unlock the flow for them,
    including creating new tickets in it — not just acting on existing
    ones. An employee not on any such whitelist still can't create."""
    if user.role in ("ADMIN", "MANAGER"):
        return True
    if not flow or not flow.restrict_to_assignee:
        return False
    import json as _json_gate
    try:
        allowed_ids = set(_json_gate.loads(flow.allowed_opener_ids_json or "[]"))
    except Exception:
        allowed_ids = set()
    return user.id in allowed_ids

def _get_ticket(db, ticket_id, tenant_id) -> FMSTicket:
    t = db.query(FMSTicket).filter(
        FMSTicket.id == ticket_id,
        FMSTicket.tenant_id == tenant_id,
        FMSTicket.is_deleted == False,
    ).first()
    if not t:
        raise HTTPException(404, "Ticket not found")
    return t

def _open_history(db, ticket_id, split_id: str = None) -> Optional[FMSStageHistory]:
    """The currently active stage history row (no exited_at).
    For single-split tickets (the common case) this is unambiguous and behaves
    exactly as before. When split_id is given, scoped to that split. When not
    given on a multi-split ticket, returns the most-recently-entered open row
    (a reasonable "primary" pick for legacy call sites — see _open_histories_for_ticket
    for code that must consider every active split)."""
    q = db.query(FMSStageHistory).filter(
        FMSStageHistory.ticket_id == ticket_id,
        FMSStageHistory.exited_at == None,
    )
    if split_id:
        q = q.filter(FMSStageHistory.split_id == split_id)
    return q.order_by(FMSStageHistory.entered_at.desc()).first()

def _open_histories_for_ticket(db, ticket_id) -> list:
    """All currently-open stage history rows for a ticket — one per active split.
    Use this (not _open_history) anywhere that must not assume a single active stage."""
    return db.query(FMSStageHistory).filter(
        FMSStageHistory.ticket_id == ticket_id,
        FMSStageHistory.exited_at == None,
    ).order_by(FMSStageHistory.entered_at.desc()).all()

def _stage_cumulative_qty(db, ticket_id, stage_id) -> int:
    result = db.query(func.sum(FMSStageHistory.qty_completed)).filter(
        FMSStageHistory.ticket_id == ticket_id,
        FMSStageHistory.stage_id  == stage_id,
    ).scalar()
    return result or 0


# ── Phase 0: Split Flows — core helpers ─────────────────────────────────────

def _active_splits(db, ticket_id) -> list:
    """Leaf, still-live splits for a ticket. A split is marked is_deleted=True
    the instant its qty is fully carved into a child split (brief §5) — so
    'active' == 'is_deleted == False', ordered oldest-first for stable
    split_label numbering."""
    return db.query(FMSTicketSplit).filter(
        FMSTicketSplit.ticket_id == ticket_id,
        FMSTicketSplit.is_deleted == False,
    ).order_by(FMSTicketSplit.created_at).all()


def _split_lineage_ids(db, ticket_id: str, split_id: str) -> list:
    """Walk parent_split_id upward from `split_id` to the root split (S1).
    Evidence uploaded on an earlier stage attaches to the split that existed
    at that point in time — once an auto-split carves off a new split entity,
    that history (and its evidence) belongs to an ancestor, not the new
    split's own id. Anything scoped to "this split" that should still be
    reachable after the ticket has moved on (evidence, in particular) needs
    to look across the whole lineage, not just the current split row."""
    ids = []
    all_splits = {s.id: s for s in db.query(FMSTicketSplit).filter(
        FMSTicketSplit.ticket_id == ticket_id,
    ).all()}
    cur = all_splits.get(split_id)
    seen = set()
    while cur and cur.id not in seen:
        seen.add(cur.id)
        ids.append(cur.id)
        cur = all_splits.get(cur.parent_split_id) if cur.parent_split_id else None
    return ids


def _ensure_ticket_has_split(db, ticket: FMSTicket) -> FMSTicketSplit:
    """Every ticket must have exactly one split covering its progress at all
    times (brief §3) — this is the backfill/self-heal path for tickets that
    existed before Phase 0, and a no-op safety net after that. Idempotent:
    if the ticket already has an active split, returns it unchanged.

    Also backfills split_id on any of the ticket's pre-existing FMSStageHistory
    rows that don't have one yet (deterministic 1:1, since only one history
    line was ever open per ticket before splits existed — brief §4)."""
    existing = _active_splits(db, ticket.id)
    if existing:
        return existing[0]

    split = FMSTicketSplit(
        tenant_id=ticket.tenant_id,
        ticket_id=ticket.id,
        split_label="S1",
        qty=ticket.target_qty,
        current_stage_id=ticket.current_stage_id,
        current_assignee_id=ticket.current_assignee_id,
        status=ticket.status if ticket.status in FMS_STATUSES else "ACTIVE",
    )
    db.add(split)
    db.flush()

    db.query(FMSStageHistory).filter(
        FMSStageHistory.ticket_id == ticket.id,
        FMSStageHistory.split_id == None,
    ).update({FMSStageHistory.split_id: split.id})

    return split


def _init_first_split(db, ticket: FMSTicket, stage_id: str, assignee_id: str) -> FMSTicketSplit:
    """Every newly-created ticket gets exactly one split ('S1') covering its
    full target_qty at the starting stage — brief §3. Call this right after
    the ticket row is flushed (has an id) and before creating its first
    FMSStageHistory row, which should then carry split_id=<this split's id>."""
    split = FMSTicketSplit(
        tenant_id=ticket.tenant_id, ticket_id=ticket.id,
        split_label="S1", qty=ticket.target_qty,
        current_stage_id=stage_id, current_assignee_id=assignee_id,
        status="ACTIVE",
    )
    db.add(split)
    db.flush()
    return split


def _resolve_linked_flow(db, tenant_id: str, library_flow_id: str):
    """Resolve a LibraryFlowTemplate id (stored on FMSFlow.next_library_flow_id /
    FMSStage.linked_library_flow_id) to this tenant's live deployed FMSFlow, if any."""
    if not library_flow_id:
        return None
    return db.query(FMSFlow).filter(
        FMSFlow.tenant_id == tenant_id,
        FMSFlow.library_flow_id == library_flow_id,
        FMSFlow.is_active == True,
        FMSFlow.is_deleted == False,
    ).first()


def _spawn_linked_ticket(db, source: FMSTicket, target_flow: FMSFlow, user: User, title: str = None) -> FMSTicket:
    """Create a new FMSTicket on target_flow, carrying over title/description/
    priority/qty from `source`. Seeds the first stage/split exactly like a
    normal manual ticket creation (_init_first_split). Used by both
    close-and-continue and send-to-linked-flow."""
    first_stage = min(
        [s for s in target_flow.stages if not s.is_deleted],
        key=lambda s: s.order, default=None,
    )
    if not first_stage:
        raise HTTPException(400, f"Linked flow '{target_flow.name}' has no stages")
    assignee_id = first_stage.default_assignee_id or source.created_by_id

    new_ticket = FMSTicket(
        tenant_id=user.tenant_id, flow_id=target_flow.id,
        current_stage_id=first_stage.id,
        title=title or source.title,
        description=source.description,
        priority=source.priority,
        target_qty=source.target_qty, qty_unit=source.qty_unit,
        current_assignee_id=assignee_id,
        created_by_id=user.id, status="ACTIVE",
    )
    db.add(new_ticket)
    db.flush()

    tenant = db.query(Tenant).get(user.tenant_id)
    new_ticket.display_id = _next_fms_display_id(db, tenant)

    split = _init_first_split(db, new_ticket, first_stage.id, assignee_id)
    db.add(FMSStageHistory(
        ticket_id=new_ticket.id, split_id=split.id, stage_id=first_stage.id,
        stage_name=first_stage.name, assignee_id=assignee_id,
        direction="FORWARD",
    ))
    _log(db, new_ticket.id, user.id, "CREATED",
         f"Auto-created from {source.display_id} on '{source.flow.name}'")
    _log(db, new_ticket.id, user.id, "STAGE_ENTERED", f"Stage: {first_stage.name}")
    return new_ticket


def _notify_linked_parent_if_ready(db, ticket: FMSTicket) -> None:
    """When a ticket that was spawned via 'send to linked flow' reaches
    COMPLETED/CLOSED, flag the original (parent) ticket as ready to resume.
    Resuming stays a manual action (existing 'resume' sub-action) — this only
    surfaces it."""
    if not ticket.linked_parent_ticket_id or ticket.status not in ("COMPLETED", "CLOSED"):
        return
    parent = db.query(FMSTicket).get(ticket.linked_parent_ticket_id)
    if not parent or parent.linked_child_ticket_id != ticket.id:
        return
    parent.is_flagged = True
    parent.flagged_reason = f"Linked ticket {ticket.display_id} completed — ready to resume"


def _next_split_label(db, ticket_id) -> str:
    """Next auto-numbered split label (S1, S2, ...) — counts every split ever
    created for this ticket (including consumed/inactive ones) so labels never
    collide even after a split is fully carved away."""
    n = db.query(FMSTicketSplit).filter(FMSTicketSplit.ticket_id == ticket_id).count()
    return f"S{n + 1}"


# ── FMS Auto-Split Engine (R1-R6) ────────────────────────────────────────────

def _next_auto_split_sequence(db, ticket_id: str, parent_split_id: str) -> int:
    """Sibling order under a given parent split, for split_display_id suffixes."""
    n = db.query(FMSTicketSplit).filter(
        FMSTicketSplit.ticket_id == ticket_id,
        FMSTicketSplit.parent_split_id == parent_split_id,
    ).count()
    return n + 1


def _resolve_split_field_value(field_id: str, custom_fields_data: dict,
                                formula_lookup: dict, fallback):
    """Resolve a configured split_target_field/split_actual_field's numeric
    value from the entry just submitted (custom_fields_data), falling back to
    cross-stage/ticket values (formula_lookup), falling back to `fallback`
    when unconfigured or unparsable. `fallback` is returned as-is (not
    coerced) so callers can pass None to detect "no usable value"."""
    if not field_id:
        return float(fallback) if fallback is not None else None
    raw = None
    if custom_fields_data and field_id in custom_fields_data:
        raw = custom_fields_data.get(field_id)
    elif formula_lookup and field_id in formula_lookup:
        raw = formula_lookup.get(field_id)
    if raw is None:
        return float(fallback) if fallback is not None else None
    try:
        return float(raw)
    except (TypeError, ValueError):
        return float(fallback) if fallback is not None else None


def _evaluate_auto_split(db, ticket, split, cur_stage, qty: int,
                          custom_fields_data: dict, formula_lookup: dict,
                          next_stage_id: str, new_assignee_id: str, user):
    """
    FMS Auto-Split Engine — inline evaluation at the point a stage data entry
    is submitted (brief §5), called from the FORWARD/non-terminal branch of
    fms_transition, right before the current split progresses to next_stage.

    R1  Opt-in: no-op unless cur_stage.split_enabled.
    R2  entered_value < target_value -> auto-split: the newly-received
        portion moves forward (returned split); the shortfall remains on
        `split` itself, which stays at cur_stage as a remainder
        (is_remainder=True) and keeps an OPEN stage-history row there so it
        can receive further entries.
    R3  Recursive: each subsequent partial entry against that same remainder
        creates another moved-forward split — no upper bound, handled simply
        by this function being called again on the next submit.
    R4  Cumulative across visits: state lives on the persistent remainder
        split row (never reset), so non-linear re-entry to this stage just
        resumes it via the normal _open_history()/split lookup path.
    R5  Over-delivery (entered_value >= target_value): no split; the caller's
        existing progression logic just moves the whole `split` forward /
        completes it cleanly — no flag, no exception UI.

    The configured "actual/entered" field is THIS VISIT'S INCREMENT — the
    quantity that belongs to the split being worked on right now, not a
    running cumulative total (asking users for a cumulative figure was
    confusing once a ticket had multiple splits in flight, since each split
    only ever has its own partial fulfilment in front of it, not the whole
    ticket's history). The engine tracks the true cumulative internally on
    the split row: cumulative_val = split.last_cumulative_entered + delta.
    `target_value` is the fixed original target (falls back to
    ticket.target_qty), not a per-visit shrinking number — the remainder's
    own `qty` is what tracks "how much is still outstanding" for display.

    Returns the FMSTicketSplit that should progress to next_stage_id: either
    `split` unchanged (no split triggered) or a newly created moved-forward
    split (remainder case).
    """
    if not getattr(cur_stage, "split_enabled", False):
        return split

    target_val = _resolve_split_field_value(
        cur_stage.split_target_field, custom_fields_data, formula_lookup, ticket.target_qty)
    delta = _resolve_split_field_value(
        cur_stage.split_actual_field, custom_fields_data, formula_lookup, qty)

    if target_val is None or delta is None:
        return split  # unconfigured / unparsable — behave as today

    prev_cumulative = split.last_cumulative_entered or 0
    if delta <= 0:
        # Nothing entered this visit (zero/blank/negative) — nothing new to
        # split off; leave state as-is.
        return split
    cumulative_val = prev_cumulative + delta

    if cumulative_val >= target_val:
        # R5: fully satisfied (or over-delivered) as of this visit — the
        # whole remainder `split` completes and moves forward as-is, no new
        # split entity, no flag/exception UI. entered_value records the
        # TOTAL received across every visit (not just this final increment)
        # since that's what "Value Entered" means once the split is done —
        # a partial-visit number here would misleadingly look like the split
        # only ever received its last delta. qty is set to this visit's own
        # increment (mirroring the moved-split case below) since everything
        # entered on earlier visits already left as separate moved splits —
        # this is the piece of the target that completes it, and is what the
        # Target Qty column falls back to for a split with no
        # target_value_at_split of its own.
        split.entered_value = cumulative_val
        split.last_cumulative_entered = cumulative_val
        split.qty = int(delta) if delta == int(delta) else delta
        split.is_remainder = False
        return split

    # R2/R3: shortfall -> create the moved-forward split for THIS VISIT's
    # increment (delta), not the running total.
    base_display = split.split_display_id or ticket.display_id or ticket.id[:8]
    seq = _next_auto_split_sequence(db, ticket.id, split.id)
    moved = FMSTicketSplit(
        tenant_id=ticket.tenant_id, ticket_id=ticket.id,
        root_ticket_id=split.root_ticket_id or ticket.id,
        parent_split_id=split.id,
        split_label=_next_split_label(db, ticket.id),
        split_display_id=f"{base_display}-{seq}",
        split_sequence=seq,
        split_stage_id=cur_stage.id,
        target_value_at_split=target_val,
        entered_value=delta,
        is_remainder=False,
        is_auto_split=True,
        qty=int(delta),
        current_stage_id=next_stage_id,
        current_assignee_id=new_assignee_id,
        status="ACTIVE",
    )
    db.add(moved)
    db.flush()

    # R4: the remainder stays on the ORIGINAL split row (never a new entity),
    # at the current stage, with a fresh open history row awaiting the next
    # entry — this is what makes cumulative-across-visits state work without
    # any extra bookkeeping: the normal _open_history(split_id=split.id, ...)
    # lookup just finds it next time. last_cumulative_entered is what lets
    # the NEXT visit compute its own delta correctly.
    remaining_qty = target_val - cumulative_val
    split.qty = int(remaining_qty) if remaining_qty == int(remaining_qty) else remaining_qty
    # Explicitly cleared (not just "left alone") — the remainder itself
    # hasn't been "entered" yet, it's still open awaiting further entries.
    # entered_value belongs to the moved-forward split that delta actually
    # produced; the remainder's row shows "—" (like target_value_at_split
    # already does) until it fully closes. This used to be set to the last
    # delta, which made the splits table show a confusing stale number on
    # the remainder — explicit None here also self-heals any row that still
    # carries that stale value from before this fix.
    split.entered_value = None
    split.last_cumulative_entered = cumulative_val
    split.is_remainder = True
    split.status = "ACTIVE"
    split.current_stage_id = cur_stage.id
    split.updated_at = datetime.utcnow()
    db.add(FMSStageHistory(
        ticket_id=ticket.id, split_id=split.id, stage_id=cur_stage.id,
        stage_name=cur_stage.name, assignee_id=split.current_assignee_id,
        direction="FORWARD", from_stage_id=cur_stage.id, from_stage_name=cur_stage.name,
    ))

    _log(db, ticket.id, user.id, "SPLIT_CREATED",
         f"Auto-split at '{cur_stage.name}': +{delta:g} this visit (cumulative {cumulative_val:g} of target {target_val:g}) — "
         f"{moved.split_display_id} moves forward, remainder "
         f"({remaining_qty:g}) stays as {split.split_display_id or split.split_label}",
         meta={"moved_split_id": moved.id, "remainder_split_id": split.id,
               "entered_value": delta, "cumulative_value": cumulative_val, "target_value": target_val,
               "auto_split": True})

    # §5/§9-E: real-time broadcast; §5/Section-9-F non-blocking wrapper — a
    # notification failure must never roll back or block split creation.
    try:
        admins = _admin_ids(db, ticket.tenant_id)
        managers = _manager_ids_for(db, new_assignee_id)
        notify_fms_split_created(
            ticket.tenant_id, ticket.id, ticket.display_id or ticket.id[:8],
            moved.split_display_id, cur_stage.name, user.id,
            admins, managers, new_assignee_id)
    except Exception:
        logger.exception("notify_fms_split_created failed (non-blocking, split already committed)")

    return moved


_STATUS_PRIORITY = ["FLAGGED", "HELP_REQUESTED", "ON_HOLD", "IN_TRANSITION",
                     "STAGE_COMPLETE", "ACTIVE"]

def _rollup_ticket_status(splits: list) -> str:
    """Brief §6: ticket status is derived from its (leaf) splits once there's
    more than one. COMPLETED only when every leaf split is terminal
    (COMPLETED/CLOSED); otherwise the 'busiest' meaningful state wins, so a
    manager scanning the ticket list sees e.g. FLAGGED without opening every
    ticket."""
    if not splits:
        return "ACTIVE"
    statuses = {s.status for s in splits}
    if statuses <= {"COMPLETED", "CLOSED"}:
        return "COMPLETED" if "COMPLETED" in statuses or len(statuses) == 1 else "CLOSED"
    for candidate in _STATUS_PRIORITY:
        if candidate in statuses:
            return candidate
    return "ACTIVE"


def _sync_ticket_cache(db, ticket: FMSTicket, splits: list = None) -> None:
    """Keeps FMSTicket.current_stage_id/current_assignee_id/status as a
    convenience cache mirroring the splits, per brief §3. For the common
    single-split case this reproduces today's exact behavior byte-for-byte.
    For multi-split tickets these fields stop being authoritative (downstream
    code should use the split rows / rollup status directly) — we still set
    them defensively to the furthest-along split so nothing reading them
    directly sees wildly stale data."""
    if splits is None:
        splits = _active_splits(db, ticket.id)
    ticket.status = _rollup_ticket_status(splits)
    if not splits:
        return
    if len(splits) == 1:
        primary = splits[0]
    else:
        # "Furthest along" = highest stage order; ties broken by most recently updated
        def _order(s):
            return (s.current_stage.order if s.current_stage else -1, s.updated_at or datetime.min)
        primary = max(splits, key=_order)
    ticket.current_stage_id = primary.current_stage_id
    ticket.current_assignee_id = primary.current_assignee_id
    if ticket.status == "COMPLETED" and not ticket.completed_at:
        ticket.completed_at = datetime.utcnow()


def _check_qty_discrepancy(db, ticket: FMSTicket, actor_id: str = None) -> bool:
    """Brief §5: soft warning, never blocks. Sums qty across leaf splits
    (is_deleted==False, status != CLOSED — a CLOSED split is an intentional
    write-off, not part of the expected total) and compares to target_qty.
    Logs QTY_DISCREPANCY only on the False→True transition to avoid event-log
    spam; silently clears the flag on match. Returns the resulting flag value."""
    if ticket.target_qty is None:
        return False
    splits = [s for s in _active_splits(db, ticket.id) if s.status != "CLOSED"]
    actual = sum(s.qty or 0 for s in splits)
    mismatched = actual != ticket.target_qty
    if mismatched and not ticket.has_qty_discrepancy:
        _log(db, ticket.id, actor_id, "QTY_DISCREPANCY",
             f"Expected {ticket.target_qty}, found {actual} across {len(splits)} active split(s)")
    ticket.has_qty_discrepancy = mismatched
    return mismatched


def _ticket_closing_rule_check(db, ticket: FMSTicket, stages: list, rule: dict,
                                in_progress_split_id=None,
                                in_progress_values: dict = None):
    """A flow's closing_rule_json (e.g. 'excess quantity == 0') gates whether
    a ticket may reach its terminal stage. Per product decision: once a
    ticket has more than one split, the rule must hold for the ticket AS A
    WHOLE — the aggregate (sum) of the rule's column across every active
    split — not just whichever single split happens to be completing right
    now. For the common single-split case this is exactly the old
    behavior (sum of one number is that number).

    Returns (ok: bool, error_message: str | None).

    Boundary cases handled:
      - CLOSED splits are write-offs (same convention as the qty-discrepancy
        check) — excluded from the aggregate entirely.
      - If ANY active leaf split hasn't captured a numeric value for the
        rule's column yet, the aggregate is unknowable — this blocks with a
        clear message naming the split, rather than silently treating the
        missing value as 0 and under-counting.
      - A ticket with zero active leaf splits (shouldn't happen given
        _ensure_ticket_has_split, but defensive) passes — nothing to check.
      - The split(s) currently mid-transition may have form values not yet
        committed to the DB; pass their id(s) via in_progress_split_id
        (a single id or an iterable of ids — bulk actions can move several
        splits at once with the same captured values) plus
        in_progress_values so they're judged on what they're *about* to
        have, not stale DB state.
    """
    col_id = rule.get("col_id")
    if not col_id:
        return True, None
    if isinstance(in_progress_split_id, str):
        in_progress_ids = {in_progress_split_id}
    else:
        in_progress_ids = set(in_progress_split_id or [])
    leaf_splits = [s for s in _active_splits(db, ticket.id) if s.status != "CLOSED"]
    if not leaf_splits:
        return True, None
    total = 0.0
    for sp in leaf_splits:
        if sp.id in in_progress_ids:
            lookup = in_progress_values or {}
        else:
            lookup = _cross_stage_cf(db, ticket.id, stages, split_id=_split_lineage_ids(db, ticket.id, sp.id))
        raw = lookup.get(col_id, "")
        try:
            total += float(raw)
        except (ValueError, TypeError):
            return False, (
                f"Cannot close ticket: split {sp.split_label} hasn't captured "
                f"the closing-rule column yet."
            )
    op = rule.get("op", "==")
    target = rule.get("value", 0)
    ok = (
        total == target if op == "==" else
        total != target if op == "!=" else
        total <  target if op == "<"  else
        total <= target if op == "<=" else
        total >  target if op == ">"  else
        total >= target if op == ">=" else True
    )
    if not ok:
        return False, (
            f"Cannot close ticket: closing rule not met across all splits "
            f"(aggregate {total} {op} {target} is false)."
        )
    return True, None


# ── 2-B: Flow Builder ────────────────────────────────────────────────────────

@router.get("/flows", response_class=HTMLResponse)
def fms_flows(request: Request, user: User = Depends(require_admin_or_redirect),
              db: Session = Depends(get_db)):
    """2-B-1: Flow list — admin only."""
    flows = db.query(FMSFlow).filter(
        FMSFlow.tenant_id == user.tenant_id,
        FMSFlow.is_deleted == False,
    ).order_by(FMSFlow.created_at).all()

    # Annotate with counts
    flow_info = []
    for f in flows:
        active_stages = [s for s in f.stages if not s.is_deleted]
        active_tickets = db.query(FMSTicket).filter(
            FMSTicket.flow_id == f.id,
            FMSTicket.is_deleted == False,
            FMSTicket.status.notin_(_INACTIVE_STATUSES),
        ).count()
        flow_info.append({"flow": f, "stage_count": len(active_stages),
                           "active_tickets": active_tickets})

    return templates.TemplateResponse(request, "fms/flow_list.html", _ctx(
        request, user, db,
        flow_info=flow_info,
    ))


@router.get("/flows/new", response_class=HTMLResponse)
def fms_flow_new(request: Request, user: User = Depends(require_admin_or_redirect),
                 db: Session = Depends(get_db)):
    # Flow creation is SA-only — redirect client admins
    return _redirect("/fms/flows?err=Flows+are+configured+by+your+OmniFlow+account+manager.+Contact+support+to+add+a+new+flow.")


@router.post("/flows/new")
def fms_flow_create(
    name: str = Form(...), description: str = Form(""),
    color: str = Form("#3b82f6"), is_active: bool = Form(True),
    stages_json: str = Form("[]"),
    user: User = Depends(require_admin), db: Session = Depends(get_db),
):
    """Flow creation is SA-only — block silently."""
    return _redirect("/fms/flows?err=Flow+creation+is+managed+by+your+OmniFlow+account+manager.")


@router.get("/flows/{flow_id}", response_class=HTMLResponse)
def fms_flow_edit(flow_id: str, request: Request,
                  user: User = Depends(require_admin_or_redirect),
                  db: Session = Depends(get_db)):
    flow = _get_flow(db, flow_id, user.tenant_id)
    employees = db.query(User).filter(
        User.tenant_id == user.tenant_id,
        User.is_deleted == False, User.is_active == True).all()
    deployed_submodules = _get_deployed_submodules(db, user.tenant_id)
    # Client admins get read-only view — only SA can edit flow definitions
    return templates.TemplateResponse(request, "fms/flow_edit.html", _ctx(
        request, user, db, flow=flow, employees=employees,
        mode="readonly",  # signals template to hide all edit controls
        deployed_submodules=deployed_submodules))


@router.post("/flows/{flow_id}")
def fms_flow_update(
    flow_id: str,
    name: str = Form(...), description: str = Form(""),
    color: str = Form("#3b82f6"), is_active: bool = Form(True),
    stages_json: str = Form("[]"),
    user: User = Depends(require_admin), db: Session = Depends(get_db),
):
    """Flow editing is SA-only — block silently."""
    return _redirect(f"/fms/flows/{flow_id}?err=Flow+definitions+are+managed+by+your+OmniFlow+account+manager.")


@router.post("/flows/{flow_id}/ticket-form")
async def fms_flow_save_ticket_form(
    flow_id: str,
    request: Request,
    user: User = Depends(require_admin),
    db: Session = Depends(get_db),
):
    """Save the custom ticket-creation form fields for a flow. Accessible to all admins."""
    from fastapi.responses import JSONResponse
    flow = _get_flow(db, flow_id, user.tenant_id)
    body = await request.json()
    fields = body.get("fields", [])

    # Validate and normalise each field definition
    valid_types = {"text", "number", "date", "longtext", "select", "ref_list", "__priority__", "__due_date__"}
    clean = []
    for f in fields:
        ftype = (f.get("field_type") or "text").strip().lower()
        label = (f.get("label") or "").strip()
        if not label or ftype not in valid_types:
            continue
        builtin_types = {"__priority__", "__due_date__"}
        field_id = ftype if ftype in builtin_types else (f.get("id") or new_id())
        entry = {
            "id": field_id,
            "label": label,
            "field_type": ftype,
            "required": bool(f.get("required", False)),
            "order": int(f.get("order", len(clean))),
            "show_in_header": bool(f.get("show_in_header", False)),
        }
        if ftype == "select":
            raw_opts = f.get("options", [])
            entry["options"] = [o.strip() for o in raw_opts if str(o).strip()]
        elif ftype == "ref_list":
            entry["ref_list_id"]   = (f.get("ref_list_id") or "").strip()
            entry["ref_list_name"] = (f.get("ref_list_name") or "").strip()
        clean.append(entry)

    flow.ticket_form_fields_json = _json.dumps(clean)
    flow.updated_at = datetime.utcnow()
    db.commit()
    return JSONResponse({"ok": True, "field_count": len(clean)})


@router.post("/flows/{flow_id}/delete")
def fms_flow_delete(flow_id: str, user: User = Depends(require_admin),
                    db: Session = Depends(get_db)):
    """Flow deletion is SA-only — block client admins."""
    return _redirect("/fms/flows?err=Flow+deletion+is+managed+by+your+OmniFlow+account+manager.")


@router.post("/flows/deploy-library")
def fms_deploy_library(user: User = Depends(require_admin)):
    """Flow deployment is SA-only — block client admins."""
    return _redirect("/fms/flows?err=Flows+are+deployed+by+your+OmniFlow+account+manager.+Contact+support+to+request+a+new+flow.")


@router.post("/flows/import-csv")
async def fms_flow_import(
    file: UploadFile = File(...),
    user: User = Depends(require_admin), db: Session = Depends(get_db),
):
    """2-B-6: Bulk import flow definitions via CSV."""
    content = (await file.read()).decode("utf-8-sig", errors="replace").lstrip(chr(65279))
    reader  = csv.DictReader(io.StringIO(content))
    flows_created = 0
    for row in reader:
        fname = (row.get("flow_name") or "").strip()
        if not fname:
            continue
        flow = FMSFlow(
            tenant_id=user.tenant_id, name=fname,
            description=(row.get("description") or "").strip() or None,
            color=(row.get("color") or "#3b82f6").strip(),
            created_by_id=user.id,
        )
        db.add(flow)
        db.flush()
        # Parse comma-separated stage names
        stage_names = [(s.strip()) for s in (row.get("stages") or "").split("|") if s.strip()]
        for i, sname in enumerate(stage_names):
            is_terminal = (i == len(stage_names) - 1)
            db.add(FMSStage(
                flow_id=flow.id, tenant_id=user.tenant_id,
                name=sname, order=i, is_terminal=is_terminal,
            ))
        flows_created += 1
    db.commit()
    return _redirect(f"/fms/flows?imported={flows_created}")


# ── 2-C/D: Ticket Lifecycle ───────────────────────────────────────────────────

@router.get("/", response_class=HTMLResponse)
def fms_root(user: User = Depends(get_current_user_or_redirect)):
    return _redirect("/fms/dashboard")


def _submodule_cols(db: Session, ticket: FMSTicket, sub_tag: Optional[str]) -> dict:
    """Return a dict of extra sub-module columns for a ticket row (P7-04)."""
    if not sub_tag:
        return {}
    tid = ticket.id
    if sub_tag == "PMS":
        logs = db.query(PMSDailyLog).filter(
            PMSDailyLog.ticket_id == tid,
            PMSDailyLog.event_type == "DAILY_LOG",
        ).order_by(PMSDailyLog.log_date.desc()).all()
        cum_qty = sum(l.qty_done for l in logs)
        pct = 0
        if ticket.target_qty and ticket.target_qty > 0:
            pct = min(int(cum_qty / ticket.target_qty * 100), 100)
        has_blockers = any(l.has_blockers for l in logs)
        last_date = logs[0].log_date if logs else None
        return {
            "target_qty": ticket.target_qty,
            "cum_qty": cum_qty,
            "pct": pct,
            "blockers": has_blockers,
            "last_entry": last_date,
        }
    if sub_tag == "DISPATCH":
        recs = db.query(DispatchRecord).filter(
            DispatchRecord.ticket_id == tid).order_by(DispatchRecord.created_at.desc()).all()
        total_disp = sum(r.qty_dispatched for r in recs)
        remaining = (ticket.target_qty or 0) - total_disp
        pod_up = any(r.proof_photo_url for r in recs)
        last_date = recs[0].created_at if recs else None
        return {
            "total_dispatched": total_disp,
            "remaining": remaining,
            "last_dispatch": last_date,
            "pod_uploaded": pod_up,
        }
    if sub_tag == "INVOICE":
        recs = db.query(InvoiceRecord).filter(
            InvoiceRecord.ticket_id == tid,
            InvoiceRecord.is_deleted == False,
        ).order_by(InvoiceRecord.due_date).all()
        total_inv = sum(r.amount for r in recs)
        total_paid = sum(r.amount for r in recs if r.is_paid)
        outstanding = total_inv - total_paid
        oldest_due = min((r.due_date for r in recs if r.due_date and not r.is_paid), default=None)
        return {
            "total_invoiced": total_inv,
            "total_received": total_paid,
            "outstanding": outstanding,
            "oldest_due": oldest_due,
        }
    return {}


@router.get("/tickets/export")
def fms_tickets_export(
    request: Request,
    flow_id: Optional[str] = None,
    status_filter: Optional[str] = None,
    f_priority: List[str] = Query([]),
    f_assignee_id: List[str] = Query([]),
    date_from: Optional[str] = None,
    date_to: Optional[str] = None,
    user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    tenant = db.query(Tenant).get(user.tenant_id)
    if not has_feature(tenant, "CSV_EXPORT", db):
        return RedirectResponse("/plan?upgrade=CSV_EXPORT", status_code=302)
    tid = user.tenant_id
    q = db.query(FMSTicket).filter(FMSTicket.tenant_id == tid, FMSTicket.is_deleted == False)
    if user.role == "MANAGER":
        team_ids = [u.id for u in db.query(User).filter(
            User.manager_id == user.id, User.is_deleted == False).all()]
        team_ids.append(user.id)
        hist_tids = [h.ticket_id for h in db.query(FMSStageHistory).filter(
            FMSStageHistory.assignee_id.in_(team_ids)).all()]
        help_tids = [h.ticket_id for h in db.query(FMSTicketHelper).filter(
            FMSTicketHelper.user_id.in_(team_ids)).all()]
        all_ids = set(hist_tids) | set(help_tids)
        q = q.filter((FMSTicket.current_assignee_id.in_(team_ids)) | (FMSTicket.id.in_(all_ids)))
    elif user.role == "EMPLOYEE":
        help_tids = [h.ticket_id for h in db.query(FMSTicketHelper).filter(
            FMSTicketHelper.user_id == user.id).all()]
        hist_tids = [h.ticket_id for h in db.query(FMSStageHistory).filter(
            FMSStageHistory.assignee_id == user.id).all()]
        emp_ids = set(help_tids) | set(hist_tids)
        q = q.filter((FMSTicket.current_assignee_id == user.id) | (FMSTicket.id.in_(emp_ids)))
    if flow_id:
        q = q.filter(FMSTicket.flow_id == flow_id)
    if status_filter:
        q = q.filter(FMSTicket.status == status_filter)
    if f_priority:
        q = q.filter(FMSTicket.priority.in_(f_priority))
    if f_assignee_id:
        q = q.filter(FMSTicket.current_assignee_id.in_(f_assignee_id))
    if date_from:
        try:
            q = q.filter(FMSTicket.created_at >= datetime.fromisoformat(date_from))
        except Exception:
            pass
    if date_to:
        try:
            q = q.filter(FMSTicket.created_at <= datetime.fromisoformat(date_to).replace(hour=23, minute=59, second=59))
        except Exception:
            pass
    tickets = q.order_by(FMSTicket.created_at.desc()).all()

    def fmt_dt(dt):
        return dt.strftime("%d %b %Y") if dt else ""

    output = io.StringIO()
    w = csv.writer(output)
    w.writerow(["Ticket ID", "Title", "Flow", "Current Stage", "Status", "Priority",
                "Assigned To", "Created By", "Created Date", "Due Date", "Work Order No."])
    for t in tickets:
        assignee = t.current_assignee
        w.writerow([
            t.display_id or "", t.title,
            t.flow.name if t.flow else "",
            t.current_stage.name if t.current_stage else "",
            t.status, t.priority,
            assignee.name if assignee else "",
            t.created_by.name if t.created_by else "",
            fmt_dt(t.created_at), fmt_dt(t.due_at),
            t.wo_number or "",
        ])
    filename = f"fms_tickets_{datetime.utcnow().strftime('%Y-%m-%d')}.csv"
    return StreamingResponse(
        iter([output.getvalue()]),
        media_type="text/csv",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


@router.get("/flow-groups/{group_id}/members")
def fms_flow_group_members(
    group_id: str,
    user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """R4: second-level selector data — member flows only, no swimlane/ticket
    data resolved at this step."""
    group = db.query(FMSFlowGroup).filter(
        FMSFlowGroup.id == group_id,
        FMSFlowGroup.tenant_id == user.tenant_id,
        FMSFlowGroup.is_deleted == False,
    ).first()
    if not group:
        raise HTTPException(status_code=404, detail="Group not found")

    members = db.query(FMSFlow).filter(
        FMSFlow.group_id == group_id,
        FMSFlow.is_active == True,
        FMSFlow.is_deleted == False,
    ).order_by(FMSFlow.name).all()
    return JSONResponse({
        "group": {"id": group.id, "name": group.name},
        "members": [{"id": f.id, "name": f.name} for f in members],
    })


@router.get("/dashboard", response_class=HTMLResponse)
def fms_dashboard(
    request: Request,
    flow_id: Optional[str] = None,
    stage_id: Optional[str] = None,
    view: str = "stage_table",
    dept_id: List[str] = Query([]),
    manager_id: List[str] = Query([]),
    branch_id: List[str] = Query([]),
    month: Optional[str] = None,
    status_filter: Optional[str] = None,
    f_priority: List[str] = Query([]),
    f_assignee_id: List[str] = Query([]),
    date_from: Optional[str] = None,
    date_to: Optional[str] = None,
    my_work: int = 0,
    log_event_type: List[str] = Query([]),
    log_actor_id: Optional[str] = None,
    log_search: Optional[str] = None,
    user: User = Depends(get_current_user_or_redirect),
    db: Session = Depends(get_db),
):
    """FMS Dashboard — summary strip + flow cards + swimlane/stage-table/consolidated view."""
    # View name normalisation — map legacy names and accept new names
    _view_map = {"list": "stage", "stage_table": "stage", "consolidated": "table"}
    view = _view_map.get(view, view)
    if view not in ("table", "stage", "timeline", "swimlane", "log"):
        view = "stage"
    import logging as _log, traceback as _tb
    try:
        return _fms_dashboard_inner(
            request=request, flow_id=flow_id, stage_id=stage_id, view=view,
            dept_id=dept_id, manager_id=manager_id, branch_id=branch_id,
            month=month, status_filter=status_filter,
            f_priority=f_priority, f_assignee_id=f_assignee_id,
            date_from=date_from, date_to=date_to,
            my_work=my_work,
            log_event_type=log_event_type, log_actor_id=log_actor_id, log_search=log_search,
            user=user, db=db,
        )
    except Exception as _exc:
        _log.getLogger("fms.dashboard").error(
            "FMS DASHBOARD CRASH:\n%s", _tb.format_exc()
        )
        raise


def _fms_dashboard_inner(
    request, flow_id, stage_id, view, dept_id, manager_id, branch_id,
    month, status_filter, f_priority, f_assignee_id, date_from, date_to, user, db,
    my_work: int = 0,
    log_event_type: List[str] = [], log_actor_id: Optional[str] = None, log_search: Optional[str] = None,
):
    tid = user.tenant_id
    now = datetime.utcnow()
    tenant = db.query(Tenant).get(tid)

    # All active flows for this tenant
    all_flows = db.query(FMSFlow).filter(
        FMSFlow.tenant_id == tid, FMSFlow.is_active == True,
        FMSFlow.is_deleted == False,
    ).order_by(FMSFlow.created_at).all()

    # Role-based flow visibility: Admin sees all; Manager sees team-involved flows;
    # Employee sees only flows they were ever part of.
    if user.role == "EMPLOYEE":
        emp_ticket_flow_ids: set = set()
        for t in db.query(FMSTicket.flow_id).filter(
            FMSTicket.tenant_id == tid,
            FMSTicket.is_deleted == False,
        ).filter(
            (FMSTicket.current_assignee_id == user.id) |
            FMSTicket.id.in_(
                db.query(FMSStageHistory.ticket_id).filter(
                    FMSStageHistory.assignee_id == user.id)
            ) |
            FMSTicket.stage_assignees_json.like(f'%"{user.id}"%')
        ).distinct():
            emp_ticket_flow_ids.add(t.flow_id)
        # Also include flows where the employee is configured as a stage's
        # default assignee, even if no ticket has reached that stage yet.
        for s in db.query(FMSStage.flow_id).filter(
            FMSStage.tenant_id == tid,
            FMSStage.is_deleted == False,
            (FMSStage.default_assignee_id == user.id) |
            FMSStage.default_assignee_ids_json.like(f'%"{user.id}"%'),
        ).distinct():
            emp_ticket_flow_ids.add(s.flow_id)
        flows = [f for f in all_flows if f.id in emp_ticket_flow_ids]
    elif user.role == "MANAGER":
        mgr_team_ids = [u.id for u in db.query(User).filter(
            User.manager_id == user.id, User.is_deleted == False).all()]
        mgr_team_ids.append(user.id)
        mgr_flow_ids: set = set()
        # Flows created by the manager
        for f in db.query(FMSFlow.id).filter(FMSFlow.created_by_id == user.id):
            mgr_flow_ids.add(f.id)
        # Flows where team members are/were assigned to any ticket
        for t in db.query(FMSTicket.flow_id).filter(
            FMSTicket.tenant_id == tid,
            FMSTicket.is_deleted == False,
        ).filter(
            (FMSTicket.current_assignee_id.in_(mgr_team_ids)) |
            FMSTicket.id.in_(
                db.query(FMSStageHistory.ticket_id).filter(
                    FMSStageHistory.assignee_id.in_(mgr_team_ids))
            ) |
            FMSTicket.id.in_(
                db.query(FMSTicketHelper.ticket_id).filter(
                    FMSTicketHelper.user_id.in_(mgr_team_ids))
            )
        ).distinct():
            mgr_flow_ids.add(t.flow_id)
        # Flows where a team member is configured as a stage's default
        # assignee, even if no ticket has reached that stage yet.
        mgr_team_ids_set = set(mgr_team_ids)
        for s in db.query(FMSStage).filter(
            FMSStage.tenant_id == tid,
            FMSStage.is_deleted == False,
            (FMSStage.default_assignee_id.in_(mgr_team_ids)) |
            (FMSStage.default_assignee_ids_json != None),
        ):
            if s.default_assignee_id in mgr_team_ids_set or mgr_team_ids_set & set(_stage_default_assignee_ids(s)):
                mgr_flow_ids.add(s.flow_id)
        flows = [f for f in all_flows if f.id in mgr_flow_ids]
    else:
        flows = all_flows

    # Select active flow (tab)
    active_flow = None
    if flow_id:
        active_flow = next((f for f in flows if f.id == flow_id), None)
    if active_flow is None and flows:
        active_flow = flows[0]

    # ── Summary strip ─────────────────────────────────────────────────────────
    base_q = db.query(FMSTicket).filter(
        FMSTicket.tenant_id == tid, FMSTicket.is_deleted == False)

    # Role scoping — build team_ids once, reuse for all three queries
    team_ids = []
    mgr_all_fms_ids: set = set()
    emp_all_fms_ids: set = set()
    emp_upcoming_ids: set = set()
    if user.role == "MANAGER":
        team_ids = [u.id for u in db.query(User).filter(
            User.manager_id == user.id, User.is_deleted == False).all()]
        team_ids.append(user.id)
        # 'Ever worked on' — include historical stage assignees + helpers
        hist_tids = [h.ticket_id for h in db.query(FMSStageHistory).filter(
            FMSStageHistory.assignee_id.in_(team_ids)).distinct().all()]
        help_tids = [h.ticket_id for h in db.query(FMSTicketHelper).filter(
            FMSTicketHelper.user_id.in_(team_ids)).all()]
        mgr_all_fms_ids = set(hist_tids) | set(help_tids)
        base_q = base_q.filter(
            (FMSTicket.current_assignee_id.in_(team_ids)) |
            (FMSTicket.id.in_(mgr_all_fms_ids))
        )
    elif user.role == "EMPLOYEE":
        helper_tids = [h.ticket_id for h in db.query(FMSTicketHelper).filter(
            FMSTicketHelper.user_id == user.id).all()]
        hist_tids_emp = [h.ticket_id for h in db.query(FMSStageHistory).filter(
            FMSStageHistory.assignee_id == user.id).all()]
        # pre-assigned to a future stage on any active ticket
        upcoming_tids_emp = [t.id for t in db.query(FMSTicket).filter(
            FMSTicket.tenant_id == tid,
            FMSTicket.is_deleted == False,
            FMSTicket.stage_assignees_json.like(f'%"{user.id}"%'),
        ).all()]
        emp_all_fms_ids = set(helper_tids) | set(hist_tids_emp)
        emp_upcoming_ids = set(upcoming_tids_emp) - emp_all_fms_ids
        base_q = base_q.filter(
            (FMSTicket.current_assignee_id == user.id) |
            (FMSTicket.id.in_(emp_all_fms_ids)) |
            (FMSTicket.id.in_(emp_upcoming_ids))
        )

    # ── Apply filter-bar selections to KPI base query ────────────────────────
    if active_flow:
        base_q = base_q.filter(FMSTicket.flow_id == active_flow.id)
    if f_priority:
        base_q = base_q.filter(FMSTicket.priority.in_(f_priority))
    # Resolve assignee/dept/manager/branch filter (same logic as ticket list)
    _kpi_assignee_ids = None
    if f_assignee_id:
        _kpi_assignee_ids = list(f_assignee_id)
    elif dept_id or manager_id or branch_id:
        _fq = db.query(User).filter(
            User.tenant_id == tid, User.is_deleted == False, User.is_active == True)
        if dept_id:
            _fq = _fq.filter(User.department_id.in_(dept_id))
        if manager_id:
            _mgr_team = []
            for _mid in manager_id:
                _mgr_team += [u.id for u in db.query(User).filter(
                    User.manager_id == _mid, User.tenant_id == tid,
                    User.is_deleted == False).all()]
                _mgr_team.append(_mid)
            _fq = _fq.filter(User.id.in_(_mgr_team))
        if branch_id:
            _br_dept_ids = [d.id for d in db.query(Department).filter(
                Department.branch_id.in_(branch_id), Department.tenant_id == tid,
                Department.is_deleted == False).all()]
            _fq = _fq.filter(User.department_id.in_(_br_dept_ids))
        _kpi_assignee_ids = [u.id for u in _fq.all()]
    if _kpi_assignee_ids is not None:
        base_q = base_q.filter(FMSTicket.current_assignee_id.in_(_kpi_assignee_ids))

    active_tickets = base_q.filter(
        FMSTicket.status.notin_(_INACTIVE_STATUSES)).count()
    flagged_count  = base_q.filter(FMSTicket.is_flagged == True).count()
    awaiting_count = base_q.filter(FMSTicket.status == "ACTIVE").count()

    tat_breaches = 0
    open_tickets = base_q.filter(
        FMSTicket.status.notin_(_INACTIVE_STATUSES)).all()
    for t in open_tickets:
        # Phase 0: check every currently-open split, not just one — two splits
        # of the same ticket can breach TAT independently at different stages.
        for h in _open_histories_for_ticket(db, t.id):
            stage_for_h = h.stage
            # Prefer ticket-specific planned_end; fall back to stage target
            if h.planned_end:
                if now > h.planned_end:
                    tat_breaches += 1
            elif stage_for_h and stage_for_h.target_tat_hours:
                elapsed = (now - h.entered_at).total_seconds() / 3600
                if elapsed > stage_for_h.target_tat_hours:
                    tat_breaches += 1

    # Compliance: completed stage history rows with a planned window or stage target
    all_history = db.query(FMSStageHistory).join(
        FMSTicket, FMSStageHistory.ticket_id == FMSTicket.id
    ).filter(
        FMSTicket.tenant_id == tid,
        FMSStageHistory.exited_at != None,
    ).all()
    scoreable = [
        h for h in all_history
        if (h.planned_start and h.planned_end) or
           (h.stage and h.stage.target_tat_hours)
    ]
    if scoreable:
        def _on_time(h):
            actual_h = (h.exited_at - h.entered_at).total_seconds() / 3600
            if h.planned_start and h.planned_end:
                target_h = (h.planned_end - h.planned_start).total_seconds() / 3600
            else:
                target_h = h.stage.target_tat_hours
            return actual_h <= target_h
        on_time = sum(1 for h in scoreable if _on_time(h))
        compliance = int(on_time / len(scoreable) * 100)
    else:
        compliance = 0

    # ── Per-flow ticket counts for flow cards ─────────────────────────────────
    flow_counts = {}
    for f in flows:
        flow_counts[f.id] = db.query(FMSTicket).filter(
            FMSTicket.flow_id == f.id, FMSTicket.is_deleted == False,
            FMSTicket.status.notin_(_INACTIVE_STATUSES),
        ).count()

    # ── Flow dropdown: ungrouped flows shown individually, grouped flows
    # collapse into their group entry (R3) ────────────────────────────────────
    dropdown_ungrouped_flows = [f for f in flows if not f.group_id]
    _visible_group_ids = {f.group_id for f in flows if f.group_id}
    dropdown_flow_groups = []
    if _visible_group_ids:
        dropdown_flow_groups = db.query(FMSFlowGroup).filter(
            FMSFlowGroup.id.in_(_visible_group_ids),
            FMSFlowGroup.is_active == True,
            FMSFlowGroup.is_deleted == False,
        ).order_by(FMSFlowGroup.name).all()

    # ── Filter data (loaded for all views) ───────────────────────────────────
    _depts_raw = db.query(Department).filter(
        Department.tenant_id == tid, Department.is_deleted == False
    ).order_by(Department.name).all()
    _seen_d: set = set()
    departments = [d for d in _depts_raw if d.name not in _seen_d and not _seen_d.add(d.name)]

    _mgrs_raw = db.query(User).filter(
        User.tenant_id == tid, User.role.in_(["MANAGER", "ADMIN"]),
        User.is_deleted == False, User.is_active == True,
    ).order_by(User.name).all()
    _seen_m: set = set()
    managers = [m for m in _mgrs_raw if m.name not in _seen_m and not _seen_m.add(m.name)]

    _branches_raw = db.query(Branch).filter(
        Branch.tenant_id == tid, Branch.is_deleted == False
    ).order_by(Branch.name).all()
    _seen_b: set = set()
    branches = [b for b in _branches_raw if b.name not in _seen_b and not _seen_b.add(b.name)]

    # Resolve combined assignee filter from multi-value params (all are lists now)
    filter_assignee_ids = None  # None = no filter applied
    if f_assignee_id:
        filter_assignee_ids = list(f_assignee_id)
    elif dept_id or manager_id or branch_id:
        filt_q = db.query(User).filter(
            User.tenant_id == tid, User.is_deleted == False, User.is_active == True
        )
        if dept_id:
            filt_q = filt_q.filter(User.department_id.in_(dept_id))
        if manager_id:
            mgr_team = []
            for mid in manager_id:
                mgr_team += [u.id for u in db.query(User).filter(
                    User.manager_id == mid, User.tenant_id == tid,
                    User.is_deleted == False).all()]
                mgr_team.append(mid)
            filt_q = filt_q.filter(User.id.in_(mgr_team))
        if branch_id:
            br_dept_ids = [d.id for d in db.query(Department).filter(
                Department.branch_id.in_(branch_id), Department.tenant_id == tid,
                Department.is_deleted == False).all()]
            filt_q = filt_q.filter(User.department_id.in_(br_dept_ids))
        filter_assignee_ids = [u.id for u in filt_q.all()]

    # Priority filter list
    priority_filter = list(f_priority) if f_priority else []

    # Parse date range filters
    from datetime import datetime as _dtp
    filter_date_from = None
    filter_date_to = None
    if date_from:
        try:
            filter_date_from = _dtp.strptime(date_from, "%Y-%m-%d")
        except ValueError:
            pass
    if date_to:
        try:
            filter_date_to = _dtp.strptime(date_to, "%Y-%m-%d").replace(
                hour=23, minute=59, second=59)
        except ValueError:
            pass

    # ── Swimlane data ─────────────────────────────────────────────────────────
    tickets_by_stage: dict = {}
    tat_info: dict = {}
    if active_flow and view == "swimlane":  # swimlane kept for legacy access
        active_stages = [s for s in active_flow.stages if not s.is_deleted]
        for stage in active_stages:
            tickets_by_stage[stage.id] = []

        swimlane_q = db.query(FMSTicket).filter(
            FMSTicket.flow_id == active_flow.id,
            FMSTicket.is_deleted == False,
            FMSTicket.status.notin_(_INACTIVE_STATUSES),
        )
        if user.role == "MANAGER":
            swimlane_q = swimlane_q.filter(
                (FMSTicket.current_assignee_id.in_(team_ids)) |
                (FMSTicket.id.in_(mgr_all_fms_ids))
            )
        elif user.role == "EMPLOYEE":
            swimlane_q = swimlane_q.filter(
                (FMSTicket.current_assignee_id == user.id) |
                (FMSTicket.id.in_(emp_all_fms_ids)) |
                (FMSTicket.id.in_(emp_upcoming_ids))
            )
        if filter_assignee_ids is not None:
            swimlane_q = swimlane_q.filter(
                FMSTicket.current_assignee_id.in_(filter_assignee_ids))
        if priority_filter:
            swimlane_q = swimlane_q.filter(FMSTicket.priority.in_(priority_filter))
        if filter_date_from:
            swimlane_q = swimlane_q.filter(FMSTicket.created_at >= filter_date_from)
        if filter_date_to:
            swimlane_q = swimlane_q.filter(FMSTicket.created_at <= filter_date_to)

        for t in swimlane_q.all():
            sid = t.current_stage_id
            if sid in tickets_by_stage:
                tickets_by_stage[sid].append(t)
            h = _open_history(db, t.id)
            if h and t.current_stage and t.current_stage.target_tat_hours:
                pct = _tat_pct(h, t.current_stage)
                color = "green" if pct < 50 else "amber" if pct < 90 else "red"
            else:
                pct, color = None, "gray"
            tat_info[t.id] = {"pct": pct, "color": color,
                               "entered_at": h.entered_at if h else None}

    # ── List view data (legacy — kept for any remaining references) ───────────
    list_tickets = []
    if view == "_legacy_list":
        list_q = db.query(FMSTicket).filter(
            FMSTicket.tenant_id == tid, FMSTicket.is_deleted == False,
        )
        # Role scoping — 'ever worked on' for Manager
        if user.role == "MANAGER":
            list_q = list_q.filter(
                (FMSTicket.current_assignee_id.in_(team_ids)) |
                (FMSTicket.id.in_(mgr_all_fms_ids))
            )
        elif user.role == "EMPLOYEE":
            list_q = list_q.filter(
                (FMSTicket.current_assignee_id == user.id) |
                (FMSTicket.id.in_(emp_all_fms_ids)) |
                (FMSTicket.id.in_(emp_upcoming_ids))
            )
        # Flow filter
        if flow_id and active_flow:
            list_q = list_q.filter(FMSTicket.flow_id == active_flow.id)
        # Status filter
        if status_filter == "open":
            list_q = list_q.filter(FMSTicket.status.notin_(_INACTIVE_STATUSES))
        elif status_filter == "closed":
            list_q = list_q.filter(FMSTicket.status.in_(_INACTIVE_STATUSES))
        elif status_filter:
            list_q = list_q.filter(FMSTicket.status == status_filter.upper())
        # Assignee/dept/manager/branch filter (shared computation above)
        if filter_assignee_ids is not None:
            list_q = list_q.filter(
                FMSTicket.current_assignee_id.in_(filter_assignee_ids))
        # Priority filter
        if priority_filter:
            list_q = list_q.filter(FMSTicket.priority.in_(priority_filter))
        # Date range filter (takes priority over month)
        if filter_date_from or filter_date_to:
            if filter_date_from:
                list_q = list_q.filter(FMSTicket.created_at >= filter_date_from)
            if filter_date_to:
                list_q = list_q.filter(FMSTicket.created_at <= filter_date_to)
        elif month:
            try:
                y, m = int(month[:4]), int(month[5:7])
                month_start = datetime(y, m, 1)
                month_end = datetime(y + (m // 12), (m % 12) + 1, 1)
                list_q = list_q.filter(
                    FMSTicket.created_at >= month_start,
                    FMSTicket.created_at < month_end,
                )
            except (ValueError, IndexError):
                pass

        list_q = list_q.order_by(FMSTicket.created_at.desc())
        raw_tickets = list_q.limit(200).all()

        for t in raw_tickets:
            h = _open_history(db, t.id)
            days_open = (now - t.created_at).days if t.created_at else 0
            if h and t.current_stage and t.current_stage.target_tat_hours:
                pct = _tat_pct(h, t.current_stage)
                tat_color = "green" if pct < 50 else "amber" if pct < 90 else "red"
            else:
                pct, tat_color = None, "gray"
            # Determine previous stage via history (most recent exited row)
            prev_h = db.query(FMSStageHistory).filter(
                FMSStageHistory.ticket_id == t.id,
                FMSStageHistory.exited_at != None,
            ).order_by(FMSStageHistory.exited_at.desc()).first()
            prev_stage = prev_h.stage if prev_h else None
            list_tickets.append({
                "ticket": t,
                "days_open": days_open,
                "stage_name": t.current_stage.name if t.current_stage else "—",
                "flow_name": t.flow.name if t.flow else "—",
                "flow_color": t.flow.color if t.flow else "#64748b",
                "assignee_name": t.current_assignee.name if t.current_assignee else "—",
                "tat_pct": pct,
                "tat_color": tat_color,
                "stage_entered_at": h.entered_at if h else None,
                "prev_stage_id": prev_stage.id if prev_stage else None,
                "prev_stage_name": prev_stage.name if prev_stage else None,
            })

    # Flagged tickets for escalations panel (admin/manager)
    flagged_tickets = []
    if user.role in ("ADMIN", "MANAGER"):
        flagged_tickets = base_q.filter(
            FMSTicket.is_flagged == True,
            FMSTicket.status.notin_(_INACTIVE_STATUSES),
        ).limit(10).all()

    can_drag = user.role in ("ADMIN", "MANAGER")

    employees = db.query(User).filter(
        User.tenant_id == tid, User.is_deleted == False, User.is_active == True,
    ).order_by(User.name).all()

    # ── Log view — consolidated audit trail (all events + field edits) across
    # every ticket the user can see, filterable by event type / user / date /
    # ticket search / flow. Reuses base_q, which is already role- and
    # filter-bar-scoped, and (when a flow tab is selected) flow-scoped too —
    # so log_columns below reflects that one flow's actual custom columns.
    log_rows = []
    log_event_types = []
    log_columns = []
    if view == "log":
        log_event_types = sorted({
            r[0] for r in db.query(FMSEvent.event_type).distinct().all()
        } | {"FIELD_EDITED", "FIELD_RECALCULATED"})

        # Dynamic columns — every custom-field label defined on the selected
        # flow (ticket-creation fields + every stage's own columns), so the
        # log table's columns always match what that specific flow captures.
        if active_flow:
            seen_labels = set()
            if active_flow.ticket_form_fields_json:
                try:
                    for fd in _json.loads(active_flow.ticket_form_fields_json):
                        lbl = fd.get("label")
                        if lbl and lbl not in seen_labels and fd.get("field_type") not in ("__priority__", "__due_date__"):
                            seen_labels.add(lbl); log_columns.append(lbl)
                except Exception:
                    pass
            for s in sorted([s for s in active_flow.stages if not s.is_deleted], key=lambda s: s.order):
                if not s.custom_fields_json:
                    continue
                try:
                    for fd in _json.loads(s.custom_fields_json):
                        lbl = fd.get("label")
                        if lbl and lbl not in seen_labels:
                            seen_labels.add(lbl); log_columns.append(lbl)
                except Exception:
                    pass

        log_tids = [row[0] for row in base_q.with_entities(FMSTicket.id).all()]
        if log_tids:
            tickets_by_id = {t.id: t for t in db.query(FMSTicket).filter(FMSTicket.id.in_(log_tids)).all()}

            ev_q = db.query(FMSEvent).filter(FMSEvent.ticket_id.in_(log_tids))
            edit_q = db.query(FMSFieldEditLog).filter(
                FMSFieldEditLog.tenant_id == tid, FMSFieldEditLog.ticket_id.in_(log_tids))
            if log_actor_id:
                ev_q = ev_q.filter(FMSEvent.actor_id == log_actor_id)
                edit_q = edit_q.filter(FMSFieldEditLog.edited_by_id == log_actor_id)
            if date_from:
                try:
                    _df = datetime.fromisoformat(date_from)
                    ev_q = ev_q.filter(FMSEvent.created_at >= _df)
                    edit_q = edit_q.filter(FMSFieldEditLog.edited_at >= _df)
                except ValueError:
                    pass
            if date_to:
                try:
                    _dtt = datetime.fromisoformat(date_to) + timedelta(days=1)
                    ev_q = ev_q.filter(FMSEvent.created_at < _dtt)
                    edit_q = edit_q.filter(FMSFieldEditLog.edited_at < _dtt)
                except ValueError:
                    pass

            _search = (log_search or "").strip().lower()

            def _matches_search(t):
                if not _search:
                    return True
                return _search in (t.display_id or "").lower() or _search in (t.title or "").lower()

            for e in ev_q.order_by(FMSEvent.created_at.desc()).limit(500).all():
                if log_event_type and e.event_type not in log_event_type:
                    continue
                t = tickets_by_id.get(e.ticket_id)
                if not t or not _matches_search(t):
                    continue
                text, meta = _split_log_meta(e.detail)
                row_cols = {lbl: v for lbl, v in (meta.get("custom_fields") or {}).items() if lbl in log_columns}
                if meta.get("from_stage_name") and meta.get("stage_name"):
                    stage_display = f"{meta['from_stage_name']} → {meta['stage_name']}"
                else:
                    stage_display = meta.get("stage_name")
                log_rows.append({
                    "at": e.created_at, "ticket": t, "event_type": e.event_type,
                    "detail": text, "actor_name": e.actor.name if e.actor else "System",
                    "stage_name": stage_display,
                    "assignee_name": meta.get("assignee_name"),
                    "reason_note": meta.get("reason") or meta.get("note"),
                    "tat_window": meta.get("tat_window"),
                    "qty": meta.get("qty"),
                    "cols": row_cols,
                })

            if not log_event_type or "FIELD_EDITED" in log_event_type or "FIELD_RECALCULATED" in log_event_type:
                for ed in edit_q.order_by(FMSFieldEditLog.edited_at.desc()).limit(500).all():
                    ev_type = "FIELD_RECALCULATED" if ed.is_cascade else "FIELD_EDITED"
                    if log_event_type and ev_type not in log_event_type:
                        continue
                    t = tickets_by_id.get(ed.ticket_id)
                    if not t or not _matches_search(t):
                        continue
                    row_cols = {}
                    if ed.field_label in log_columns:
                        row_cols[ed.field_label] = f"{ed.old_value or '—'} → {ed.new_value or '—'}"
                    log_rows.append({
                        "at": ed.edited_at, "ticket": t, "event_type": ev_type,
                        "detail": f"{ed.field_label or ed.field_id}: {ed.old_value or '—'} → {ed.new_value or '—'}",
                        "actor_name": ed.edited_by.name if ed.edited_by else "System",
                        "stage_name": ed.stage.name if ed.stage else None,
                        "assignee_name": None,
                        "reason_note": ed.reason,
                        "tat_window": None,
                        "qty": None,
                        "cols": row_cols,
                    })

            log_rows.sort(key=lambda r: r["at"], reverse=True)
            log_rows = log_rows[:400]

    # ── P7-03/04: Stage-table view ────────────────────────────────────────────
    stage_table_stages = []
    active_stage = None
    stage_tickets = []
    stage_ticket_counts: dict = {}
    my_work_stage_ids: set = set()

    if active_flow:
        stage_table_stages = sorted(
            [s for s in active_flow.stages if not s.is_deleted], key=lambda s: s.order
        )
        # Per-stage counts for badges — Phase 0: count active splits at that
        # stage, not tickets, so a 2-split ticket contributes to both buckets.
        # Role-scoped the same way as the stage-tab ticket list below (a
        # MANAGER/EMPLOYEE previously saw a raw tenant-wide count here that
        # the list underneath could never actually match, e.g. a badge of
        # "15" next to a stage where their own visible list was empty).
        for s in stage_table_stages:
            _badge_rows = db.query(
                FMSTicketSplit.ticket_id, FMSTicketSplit.current_assignee_id
            ).filter(
                FMSTicketSplit.current_stage_id == s.id,
                FMSTicketSplit.is_deleted == False,
                FMSTicketSplit.status.notin_(_INACTIVE_STATUSES),
            ).all()
            if user.role == "MANAGER":
                _badge_tids = {r[0] for r in _badge_rows}
                _badge_cached = {
                    row[0] for row in db.query(FMSTicket.id).filter(
                        FMSTicket.id.in_(_badge_tids),
                        FMSTicket.current_assignee_id.in_(team_ids),
                    )
                } if _badge_tids else set()
                _badge_visible = (
                    {r[0] for r in _badge_rows if r[1] in team_ids}
                    | _badge_cached
                    | (mgr_all_fms_ids & _badge_tids)
                )
                stage_ticket_counts[s.id] = len(_badge_visible)
            elif user.role == "EMPLOYEE":
                _badge_tids = {r[0] for r in _badge_rows}
                _badge_cached = {
                    row[0] for row in db.query(FMSTicket.id).filter(
                        FMSTicket.id.in_(_badge_tids),
                        FMSTicket.current_assignee_id == user.id,
                    )
                } if _badge_tids else set()
                _badge_visible = (
                    {r[0] for r in _badge_rows if r[1] == user.id}
                    | _badge_cached
                    | (emp_all_fms_ids & _badge_tids)
                )
                stage_ticket_counts[s.id] = len(_badge_visible)
            else:
                stage_ticket_counts[s.id] = len({r[0] for r in _badge_rows})

        # "My Work" is scoped to whatever single stage tab happens to be
        # active — with 10 tickets spread across several stages, the tab
        # selected by default (or left over from browsing) is very often NOT
        # one the employee is assigned at, so the toggle silently showed 0
        # results even though matching tickets existed on other tabs. Find
        # every stage where this user currently owns an active split so we
        # can jump the tab there instead of failing quietly.
        my_work_stage_ids: set = set()
        if my_work:
            my_work_stage_ids = {
                row[0] for row in db.query(FMSTicketSplit.current_stage_id)
                .join(FMSTicket, FMSTicket.id == FMSTicketSplit.ticket_id)
                .filter(
                    FMSTicket.flow_id == active_flow.id,
                    FMSTicket.is_deleted == False,
                    FMSTicketSplit.current_assignee_id == user.id,
                    FMSTicketSplit.is_deleted == False,
                    FMSTicketSplit.status.notin_(_INACTIVE_STATUSES),
                ).distinct().all() if row[0]
            }

        # Determine active stage (default: first)
        if stage_id:
            active_stage = next((s for s in stage_table_stages if s.id == stage_id), None)
        if my_work and my_work_stage_ids and (active_stage is None or active_stage.id not in my_work_stage_ids):
            # Explicit stage_id wasn't one of the user's own stages (or none
            # was given at all) — jump to the first stage (by flow order)
            # where they actually have assigned work.
            active_stage = next((s for s in stage_table_stages if s.id in my_work_stage_ids), active_stage)
        if active_stage is None and stage_table_stages:
            active_stage = stage_table_stages[0]

        if active_stage and view in ("stage", "stage_table"):
            # Phase 0: a ticket appears under a stage tab if it has an active
            # split parked there — single-split tickets: identical result set
            # to before (their one split mirrors current_stage_id). Multi-split
            # tickets now correctly show up under every stage they have a live
            # split in.
            _stage_splits_here = db.query(
                FMSTicketSplit.ticket_id, FMSTicketSplit.current_assignee_id
            ).filter(
                FMSTicketSplit.current_stage_id == active_stage.id,
                FMSTicketSplit.is_deleted == False,
                FMSTicketSplit.status.notin_(_INACTIVE_STATUSES),
            ).all()
            _stage_split_ticket_ids = list({row[0] for row in _stage_splits_here})
            # Ticket -> set of assignee ids among the splits actually parked at
            # THIS stage, so role/assignee scoping can check who owns the work
            # sitting here, not just the ticket's ticket-wide cached assignee
            # (which mirrors whichever split is furthest along for multi-split
            # tickets and can misrepresent ownership at earlier stages).
            _stage_split_assignees_by_ticket: dict = {}
            for _tid_row, _aid_row in _stage_splits_here:
                _stage_split_assignees_by_ticket.setdefault(_tid_row, set()).add(_aid_row)
            q = db.query(FMSTicket).filter(
                FMSTicket.id.in_(_stage_split_ticket_ids),
                FMSTicket.is_deleted == False,
                FMSTicket.status.notin_(_INACTIVE_STATUSES),
            ) if _stage_split_ticket_ids else db.query(FMSTicket).filter(FMSTicket.id == None)
            if user.role == "MANAGER":
                _stage_team_ticket_ids = {
                    t_id for t_id, aids in _stage_split_assignees_by_ticket.items()
                    if aids & set(team_ids)
                }
                q = q.filter(
                    (FMSTicket.id.in_(_stage_team_ticket_ids)) |
                    (FMSTicket.current_assignee_id.in_(team_ids)) |
                    (FMSTicket.id.in_(mgr_all_fms_ids))
                )
            elif user.role == "EMPLOYEE":
                _stage_emp_ticket_ids = {
                    t_id for t_id, aids in _stage_split_assignees_by_ticket.items()
                    if user.id in aids
                }
                q = q.filter(
                    (FMSTicket.id.in_(_stage_emp_ticket_ids)) |
                    (FMSTicket.current_assignee_id == user.id) |
                    (FMSTicket.id.in_(emp_all_fms_ids))
                )
            if my_work:
                # Filter on the split actually parked at this stage, not the
                # ticket-wide cached assignee (which mirrors the furthest-along
                # split for multi-split tickets and can misrepresent who owns
                # the work sitting at active_stage).
                _my_work_ticket_ids = [
                    row[0] for row in db.query(FMSTicketSplit.ticket_id).filter(
                        FMSTicketSplit.current_stage_id == active_stage.id,
                        FMSTicketSplit.current_assignee_id == user.id,
                        FMSTicketSplit.is_deleted == False,
                        FMSTicketSplit.status.notin_(_INACTIVE_STATUSES),
                    ).distinct()
                ]
                q = q.filter(FMSTicket.id.in_(_my_work_ticket_ids))
            if priority_filter:
                q = q.filter(FMSTicket.priority.in_(priority_filter))
            if filter_assignee_ids is not None:
                _stage_filter_ticket_ids = {
                    t_id for t_id, aids in _stage_split_assignees_by_ticket.items()
                    if aids & set(filter_assignee_ids)
                }
                q = q.filter(
                    (FMSTicket.id.in_(_stage_filter_ticket_ids)) |
                    (FMSTicket.current_assignee_id.in_(filter_assignee_ids))
                )
            if filter_date_from:
                q = q.filter(FMSTicket.created_at >= filter_date_from)
            if filter_date_to:
                q = q.filter(FMSTicket.created_at <= filter_date_to)

            raw = q.order_by(FMSTicket.created_at.desc()).all()
            for t in raw:
                # Phase 0: this row represents whichever of the ticket's active
                # splits is parked at this stage (normally exactly one — the
                # rare case of two splits landing on the same stage just picks
                # the first for the row's summary columns; both are still
                # independently actionable from the Splits modal).
                all_active_splits = _active_splits(db, t.id)
                splits_here = [
                    s for s in all_active_splits
                    if s.current_stage_id == active_stage.id and s.status not in ("COMPLETED", "CLOSED")
                ]
                row_split = splits_here[0] if splits_here else _ensure_ticket_has_split(db, t)
                # brief §7: the Splits popup/button on a row should reflect only
                # the split "family" that originated from THIS row's split at
                # whatever stage it split at — not every active split the
                # ticket has anywhere. A split that itself gets split again at
                # a later stage (e.g. S2 -> S2-1/S2-2 at Stage 4) surfaces as
                # its own separate group on that later stage's row, distinct
                # from the S1/S2 group shown back at the stage where S1 split.
                split_family = [row_split] + [
                    s for s in all_active_splits if s.parent_split_id == row_split.id
                ]
                split_count = len(split_family)

                h = _open_history(db, t.id, split_id=row_split.id)
                # All evidence uploaded anywhere along this split's lineage —
                # not just this exact split id — so evidence attached before
                # an auto-split carved off the current split entity is still
                # reachable from later stages, instead of disappearing once
                # the ticket moves past the stage it was uploaded on.
                _lineage_ids = _split_lineage_ids(db, t.id, row_split.id)
                _lineage_stage_evidence = (
                    db.query(FMSStageHistory)
                    .filter(
                        FMSStageHistory.ticket_id == t.id,
                        FMSStageHistory.split_id.in_(_lineage_ids),
                        FMSStageHistory.evidence_url.isnot(None),
                    )
                    .order_by(FMSStageHistory.entered_at.desc())
                    .all()
                ) if _lineage_ids else []
                _lineage_split_evidence = (
                    db.query(FMSSplitEvidence)
                    .filter(FMSSplitEvidence.split_id.in_(_lineage_ids))
                    .order_by(FMSSplitEvidence.created_at.desc())
                    .all()
                ) if _lineage_ids else []
                evidence_payload = [
                    {
                        "url": eh.evidence_url, "filename": eh.evidence_filename,
                        "stage_name": eh.stage_name, "uploaded_at": eh.entered_at.strftime("%d %b %Y, %H:%M") if eh.entered_at else None,
                    }
                    for eh in _lineage_stage_evidence
                ] + [
                    {
                        "url": ev.file_url, "filename": ev.file_name,
                        "stage_name": None, "uploaded_at": ev.created_at.strftime("%d %b %Y, %H:%M") if ev.created_at else None,
                    }
                    for ev in _lineage_split_evidence
                ]
                latest_evidence = _lineage_stage_evidence[0] if _lineage_stage_evidence else None
                if h and active_stage.target_tat_hours:
                    pct = _tat_pct(h, active_stage)
                    tc = "green" if pct < 50 else "amber" if pct < 90 else "red"
                else:
                    pct, tc = None, "gray"
                sub_cols = _submodule_cols(db, t, active_stage.sub_module_tag)
                # Aggregate custom field values from ALL stage history entries.
                # Keys are UUID-based (per field def), so we also add label-keyed
                # entries using the stage's field definitions — this allows cross-stage
                # column display when two stages share the same label name.
                # Build a label-keyed lookup of all custom field values across every
                # stage this ticket has visited, using a direct DB query (avoids
                # lazy-load uncertainty). Also index by UUID so reused-column refs work.
                import json as _json
                cf_all: dict = {}
                if t.ticket_custom_fields_json:
                    try:
                        cf_all.update(_json.loads(t.ticket_custom_fields_json))
                    except Exception:
                        pass
                # Scoped to this row's own split's full lineage — otherwise a
                # sibling split's custom-field values (e.g. Quantity/Issued
                # Qty at a different stage) can clobber this row's via shared
                # field ids/labels. Lineage (not just the split's own id) so
                # an auto-split-created split can still see what its parent
                # captured before this split existed as its own entity —
                # e.g. Quantity/Price entered at Sales, needed for a formula
                # at Profit once a piece of the ticket splits forward.
                cf_all.update(_cross_stage_cf(
                    db, t.id, stage_table_stages,
                    split_id=_split_lineage_ids(db, t.id, row_split.id) if row_split else None,
                ))
                _live_eval_formulas(cf_all, stage_table_stages)
                # The split "actual" field is entered as each visit's own
                # increment (brief §5), so the raw value picked up by
                # _cross_stage_cf above is just the LAST delta typed — fine
                # for the auto-split engine's own bookkeeping, but wrong
                # wherever this field is displayed as a normal column/value:
                # a user reading "Actual Quantity" expects the running total
                # received so far, not whatever number happened to be typed
                # in the most recent visit. Override with the split's tracked
                # cumulative wherever this row's own split has one.
                if (active_stage.split_enabled and active_stage.split_actual_field
                        and row_split and row_split.last_cumulative_entered is not None):
                    cf_all[active_stage.split_actual_field] = row_split.last_cumulative_entered
                planned_end = None
                pd = _planned_dates(t, stage_table_stages, tenant)
                if active_stage.id in pd:
                    planned_end = pd[active_stage.id][1]
                row_assignee = row_split.current_assignee if row_split else t.current_assignee
                # R8: splits popup is a read-only table of ticket-creation columns
                # + value entered per split — evidence-indicator lookup is a single
                # cheap query per ticket row (typically 1-3 splits).
                _split_ids_for_evidence = [s.id for s in split_family]
                _evidence_split_ids = set()
                if _split_ids_for_evidence:
                    _evidence_split_ids = {
                        row[0] for row in db.query(FMSSplitEvidence.split_id).filter(
                            FMSSplitEvidence.split_id.in_(_split_ids_for_evidence)
                        ).distinct().all()
                    }
                # R8: ticket-creation columns — just Priority as a base field
                # (Title/WO#/Target Qty/Qty Unit dropped per explicit request:
                # not useful in this popup, and Target Qty already has its
                # own dedicated split-detail column further right) plus any
                # additional per-flow custom fields configured on the
                # ticket-creation form (flow.ticket_form_fields_json, used by
                # bulk-create).
                ticket_form_columns = [
                    {"label": "Priority", "value": t.priority or "—"},
                ]
                _tff_defs = []
                try:
                    _tff_defs = _json.loads(t.flow.ticket_form_fields_json or "[]") if t.flow else []
                except Exception:
                    _tff_defs = []
                _tff_vals = {}
                try:
                    _tff_vals = _json.loads(t.ticket_custom_fields_json or "{}")
                except Exception:
                    _tff_vals = {}
                ticket_form_columns += [
                    {"label": fd.get("label", ""), "value": _tff_vals.get(fd.get("id", ""), "—")}
                    for fd in _tff_defs
                    if fd.get("field_type") not in ("__priority__", "__due_date__") and fd.get("label")
                ]
                # Admin-configurable "identifying" columns (Setup > Ticket
                # Creation Form > Show in header) — shown next to the ticket
                # number in Enter Data / Complete Stage popups instead of the
                # generic auto-generated title (e.g. "Ticket-1"), so the
                # person filling in data can tell which real order/item
                # they're working on.
                header_fields = [
                    {"label": fd.get("label", ""), "value": _tff_vals.get(fd.get("id", ""), "—")}
                    for fd in _tff_defs
                    if fd.get("show_in_header") and fd.get("label")
                ]

                splits_payload = [
                    {
                        "id": s.id, "label": s.split_label, "qty": s.qty,
                        "stage_id": s.current_stage_id,
                        "stage_order": s.current_stage.order if s.current_stage else -1,
                        "stage_name": s.current_stage.name if s.current_stage else "—",
                        "assignee_id": s.current_assignee_id or "",
                        "assignee_name": s.current_assignee.name if s.current_assignee else "—",
                        "status": s.status,
                        "updated_at": s.updated_at.strftime("%d %b %Y, %H:%M") if s.updated_at else None,
                        # Ticket-creation columns (R8) — driven by this flow's
                        # actual ticket-creation form, same on every split row
                        # since they belong to the parent ticket.
                        "ticket_display_id": t.display_id or t.id[:8],
                        "ticket_target_qty": t.target_qty,
                        "ticket_form_columns": ticket_form_columns,
                        "ticket_created_at": t.created_at.strftime("%d %b %Y, %H:%M") if t.created_at else None,
                        # Auto-split engine / split-detail fields
                        "split_display_id": s.split_display_id or s.split_label,
                        "entered_value": s.entered_value,
                        "target_value_at_split": s.target_value_at_split,
                        "is_remainder": bool(s.is_remainder),
                        "is_auto_split": bool(s.is_auto_split),
                        "created_at": s.created_at.strftime("%d %b %Y, %H:%M") if s.created_at else None,
                        "has_evidence": s.id in _evidence_split_ids,
                    }
                    for s in split_family
                ] if split_count > 1 else []
                stage_tickets.append({
                    "ticket": t,
                    "tat_pct": pct,
                    "tat_color": tc,
                    "assignee_name": row_assignee.name if row_assignee else "—",
                    "sub": sub_cols,
                    "entered_at": h.entered_at if h else None,
                    "cf_all": cf_all,
                    "planned_end": planned_end,
                    "split_id": row_split.id if row_split else None,
                    "split_label": row_split.split_label if row_split else None,
                    "split_last_cumulative": row_split.last_cumulative_entered if row_split else None,
                    "split_count": split_count,
                    "header_fields": header_fields,
                    "splits_payload": splits_payload,
                    "evidence_url": latest_evidence.evidence_url if latest_evidence else None,
                    "evidence_filename": latest_evidence.evidence_filename if latest_evidence else None,
                    "evidence_payload": evidence_payload,
                })

    # Next/prev stage maps (used by Mark Done and Move Backward modals)
    next_stage_map: dict = {}
    prev_stage_map: dict = {}
    for i, s in enumerate(stage_table_stages):
        if i + 1 < len(stage_table_stages):
            next_stage_map[s.id] = stage_table_stages[i + 1]
        if i > 0:
            prev_stage_map[s.id] = stage_table_stages[i - 1]

    # ── Table view: full journey per ticket — all stages as columns ───────────
    table_tickets = []
    if view == "table" and active_flow and stage_table_stages:
        tq = db.query(FMSTicket).filter(
            FMSTicket.flow_id == active_flow.id,
            FMSTicket.is_deleted == False,
        )
        if user.role == "MANAGER":
            tq = tq.filter(
                (FMSTicket.current_assignee_id.in_(team_ids)) |
                (FMSTicket.id.in_(mgr_all_fms_ids))
            )
        elif user.role == "EMPLOYEE":
            tq = tq.filter(
                (FMSTicket.current_assignee_id == user.id) |
                (FMSTicket.id.in_(emp_all_fms_ids)) |
                (FMSTicket.id.in_(emp_upcoming_ids))
            )
        if priority_filter:
            tq = tq.filter(FMSTicket.priority.in_(priority_filter))
        if filter_assignee_ids is not None:
            tq = tq.filter(FMSTicket.current_assignee_id.in_(filter_assignee_ids))
        if filter_date_from:
            tq = tq.filter(FMSTicket.created_at >= filter_date_from)
        if filter_date_to:
            tq = tq.filter(FMSTicket.created_at <= filter_date_to)

        import json as _json
        for t in tq.order_by(FMSTicket.created_at.desc()).all():
            pd = _planned_dates(t, stage_table_stages, tenant)

            # Build latest-visit dict from history: stage_id → most recent row
            all_hist = db.query(FMSStageHistory).filter(
                FMSStageHistory.ticket_id == t.id
            ).order_by(FMSStageHistory.entered_at).all()
            visit_map: dict = {}
            for h in all_hist:
                visit_map[h.stage_id] = h  # last assignment wins (most recent visit)

            # Manual-edit audit trail, keyed by (stage_id or "" for ticket-level
            # fields) -> field_id -> latest FMSFieldEditLog row, so cell
            # tooltips ("who edited, when, why") persist across page loads.
            edit_info_by_stage: dict = {}
            for el in (
                db.query(FMSFieldEditLog)
                .filter(FMSFieldEditLog.ticket_id == t.id)
                .order_by(FMSFieldEditLog.edited_at)
                .all()
            ):
                edit_info_by_stage.setdefault(el.stage_id or "", {})[el.field_id] = el

            # Base ticket-level custom fields (from creation form) — shared
            # across every stage column since they belong to the ticket, not
            # any one split.
            cf_base: dict = {}
            if t.ticket_custom_fields_json:
                try:
                    cf_base.update(_json.loads(t.ticket_custom_fields_json))
                except Exception:
                    pass
            _live_eval_formulas(cf_base, stage_table_stages)

            stages_info = []
            for s in stage_table_stages:
                ps, pe = pd.get(s.id, (None, None))
                h = visit_map.get(s.id)
                actual_start = h.entered_at if h else None
                actual_end   = h.exited_at  if h else None
                delay_h      = None
                delay_positive = None
                if actual_end and pe:
                    delay_secs = (actual_end - pe).total_seconds()
                    delay_h    = round(abs(delay_secs) / 3600, 1)
                    delay_positive = delay_secs > 0
                is_current = (s.id == t.current_stage_id)
                assignee_name = h.assignee.name if (h and h.assignee) else "—"
                # Scoped to the split that actually visited this stage —
                # otherwise a sibling split's custom-field values (e.g.
                # Quantity/Issued Qty entered at a different stage) can
                # clobber this stage's own value via shared field ids/labels.
                cf_for_stage = dict(cf_base)
                cf_for_stage.update(_cross_stage_cf(
                    db, t.id, stage_table_stages,
                    split_id=_split_lineage_ids(db, t.id, h.split_id) if (h and h.split_id) else None,
                ))
                _live_eval_formulas(cf_for_stage, stage_table_stages)
                stages_info.append({
                    "stage":          s,
                    "planned_start":  ps,
                    "planned_end":    pe,
                    "actual_start":   actual_start,
                    "actual_end":     actual_end,
                    "delay_h":        delay_h,
                    "delay_positive": delay_positive,
                    "is_current":     is_current,
                    "visited":        h is not None,
                    "cf":             cf_for_stage,
                    "assignee_name":  assignee_name,
                    "edit_info":      edit_info_by_stage.get(s.id, {}),
                })

            _t_active_splits = _active_splits(db, t.id)

            # Simplified 3-way status for the Table view's quick filter:
            # CLOSED (terminal), OVERDUE (still open but past its due date or
            # its current stage's planned end), else ACTIVE.
            if t.status in ("COMPLETED", "CLOSED"):
                display_status = "CLOSED"
            else:
                cur_planned_end = pd.get(t.current_stage_id, (None, None))[1] if t.current_stage_id else None
                is_overdue = (t.due_at and t.due_at < now) or (cur_planned_end and cur_planned_end < now)
                display_status = "OVERDUE" if is_overdue else "ACTIVE"

            table_tickets.append({
                "ticket":        t,
                "assignee_name": t.current_assignee.name if t.current_assignee else "—",
                "stages":        stages_info,
                "edit_info":     edit_info_by_stage.get("", {}),
                "display_status": display_status,
                # Phase 0 §6: stage-distribution badge for multi-split tickets
                "split_count":   len(_t_active_splits),
                "split_stage_names": [s.current_stage.name for s in _t_active_splits
                                       if s.current_stage and s.status not in ("COMPLETED", "CLOSED")],
            })

        if status_filter and status_filter.upper() in ("ACTIVE", "OVERDUE", "CLOSED"):
            table_tickets = [r for r in table_tickets if r["display_status"] == status_filter.upper()]

    # ── Timeline view ─────────────────────────────────────────────────────────
    # Phase 0 fix: a ticket's progress is tracked per-SPLIT, not on the cached
    # FMSTicket.current_stage_id (which only reflects the furthest-along
    # split once a ticket has more than one — see _sync_ticket_cache). Iterate
    # active splits at this stage, same pattern as the Stage view above, so a
    # split ticket shows up at EVERY stage one of its live splits occupies
    # (e.g. remainder still at stage 2 AND moved-forward part at stage 3/4).
    timeline_data = []
    if view == "timeline" and active_flow and stage_table_stages:
        for s in stage_table_stages:
            splitq = db.query(FMSTicketSplit).join(
                FMSTicket, FMSTicketSplit.ticket_id == FMSTicket.id
            ).filter(
                FMSTicketSplit.current_stage_id == s.id,
                FMSTicketSplit.is_deleted == False,
                FMSTicketSplit.status.notin_(_INACTIVE_STATUSES),
                FMSTicket.flow_id == active_flow.id,
                FMSTicket.is_deleted == False,
            )
            if user.role == "MANAGER":
                splitq = splitq.filter(
                    (FMSTicketSplit.current_assignee_id.in_(team_ids)) |
                    (FMSTicketSplit.ticket_id.in_(mgr_all_fms_ids))
                )
            elif user.role == "EMPLOYEE":
                splitq = splitq.filter(
                    (FMSTicketSplit.current_assignee_id == user.id) |
                    (FMSTicketSplit.ticket_id.in_(emp_all_fms_ids))
                )
            split_list = splitq.all()
            stage_items = []
            for sp in split_list:
                t = sp.ticket
                h = _open_history(db, t.id, split_id=sp.id)
                time_at_s = int((now - h.entered_at).total_seconds()) if h else 0
                total_active = len(_active_splits(db, t.id))
                stage_items.append({
                    "ticket": t,
                    "time_at_s": time_at_s,
                    "assignee_name": sp.current_assignee.name if sp.current_assignee else "—",
                    "split_id": sp.id,
                    "split_label": sp.split_display_id or sp.split_label,
                    "split_count": total_active,
                })
            timeline_data.append({
                "stage": s,
                "count": len(stage_items),
                "tickets": stage_items,
            })

    # ── Consolidated table view (legacy — maps to 'table' now) ───────────────
    import json as _json
    consolidated_rows = []
    if view == "_legacy_consolidated" and active_flow and stage_table_stages:
        # All tickets in this flow (role-scoped)
        cq = db.query(FMSTicket).filter(
            FMSTicket.flow_id == active_flow.id,
            FMSTicket.is_deleted == False,
        )
        if user.role == "MANAGER":
            cq = cq.filter(
                (FMSTicket.current_assignee_id.in_(team_ids)) |
                (FMSTicket.id.in_(mgr_all_fms_ids))
            )
        elif user.role == "EMPLOYEE":
            cq = cq.filter(
                (FMSTicket.current_assignee_id == user.id) |
                (FMSTicket.id.in_(emp_all_fms_ids)) |
                (FMSTicket.id.in_(emp_upcoming_ids))
            )
        if priority_filter:
            cq = cq.filter(FMSTicket.priority.in_(priority_filter))
        if filter_assignee_ids is not None:
            cq = cq.filter(FMSTicket.current_assignee_id.in_(filter_assignee_ids))
        if filter_date_from:
            cq = cq.filter(FMSTicket.created_at >= filter_date_from)
        if filter_date_to:
            cq = cq.filter(FMSTicket.created_at <= filter_date_to)
        all_flow_tickets = cq.order_by(FMSTicket.created_at.desc()).all()

        for t in all_flow_tickets:
            # Get the most recent history row per stage
            histories = db.query(FMSStageHistory).filter(
                FMSStageHistory.ticket_id == t.id
            ).order_by(FMSStageHistory.entered_at).all()

            stage_data = {}  # stage_id -> dict
            for h in histories:
                cf_data = {}
                if h.custom_fields_data_json:
                    try:
                        cf_data = _json.loads(h.custom_fields_data_json)
                    except Exception:
                        cf_data = {}
                duration_h = None
                if h.entered_at and h.exited_at:
                    duration_h = round((h.exited_at - h.entered_at).total_seconds() / 3600, 1)
                # Keep latest visit per stage
                stage_data[h.stage_id] = {
                    "assignee_name": h.assignee.name if h.assignee else "—",
                    "entered_at":    h.entered_at,
                    "exited_at":     h.exited_at,
                    "duration_h":    duration_h,
                    "planned_start": h.planned_start,
                    "planned_end":   h.planned_end,
                    "cf":            cf_data,
                    "is_active":     h.exited_at is None,
                }

            consolidated_rows.append({
                "ticket": t,
                "stage_data": stage_data,
            })

    from .linked_entities import get_linked_entity_options as _geo
    entity_options = _geo(db, tid)

    # Whether the current user may create new tickets in the active flow —
    # admins/managers always can; an employee can too if this flow's
    # "Allowed Employees" whitelist includes them (see _can_create_in_flow).
    can_create_ticket = _can_create_in_flow(user, active_flow) if active_flow else (user.role in ("ADMIN", "MANAGER"))

    # Per-ticket manager-override window flag for stage view (2h after BACKWARD move)
    override_eligible: set = set()
    if user.role in ("ADMIN", "MANAGER") and active_stage and view == "stage":
        for row in stage_tickets:
            t = row["ticket"]
            last_back = db.query(FMSStageHistory).filter(
                FMSStageHistory.ticket_id == t.id,
                FMSStageHistory.direction == "BACKWARD",
            ).order_by(FMSStageHistory.entered_at.desc()).first()
            if last_back and (now - last_back.entered_at).total_seconds() < 7200:
                override_eligible.add(t.id)

    ticket_form_fields = []
    if active_flow and active_flow.ticket_form_fields_json:
        try:
            ticket_form_fields = [
                f for f in _json.loads(active_flow.ticket_form_fields_json)
                if f.get("field_type") not in ("__priority__", "__due_date__")
            ]
        except Exception:
            ticket_form_fields = []

    split_last_cumulative_json = _json_module.dumps({
        row["split_id"]: row["split_last_cumulative"]
        for row in stage_tickets
        if row.get("split_id") and row.get("split_last_cumulative") is not None
    })

    template_name = "fms/dashboard.html"
    return templates.TemplateResponse(request, template_name, _ctx(
        request, user, db,
        can_create_ticket=can_create_ticket,
        flows=flows, active_flow=active_flow,
        flow_counts=flow_counts,
        dropdown_ungrouped_flows=dropdown_ungrouped_flows,
        dropdown_flow_groups=dropdown_flow_groups,
        view=view,
        ticket_form_fields=ticket_form_fields,
        split_last_cumulative_json=split_last_cumulative_json,
        # Stage view (formerly stage_table)
        stage_table_stages=stage_table_stages,
        active_stage=active_stage,
        my_work_stage_ids=my_work_stage_ids,
        stage_tickets=stage_tickets,
        stage_ticket_counts=stage_ticket_counts,
        next_stage_map=next_stage_map,
        prev_stage_map=prev_stage_map,
        override_eligible=override_eligible,
        # Table view (per-ticket flat list with planned/actual dates)
        table_tickets=table_tickets,
        # Timeline view
        timeline_data=timeline_data,
        # Log view — consolidated audit trail
        log_rows=log_rows,
        log_columns=log_columns,
        log_event_types=log_event_types,
        f_log_event_type=list(log_event_type),
        f_log_actor_id=log_actor_id or "",
        f_log_search=log_search or "",
        # swimlane (legacy)
        tickets_by_stage=tickets_by_stage,
        tat_info=tat_info,
        flagged_tickets=flagged_tickets,
        can_drag=can_drag,
        # consolidated (legacy — no longer shown in toggle)
        consolidated_rows=consolidated_rows,
        # list (legacy)
        list_tickets=list_tickets,
        departments=departments,
        managers=managers,
        branches=branches,
        # active filters (lists for multi-select)
        f_dept_id=list(dept_id),
        f_manager_id=list(manager_id),
        f_branch_id=list(branch_id),
        f_month=month or "",
        f_status=status_filter or "",
        f_priority=list(f_priority),
        f_assignee_id=list(f_assignee_id),
        f_date_from=date_from or "",
        f_date_to=date_to or "",
        employees=employees,
        entity_options=entity_options,
        # role-relative ticket classification for employee board symbols
        emp_upcoming_ids=emp_upcoming_ids,
        emp_all_fms_ids=emp_all_fms_ids,
        my_work=my_work,
        # summary strip
        active_tickets=active_tickets,
        tat_breaches=tat_breaches,
        flagged_count=flagged_count,
        awaiting_count=awaiting_count,
        compliance=compliance,
        now=now,
        ref_lists_json=_build_ref_lists_json(user.tenant_id, db),
    ))


@router.get("/tickets/new", response_class=HTMLResponse)
def fms_ticket_new(
    request: Request, flow_id: Optional[str] = None,
    user: User = Depends(require_manager_or_redirect), db: Session = Depends(get_db),
):
    """2-C-1: Ticket creation form."""
    flows = db.query(FMSFlow).filter(
        FMSFlow.tenant_id == user.tenant_id, FMSFlow.is_active == True,
        FMSFlow.is_deleted == False).all()
    employees = db.query(User).filter(
        User.tenant_id == user.tenant_id, User.is_deleted == False,
        User.is_active == True).all()
    selected_flow = None
    if flow_id:
        selected_flow = next((f for f in flows if f.id == flow_id), None)
    from .linked_entities import get_linked_entity_options as _geo
    entity_options = _geo(db, user.tenant_id)
    return templates.TemplateResponse(request, "fms/ticket_new.html", _ctx(
        request, user, db,
        flows=flows, employees=employees,
        selected_flow=selected_flow,
        priorities=PRIORITIES,
        entity_options=entity_options,
    ))


@router.post("/tickets/new")
async def fms_ticket_create(
    request: Request,
    title: str = Form(...), description: str = Form(""),
    flow_id: str = Form(...), starting_stage_id: str = Form(...),
    priority: str = Form("MEDIUM"), assignee_id: str = Form(...),
    wo_number: str = Form(""), due_at: str = Form(""),
    target_qty: str = Form(""), qty_unit: str = Form(""),
    evidence_required: bool = Form(False),
    user: User = Depends(get_current_user), db: Session = Depends(get_db),
):
    """2-C-1 / P7-06: Create FMS ticket with evidence_required + linked entities."""
    import json as _json
    flow = _get_flow(db, flow_id, user.tenant_id)
    if not _can_create_in_flow(user, flow):
        raise HTTPException(403, "Not authorised to create tickets in this flow")
    stage = db.query(FMSStage).filter(
        FMSStage.id == starting_stage_id,
        FMSStage.flow_id == flow_id).first()
    if not stage:
        raise HTTPException(400, "Invalid starting stage")

    # Collect per-stage pre-assignments: stage_assignee_<stage_id>
    form_data = dict(await request.form())

    # Collect ticket creation form fields (defined in flow's ticket_form_fields_json)
    ticket_form_fields = _json.loads(flow.ticket_form_fields_json or "[]")
    ticket_custom_fields: dict = {}
    for cf in ticket_form_fields:
        fid   = cf.get("id", "")
        ftype = cf.get("field_type", "")
        val   = str(form_data.get(f"cf__{fid}", "") or "").strip()
        if ftype == "__priority__":
            if val and val.upper() in ("LOW", "MEDIUM", "HIGH", "CRITICAL"):
                priority = val.upper()
            continue
        if ftype == "__due_date__":
            if val:
                try:
                    due_at = val  # passed through to ticket creation below
                except Exception:
                    pass
            continue
        if val:
            ticket_custom_fields[fid] = val
    stage_assignees = {
        k[len("stage_assignee_"):]: v
        for k, v in form_data.items()
        if k.startswith("stage_assignee_") and v.strip()
    }
    stage_assignees[stage.id] = assignee_id
    stage_assignees_json = _json.dumps(stage_assignees)

    # Build stage schedule from form: stage_planned_end_<stage_id>
    # planned_start of stage N = planned_end of stage N-1
    # Auto-filled by JS from start_date + cumulative TAT, user can override per stage
    all_flow_stages = sorted(
        [s for s in flow.stages if not s.is_deleted], key=lambda s: s.order
    )
    stage_schedule: dict = {}
    start_date_str = form_data.get("schedule_start_date", "").strip()
    if start_date_str:
        from .notifications import add_business_hours
        tenant = db.query(Tenant).get(user.tenant_id)
        try:
            cursor = datetime.fromisoformat(start_date_str)
            for fs in all_flow_stages:
                p_end_str = form_data.get(f"stage_planned_end_{fs.id}", "").strip()
                if p_end_str:
                    p_end = datetime.fromisoformat(p_end_str)
                else:
                    tat_h = fs.target_tat_hours or 24
                    p_end = add_business_hours(tenant, cursor, tat_h)
                stage_schedule[fs.id] = {
                    "planned_start": cursor.isoformat(),
                    "planned_end":   p_end.isoformat(),
                }
                cursor = p_end
        except (ValueError, TypeError):
            stage_schedule = {}
    stage_schedule_json = _json.dumps(stage_schedule) if stage_schedule else None

    # Planned dates for the first stage history row
    first_sched = stage_schedule.get(stage.id, {})
    first_ps = datetime.fromisoformat(first_sched["planned_start"]) if first_sched.get("planned_start") else None
    first_pe = datetime.fromisoformat(first_sched["planned_end"])   if first_sched.get("planned_end")   else None

    ticket = FMSTicket(
        tenant_id=user.tenant_id, flow_id=flow_id,
        current_stage_id=stage.id, title=title,
        description=description or None,
        wo_number=wo_number or None, priority=priority,
        target_qty=int(target_qty) if target_qty.strip() else None,
        qty_unit=qty_unit or None,
        current_assignee_id=assignee_id,
        due_at=datetime.fromisoformat(due_at) if due_at.strip() else None,
        created_by_id=user.id, status="ACTIVE",
        stage_assignees_json=stage_assignees_json,
        stage_schedule_json=stage_schedule_json,
        ticket_custom_fields_json=_json.dumps(ticket_custom_fields) if ticket_custom_fields else None,
    )
    db.add(ticket)
    db.flush()

    tenant = db.query(Tenant).get(user.tenant_id)
    ticket.display_id = _next_fms_display_id(db, tenant)

    split = _init_first_split(db, ticket, stage.id, assignee_id)
    db.add(FMSStageHistory(
        ticket_id=ticket.id, split_id=split.id, stage_id=stage.id,
        stage_name=stage.name, assignee_id=assignee_id,
        direction="FORWARD",
        planned_start=first_ps,
        planned_end=first_pe,
    ))
    _assignee_obj = db.query(User).filter(User.id == assignee_id).first() if assignee_id else None
    _created_detail = (
        f"Title: {title} | Priority: {priority or '—'} | Qty: {target_qty or '—'} {qty_unit or ''} | "
        f"Due: {ticket.due_at.strftime('%d %b %Y') if ticket.due_at else '—'} | "
        f"Assignee: {_assignee_obj.name if _assignee_obj else '—'} | "
        f"Custom fields: {_fmt_cf(ticket_custom_fields)}"
    )
    _log(db, ticket.id, user.id, "CREATED", _created_detail, meta={
        "priority": priority, "qty": target_qty, "qty_unit": qty_unit,
        "due_at": ticket.due_at.isoformat() if ticket.due_at else None,
        "assignee_name": _assignee_obj.name if _assignee_obj else None,
        "custom_fields": _cf_by_label(ticket_custom_fields, ticket_form_fields),
    })
    _log(db, ticket.id, user.id, "STAGE_ENTERED",
         f"Stage: {stage.name} | TAT window: {_fmt_window(first_ps, first_pe)} | "
         f"Target TAT: {stage.target_tat_hours or '—'}h", meta={
             "stage_name": stage.name, "assignee_name": _assignee_obj.name if _assignee_obj else None,
             "tat_window": _fmt_window(first_ps, first_pe), "target_tat_hours": stage.target_tat_hours,
         })
    db.commit()
    db.refresh(ticket)

    # P7-06: save linked entities
    from .linked_entities import save_linked_entities_from_form as _slf
    _slf(db, form_data, "FMS_TICKET", ticket.id, user.tenant_id, user.id)

    admins   = _admin_ids(db, user.tenant_id)
    managers = _manager_ids_for(db, assignee_id)
    notify_fms_stage_transition(
        db, user.tenant_id, ticket.id, ticket.title,
        stage.name, user.id, admins, managers, assignee_id)

    assignee_obj = db.query(User).filter(User.id == assignee_id).first()
    if assignee_obj:
        send_whatsapp_for_fms_ticket_created(db, ticket, assignee_obj)
    notify_fms_ticket_opened(db, ticket, assignee_obj, admins, managers)

    return _redirect(
        f"/fms/dashboard?view=stage&flow_id={flow_id}&stage_id={stage.id}&msg=Ticket+created"
    )


# ── P7-08: Edit FMS Ticket ────────────────────────────────────────────────────

@router.post("/tickets/{ticket_id}/edit")
def fms_ticket_edit(
    ticket_id: str,
    title: str = Form(...), description: str = Form(""),
    priority: str = Form("MEDIUM"),
    due_at: str = Form(""), wo_number: str = Form(""),
    user: User = Depends(require_manager), db: Session = Depends(get_db),
):
    ticket = _get_ticket(db, ticket_id, user.tenant_id)
    if ticket.status in ("COMPLETED", "CLOSED"):
        raise HTTPException(400, "Cannot edit a completed or closed ticket")
    old = f"title={ticket.title}, priority={ticket.priority}"
    ticket.title = title
    ticket.description = description or None
    ticket.priority = priority
    ticket.wo_number = wo_number or None
    ticket.due_at = datetime.fromisoformat(due_at) if due_at.strip() else ticket.due_at
    ticket.updated_at = datetime.utcnow()
    _log(db, ticket_id, user.id, "EDITED", f"Was: {old}")
    db.commit()
    return _redirect(
        f"/fms/dashboard?view=stage"
        f"{'&flow_id=' + ticket.flow_id if ticket.flow_id else ''}"
        f"{'&stage_id=' + ticket.current_stage_id if ticket.current_stage_id else ''}"
    )


# ── P7-08: Delete FMS Ticket ──────────────────────────────────────────────────

@router.post("/tickets/{ticket_id}/delete")
def fms_ticket_delete(
    ticket_id: str,
    flow_id: str = Form(""),
    stage_id: str = Form(""),
    user: User = Depends(require_admin),
    db: Session = Depends(get_db),
):
    ticket = _get_ticket(db, ticket_id, user.tenant_id)
    if ticket.status in ("COMPLETED", "CLOSED"):
        return_url = f"/fms/dashboard?view=stage"
        if flow_id: return_url += f"&flow_id={flow_id}"
        if stage_id: return_url += f"&stage_id={stage_id}"
        return _redirect(return_url + "&err=Tickets+with+terminal+status+cannot+be+deleted")
    history_count = db.query(FMSStageHistory).filter(
        FMSStageHistory.ticket_id == ticket_id,
        FMSStageHistory.exited_at != None,
    ).count()
    if history_count > 0:
        return_url = f"/fms/dashboard?view=stage"
        if flow_id: return_url += f"&flow_id={flow_id}"
        if stage_id: return_url += f"&stage_id={stage_id}"
        return _redirect(return_url + "&err=Tickets+with+activity+cannot+be+deleted")
    ticket.is_deleted = True
    _log(db, ticket_id, user.id, "DELETED", "Soft deleted by admin")
    db.commit()
    return_url = f"/fms/dashboard?view=stage"
    if flow_id: return_url += f"&flow_id={flow_id}"
    if stage_id: return_url += f"&stage_id={stage_id}"
    return _redirect(return_url)


@router.post("/tickets/{ticket_id}/notify")
def fms_ticket_notify(
    ticket_id: str,
    flow_id: str = Form(""),
    stage_id: str = Form(""),
    user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """Send a manual in-app reminder notification to the current stage assignee."""
    if user.role not in ("ADMIN", "MANAGER"):
        raise HTTPException(403, "Not authorised")
    ticket = _get_ticket(db, ticket_id, user.tenant_id)
    if not ticket.current_assignee_id:
        return_url = f"/fms/dashboard?view=stage"
        if flow_id: return_url += f"&flow_id={flow_id}"
        if stage_id: return_url += f"&stage_id={stage_id}"
        return _redirect(return_url + "&err=Ticket+has+no+assignee+to+notify")
    stage_name = ticket.current_stage.name if ticket.current_stage else "current stage"
    create_notification(
        db, user.tenant_id,
        user_id=ticket.current_assignee_id,
        notif_type="FMS_REMINDER",
        title=f"Reminder: {ticket.title}",
        body=f"This flow ticket is waiting for your action at stage {stage_name}.",
        link=f"/fms/dashboard?view=stage&flow_id={ticket.flow_id}&stage_id={ticket.current_stage_id}",
    )
    db.commit()
    assignee = db.query(User).filter(User.id == ticket.current_assignee_id).first()
    assignee_name = assignee.name if assignee else "assignee"
    return_url = f"/fms/dashboard?view=stage"
    if flow_id: return_url += f"&flow_id={flow_id}"
    if stage_id: return_url += f"&stage_id={stage_id}"
    from urllib.parse import quote
    return _redirect(return_url + f"&msg=Notification+sent+to+{quote(assignee_name)}")


@router.get("/tickets/{ticket_id}/events")
def fms_ticket_events(
    ticket_id: str,
    user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """Append-only audit trail for a ticket — every action (by whom, when,
    what) across the ticket and all of its splits. Backs the History modal."""
    from fastapi.responses import JSONResponse
    ticket = _get_ticket(db, ticket_id, user.tenant_id)
    events = db.query(FMSEvent).filter(
        FMSEvent.ticket_id == ticket.id,
    ).order_by(FMSEvent.created_at.desc()).all()
    actor_ids = {e.actor_id for e in events if e.actor_id}
    actors = {u.id: u.name for u in db.query(User).filter(User.id.in_(actor_ids)).all()} if actor_ids else {}
    return JSONResponse({"events": [
        {
            "event_type": e.event_type,
            "detail": e.detail or "",
            "actor_name": actors.get(e.actor_id, "System"),
            "created_at": e.created_at.strftime("%d %b %Y, %H:%M") if e.created_at else "",
        }
        for e in events
    ]})


# ── P7-07: Bulk upload FMS tickets ────────────────────────────────────────────

@router.get("/tickets/bulk-template")
def fms_bulk_template(
    flow_id: str = "",
    user: User = Depends(require_manager),
    db: Session = Depends(get_db),
):
    import io as _io
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    flow = None
    if flow_id:
        flow = db.query(FMSFlow).filter(
            FMSFlow.id == flow_id,
            FMSFlow.tenant_id == user.tenant_id,
            FMSFlow.is_deleted == False,
        ).first()
    if not flow:
        flow = db.query(FMSFlow).filter(
            FMSFlow.tenant_id == user.tenant_id,
            FMSFlow.is_active == True,
            FMSFlow.is_deleted == False,
        ).order_by(FMSFlow.name).first()

    stages = sorted([s for s in flow.stages if not s.is_deleted], key=lambda s: s.order) if flow else []

    wb = Workbook()
    ws = wb.active
    ws.title = "FMS Tickets"

    base_headers = ["Title *", "Priority", "Due Date (YYYY-MM-DD)", "WO Number", "Target Qty", "Qty Unit", "TaT Unit (Days/Hours)"]
    stage_headers = []
    for s in stages:
        stage_headers.append(f"{s.name} Assignee Phone")
        stage_headers.append(f"{s.name} TaT")
    all_headers = base_headers + stage_headers

    hdr_fill = PatternFill("solid", fgColor="1E293B")
    hdr_font = Font(bold=True, color="FFFFFF", size=11)
    for col, h in enumerate(all_headers, 1):
        c = ws.cell(row=1, column=col, value=h)
        c.font = hdr_font; c.fill = hdr_fill
        c.alignment = Alignment(horizontal="center", wrap_text=True)

    inst_fill = PatternFill("solid", fgColor="374151")
    inst_font = Font(italic=True, color="9CA3AF", size=10)
    instructions = [
        "Required. Max 200 chars", "LOW / MEDIUM / HIGH / CRITICAL",
        "e.g. 2026-08-15", "Optional WO/PO ref",
        "Optional integer", "Optional: pcs/kg/m",
        "Days or Hours — applies to all TaT columns in this sheet",
    ]
    for s in stages:
        instructions.append(f"Phone of assignee for {s.name} (blank = flow default)")
        instructions.append(f"TaT for {s.name} (blank = flow default {s.target_tat_hours or '—'})")
    for col, inst in enumerate(instructions, 1):
        c = ws.cell(row=2, column=col, value=inst)
        c.font = inst_font; c.fill = inst_fill
        c.alignment = Alignment(wrap_text=True)

    sample = ["Sample Ticket WO-001", "MEDIUM", "2026-08-20", "WO-001", "100", "pcs", "Hours"]
    for s in stages:
        sample.append("")
        sample.append(str(s.target_tat_hours) if s.target_tat_hours else "24")
    for col, val in enumerate(sample, 1):
        ws.cell(row=3, column=col, value=val)

    col_widths = [30, 12, 18, 14, 10, 10, 20]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    for i in range(len(stages)):
        ws.column_dimensions[get_column_letter(8 + i * 2 - 1)].width = 22
        ws.column_dimensions[get_column_letter(8 + i * 2)].width = 12
    ws.freeze_panes = "A3"

    buf = _io.BytesIO()
    wb.save(buf); buf.seek(0)
    fname = f"fms_{flow.name.replace(' ', '_')}_template.xlsx" if flow else "fms_template.xlsx"
    return StreamingResponse(
        iter([buf.read()]),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={fname}"},
    )


def _fms_stage_cols(stages: list) -> tuple:
    base_headers = ["Title *", "Priority", "Due Date (YYYY-MM-DD)", "WO Number", "Target Qty", "Qty Unit", "TaT Unit (Days/Hours)"]
    stage_headers = []
    for s in stages:
        stage_headers.append(f"{s.name} Assignee Phone")
        stage_headers.append(f"{s.name} TaT")
    return base_headers, stage_headers


def _get_active_flow(flow_id: str, tenant_id: str, db: Session):
    flow = None
    if flow_id:
        flow = db.query(FMSFlow).filter(
            FMSFlow.id == flow_id, FMSFlow.tenant_id == tenant_id, FMSFlow.is_deleted == False,
        ).first()
    if not flow:
        flow = db.query(FMSFlow).filter(
            FMSFlow.tenant_id == tenant_id, FMSFlow.is_active == True, FMSFlow.is_deleted == False,
        ).order_by(FMSFlow.name).first()
    return flow


def _parse_fms_row(row: dict, stages: list, tenant_id: str, db: Session) -> tuple:
    """row is a dict keyed by the template's header strings. Returns (parsed_dict, error) —
    parsed_dict carries everything confirm needs to create the ticket without re-touching the DB
    beyond what's necessary; unresolvable stage phones fall back to the flow's default (non-blocking)."""
    def cell(key):
        v = row.get(key)
        return str(v).strip() if v not in (None, "") else ""

    title = cell("Title *")
    if not title:
        return None, "Title is required"
    title = title[:200]
    priority = cell("Priority").upper() or "MEDIUM"
    if priority not in ("LOW", "MEDIUM", "HIGH", "CRITICAL"):
        priority = "MEDIUM"
    due_date_str = cell("Due Date (YYYY-MM-DD)")
    due_date = None
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%m/%d/%Y", "%Y-%m-%d %H:%M"):
        try:
            due_date = datetime.strptime(due_date_str, fmt); break
        except ValueError:
            pass
    if due_date is None:
        return None, f"Invalid Due Date '{due_date_str}' — expected YYYY-MM-DD"

    wo_number = cell("WO Number")
    target_qty_str = cell("Target Qty")
    qty_unit = cell("Qty Unit")
    tat_unit_str = cell("TaT Unit (Days/Hours)").lower() or "hours"
    tat_mult = 24.0 if "day" in tat_unit_str else (1 / 60.0 if "min" in tat_unit_str else 1.0)

    from .notifications import add_business_hours
    tenant = db.query(Tenant).get(tenant_id)

    stage_assignees: dict = {}
    stage_schedule: dict = {}
    cursor = datetime.utcnow()
    for s in stages:
        assignee_id = _stage_default_assignee(s)
        phone = cell(f"{s.name} Assignee Phone")
        if phone:
            u = db.query(User).filter(
                User.tenant_id == tenant_id, User.phone == phone, User.is_deleted == False,
            ).first()
            if u:
                assignee_id = u.id
        if assignee_id:
            stage_assignees[s.id] = assignee_id

        tat_hours = float(s.target_tat_hours or 24)
        raw = cell(f"{s.name} TaT")
        if raw:
            try:
                tat_hours = float(raw) * tat_mult
            except (ValueError, TypeError):
                pass
        p_end = add_business_hours(tenant, cursor, tat_hours)
        stage_schedule[s.id] = {"planned_start": cursor.isoformat(), "planned_end": p_end.isoformat()}
        cursor = p_end

    return {
        "title": title, "priority": priority, "due_at": due_date.isoformat(),
        "wo_number": wo_number or None,
        "target_qty": int(target_qty_str) if target_qty_str.isdigit() else None,
        "qty_unit": qty_unit or None,
        "stage_assignees": stage_assignees, "stage_schedule": stage_schedule,
    }, None


def _run_fms_validation(rows_in: list, stages: list, tenant_id: str, db: Session, start_index: int = 3) -> dict:
    valid_rows, errors = [], []
    for i, row in enumerate(rows_in, start=start_index):
        if not any(v not in (None, "", "_row") for k, v in row.items() if k != "_row"):
            continue
        parsed, error = _parse_fms_row(row, stages, tenant_id, db)
        if error:
            errors.append({"row": row.get("_row", i), "error": error, "data": {k: v for k, v in row.items() if k != "_row"}})
        else:
            valid_rows.append(parsed)
    return {
        "total": len(valid_rows) + len(errors),
        "valid": len(valid_rows),
        "errors": errors,
        "rows": valid_rows,
    }


@router.get("/tickets/bulk-upload-page", response_class=HTMLResponse)
def fms_bulk_upload_page(request: Request, flow_id: str = "", user: User = Depends(require_manager_or_redirect), db: Session = Depends(get_db)):
    flows = db.query(FMSFlow).filter(
        FMSFlow.tenant_id == user.tenant_id, FMSFlow.is_active == True, FMSFlow.is_deleted == False,
    ).order_by(FMSFlow.name).all()
    flow = _get_active_flow(flow_id, user.tenant_id, db)
    columns = []
    if flow:
        stages = sorted([s for s in flow.stages if not s.is_deleted], key=lambda s: s.order)
        base_headers, stage_headers = _fms_stage_cols(stages)
        columns = base_headers + stage_headers
    return templates.TemplateResponse(request, "fms/bulk_upload.html", _ctx(
        request, user, db, flows=flows, active_flow=flow, columns=columns,
    ))


@router.post("/tickets/bulk-upload")
async def fms_bulk_upload(
    request: Request,
    file: UploadFile = File(...),
    flow_id: str = Form(...),
    user: User = Depends(require_manager),
    db: Session = Depends(get_db),
):
    import io as _io
    from openpyxl import load_workbook

    flow = db.query(FMSFlow).filter(
        FMSFlow.id == flow_id, FMSFlow.tenant_id == user.tenant_id,
        FMSFlow.is_active == True, FMSFlow.is_deleted == False,
    ).first()
    if not flow:
        raise HTTPException(400, "Flow not found")
    stages = sorted([s for s in flow.stages if not s.is_deleted], key=lambda s: s.order)
    if not stages:
        raise HTTPException(400, "Flow has no stages")

    content = await file.read()
    try:
        wb = load_workbook(filename=_io.BytesIO(content), read_only=True, data_only=True)
    except Exception:
        raise HTTPException(400, "Invalid Excel file — please use the downloaded template.")

    ws = wb.active
    all_rows = list(ws.iter_rows(values_only=True))
    if len(all_rows) < 3:
        raise HTTPException(400, "File too short.")
    if len(all_rows) - 2 > BULK_IMPORT_MAX_ROWS:
        raise HTTPException(400, f"File has too many rows — maximum {BULK_IMPORT_MAX_ROWS}.")

    headers = [str(h).strip() if h is not None else "" for h in all_rows[0]]
    base_headers, stage_headers = _fms_stage_cols(stages)
    fmt_err = check_required_headers(headers, ["Title *", "Due Date (YYYY-MM-DD)"], base_headers + stage_headers)
    if fmt_err:
        return JSONResponse({"format_error": fmt_err})
    rows = []
    for row_num, row in enumerate(all_rows[2:], start=3):
        d = {headers[i]: (row[i] if i < len(row) else None) for i in range(len(headers))}
        d["_row"] = row_num
        rows.append(d)

    return JSONResponse(_run_fms_validation(rows, stages, user.tenant_id, db))


@router.post("/tickets/bulk-upload/revalidate")
async def fms_bulk_revalidate(request: Request, user: User = Depends(require_manager), db: Session = Depends(get_db)):
    body = await request.json()
    rows_in = body.get("rows", [])
    flow_id = body.get("flow_id", "")
    flow = _get_active_flow(flow_id, user.tenant_id, db)
    if not flow:
        raise HTTPException(400, "Flow not found")
    stages = sorted([s for s in flow.stages if not s.is_deleted], key=lambda s: s.order)
    if len(rows_in) > BULK_IMPORT_MAX_ROWS:
        raise HTTPException(400, f"Too many rows — maximum allowed is {BULK_IMPORT_MAX_ROWS}.")
    return JSONResponse(_run_fms_validation(rows_in, stages, user.tenant_id, db))


@router.post("/tickets/bulk-upload/confirm")
async def fms_bulk_confirm(request: Request, user: User = Depends(require_manager), db: Session = Depends(get_db)):
    body = await request.json()
    rows = body.get("rows", [])
    flow_id = body.get("flow_id", "")
    flow = _get_active_flow(flow_id, user.tenant_id, db)
    if not flow:
        raise HTTPException(400, "Flow not found")
    stages = sorted([s for s in flow.stages if not s.is_deleted], key=lambda s: s.order)
    if not stages:
        raise HTTPException(400, "Flow has no stages")

    tenant = db.query(Tenant).get(user.tenant_id)
    created = 0
    for r in rows:
        stage_assignees = r.get("stage_assignees") or {}
        stage_schedule = r.get("stage_schedule") or {}
        first_assignee_id = stage_assignees.get(stages[0].id) or _stage_default_assignee(stages[0])
        first_sched = stage_schedule.get(stages[0].id, {})

        try:
            due_at = datetime.fromisoformat(r["due_at"])
        except (KeyError, ValueError):
            continue

        ticket = FMSTicket(
            tenant_id=user.tenant_id, flow_id=flow.id,
            current_stage_id=stages[0].id, title=r["title"],
            priority=r["priority"], wo_number=r.get("wo_number"),
            target_qty=r.get("target_qty"), qty_unit=r.get("qty_unit"),
            current_assignee_id=first_assignee_id,
            due_at=due_at, created_by_id=user.id, status="ACTIVE",
            stage_assignees_json=_json.dumps(stage_assignees) if stage_assignees else None,
            stage_schedule_json=_json.dumps(stage_schedule) if stage_schedule else None,
        )
        db.add(ticket); db.flush()
        ticket.display_id = _next_fms_display_id(db, tenant)
        split = _init_first_split(db, ticket, stages[0].id, first_assignee_id)
        db.add(FMSStageHistory(
            ticket_id=ticket.id, split_id=split.id, stage_id=stages[0].id,
            stage_name=stages[0].name, assignee_id=first_assignee_id,
            direction="FORWARD",
            planned_start=datetime.fromisoformat(first_sched["planned_start"]) if first_sched.get("planned_start") else None,
            planned_end=datetime.fromisoformat(first_sched["planned_end"]) if first_sched.get("planned_end") else None,
        ))
        _log(db, ticket.id, user.id, "CREATED", f"Bulk import: {r['title']}")
        created += 1

    try:
        db.commit()
    except Exception as e:
        db.rollback()
        raise HTTPException(400, f"Import failed — no tickets were created. {e}")
    return JSONResponse({"created": created})


@router.get("/api/flow/{flow_id}/defaults")
def fms_api_flow_defaults(
    flow_id: str,
    user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """Return stage names, default assignee names, and TaT values for bulk-create AJAX."""
    from fastapi.responses import JSONResponse
    flow = _get_flow(db, flow_id, user.tenant_id)
    stages = sorted([s for s in flow.stages if not s.is_deleted], key=lambda s: s.order)
    result = []
    for s in stages:
        default_id = _stage_default_assignee(s)
        assignee_name = None
        if default_id:
            u = db.query(User).get(default_id)
            assignee_name = u.name if u else None
        result.append({
            "id": s.id,
            "name": s.name,
            "default_assignee_name": assignee_name,
            "default_assignee_id": default_id,
            "default_assignee_ids": _stage_default_assignee_ids(s),
            "target_tat_hours": s.target_tat_hours,
            "order": s.order,
        })
    employees = db.query(User).filter(
        User.tenant_id == user.tenant_id,
        User.is_deleted == False,
        User.is_active == True,
    ).order_by(User.name).all()
    emp_list = [{"id": e.id, "name": e.name} for e in employees]
    ticket_form_fields = _json.loads(flow.ticket_form_fields_json or "[]")
    # Resolve ref_list options so the bulk-create table can render a dropdown
    for f in ticket_form_fields:
        if f.get("field_type") == "ref_list" and f.get("ref_list_id") and not f.get("options"):
            from .database import CustomReferenceList, CustomReferenceItem, Customer, Vendor, RawMaterial
            rid = f["ref_list_id"]
            if rid.startswith("__system_"):
                _sys_map = {
                    "__system_customer__": (Customer, "name"),
                    "__system_vendor__": (Vendor, "name"),
                    "__system_rawmaterial__": (RawMaterial, "name"),
                    "__system_employee__": (User, "name"),
                }
                model, col = _sys_map.get(rid, (None, None))
                if model:
                    rows = db.query(model).filter(model.tenant_id == user.tenant_id, model.is_deleted == False).all()
                    f["options"] = [getattr(r, col) for r in rows if getattr(r, col, None)]
            else:
                lst = db.query(CustomReferenceList).filter(
                    CustomReferenceList.id == rid,
                    CustomReferenceList.tenant_id == user.tenant_id,
                ).first()
                if lst:
                    f["options"] = [i.value for i in lst.items if i.is_active and not i.is_deleted]
    return JSONResponse({"stages": result, "employees": emp_list, "ticket_form_fields": ticket_form_fields})


@router.get("/tickets/bulk-create", response_class=HTMLResponse)
def fms_bulk_create_get(
    request: Request,
    flow_id: Optional[str] = Query(default=None),
    user: User = Depends(get_current_user_or_redirect),
    db: Session = Depends(get_db),
):
    """A3-1: Bulk ticket creation form."""
    # Role-filtered flows (same logic as dashboard dropdown). An employee
    # sees a flow here only if they're on that flow's "Allowed Employees"
    # whitelist (_can_create_in_flow) — being whitelisted to open/act on a
    # flow's tickets unlocks creating tickets in it too.
    all_flows = db.query(FMSFlow).filter(
        FMSFlow.tenant_id == user.tenant_id, FMSFlow.is_active == True,
        FMSFlow.is_deleted == False,
    ).order_by(FMSFlow.name).all()
    if user.role == "EMPLOYEE":
        flows = [f for f in all_flows if _can_create_in_flow(user, f)]
        if not flows:
            raise HTTPException(403, "Not authorised to create tickets in any flow")
    else:
        flows = all_flows
    # Pre-select flow passed from dashboard
    preselect_flow_id = flow_id if flow_id and any(f.id == flow_id for f in flows) else None
    return templates.TemplateResponse(request, "fms/bulk_create.html", _ctx(
        request, user, db, flows=flows, priorities=PRIORITIES,
        preselect_flow_id=preselect_flow_id,
    ))


@router.post("/tickets/bulk-create")
async def fms_bulk_create_post(
    request: Request,
    user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """A3-1: Process bulk ticket creation — all-or-nothing validation."""
    form = await request.form()
    flow_id = (form.get("flow_id") or "").strip()
    all_flows = db.query(FMSFlow).filter(
        FMSFlow.tenant_id == user.tenant_id, FMSFlow.is_active == True,
        FMSFlow.is_deleted == False,
    ).order_by(FMSFlow.name).all()
    _target_flow = next((f for f in flows if f.id == flow_id), None)
    if not _can_create_in_flow(user, _target_flow):
        raise HTTPException(403, "Not authorised to create tickets in this flow")

    def _reraise(error, row_errors=None):
        return templates.TemplateResponse(request, "fms/bulk_create.html", _ctx(
            request, user, db, flows=flows, priorities=PRIORITIES,
            error=error, row_errors=row_errors or [],
            saved_flow_id=flow_id, saved_form=form,
        ))

    if not flow_id:
        return _reraise("Please select a flow before submitting.")
    if user.role == "EMPLOYEE" and not any(f.id == flow_id for f in flows):
        raise HTTPException(403, "Not authorised to create tickets on this flow")

    flow = _get_flow(db, flow_id, user.tenant_id)
    stages = sorted([s for s in flow.stages if not s.is_deleted], key=lambda s: s.order)
    if not stages:
        return _reraise("Selected flow has no stages configured.")

    first_stage = stages[0]
    ticket_form_fields = _json.loads(flow.ticket_form_fields_json or "[]")

    try:
        row_count = int(form.get("row_count") or "0")
    except ValueError:
        row_count = 0

    if row_count < 1:
        return _reraise("At least 1 ticket row is required.")
    if row_count > 50:
        return _reraise("Maximum 50 rows per bulk create.")

    today = datetime.utcnow().date()
    errors = []
    tickets_data = []

    tat_unit = (form.get("tat_unit") or "hours").strip().lower()
    tat_mult = 24.0 if tat_unit == "days" else (1 / 60.0 if tat_unit == "minutes" else 1.0)

    for i in range(row_count):
        wo_number = (form.get(f"row_wo_number_{i}") or "").strip()
        target_qty_str = (form.get(f"row_target_qty_{i}") or "").strip()
        qty_unit = (form.get(f"row_qty_unit_{i}") or "").strip()

        row_errs = []
        due_date = None
        priority = "MEDIUM"

        # Validate custom ticket form fields (__priority__ and __due_date__ are built-in special types)
        custom_field_values: dict = {}
        for cf in ticket_form_fields:
            val = (form.get(f"row_cf_{cf['id']}_{i}") or "").strip()
            ftype = cf.get("field_type", "")
            if ftype == "__priority__":
                if val and val.upper() in PRIORITIES:
                    priority = val.upper()
                continue  # not stored in custom_fields, mapped to ticket.priority
            if ftype == "__due_date__":
                if val:
                    try:
                        due_date = datetime.strptime(val, "%Y-%m-%d")
                        if due_date.date() <= today:
                            row_errs.append("Due date must be a future date")
                    except ValueError:
                        row_errs.append("Invalid due date format")
                elif cf.get("required"):
                    row_errs.append(f"'{cf['label']}' is required")
                continue  # not stored in custom_fields, mapped to ticket.due_at
            if cf.get("required") and not val:
                row_errs.append(f"'{cf['label']}' is required")
            if val:
                custom_field_values[cf["id"]] = val

        # Title is always auto-generated from the sequence — display_id is the human identifier
        title = f"Ticket-{i + 1}"

        if row_errs:
            errors.append({"row": i + 1, "title": f"Row {i + 1}", "errors": row_errs})
        else:
            # Collect per-stage assignee and TaT
            from .notifications import add_business_hours
            _row_tenant = db.query(Tenant).get(user.tenant_id)
            stage_assignees: dict = {}
            stage_schedule: dict = {}
            cursor = datetime.utcnow()
            for s in stages:
                aid = (form.get(f"row_stage_assignee_{i}_{s.id}") or "").strip() or _stage_default_assignee(s)
                if aid:
                    stage_assignees[s.id] = aid
                tat_hours = float(s.target_tat_hours or 24)
                raw_tat = (form.get(f"row_stage_tat_{i}_{s.id}") or "").strip()
                if raw_tat:
                    try:
                        tat_hours = float(raw_tat) * tat_mult
                    except ValueError:
                        pass
                p_end = add_business_hours(_row_tenant, cursor, tat_hours)
                stage_schedule[s.id] = {"planned_start": cursor.isoformat(), "planned_end": p_end.isoformat()}
                cursor = p_end

            tickets_data.append({
                "title": title, "priority": priority,
                "due_date": due_date, "wo_number": wo_number or None,
                "target_qty": int(target_qty_str) if target_qty_str.isdigit() else None,
                "qty_unit": qty_unit or None,
                "stage_assignees": stage_assignees,
                "stage_schedule": stage_schedule,
                "custom_fields": custom_field_values,
            })

    if errors:
        return _reraise(None, errors)

    # All rows valid — create all tickets
    tenant = db.query(Tenant).get(user.tenant_id)
    admins = _admin_ids(db, user.tenant_id)

    created = []
    for td in tickets_data:
        first_assignee_id = td["stage_assignees"].get(first_stage.id) or _stage_default_assignee(first_stage)
        first_sched = td["stage_schedule"].get(first_stage.id, {})
        ticket = FMSTicket(
            tenant_id=user.tenant_id, flow_id=flow_id,
            current_stage_id=first_stage.id, title=td["title"],
            priority=td["priority"], due_at=td["due_date"],
            wo_number=td["wo_number"],
            target_qty=td["target_qty"], qty_unit=td["qty_unit"],
            current_assignee_id=first_assignee_id,
            created_by_id=user.id, status="ACTIVE",
            stage_assignees_json=_json.dumps(td["stage_assignees"]) if td["stage_assignees"] else None,
            stage_schedule_json=_json.dumps(td["stage_schedule"]) if td["stage_schedule"] else None,
            ticket_custom_fields_json=_json.dumps(td["custom_fields"]) if td.get("custom_fields") else None,
        )
        db.add(ticket); db.flush()
        ticket.display_id = _next_fms_display_id(db, tenant)
        split = _init_first_split(db, ticket, first_stage.id, first_assignee_id)
        db.add(FMSStageHistory(
            ticket_id=ticket.id, split_id=split.id, stage_id=first_stage.id,
            stage_name=first_stage.name, assignee_id=first_assignee_id,
            direction="FORWARD",
            planned_start=datetime.fromisoformat(first_sched["planned_start"]) if first_sched.get("planned_start") else None,
            planned_end=datetime.fromisoformat(first_sched["planned_end"]) if first_sched.get("planned_end") else None,
        ))
        _log(db, ticket.id, user.id, "CREATED", td["title"])
        created.append((ticket, first_assignee_id))

    db.commit()

    # Notify stage-1 assignee for each created ticket
    for ticket, first_aid in created:
        if first_aid:
            mgrs = _manager_ids_for(db, first_aid)
            notify_fms_stage_transition(
                db, user.tenant_id, ticket.id, ticket.title,
                first_stage.name, user.id, admins, mgrs, first_aid,
            )

    from urllib.parse import quote as _q
    n = len(created)
    msg = _q(f"{n} ticket{'s' if n != 1 else ''} created successfully")
    return _redirect(f"/fms/dashboard?view=stage&flow_id={flow_id}&msg={msg}")





@router.post("/tickets/bulk-transition")
async def fms_bulk_transition(
    request: Request,
    user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """A3-2: Bulk stage transition — partial success allowed."""
    if not _can_transition(user, FMSTicket(current_assignee_id=user.id)):
        # Allow admin/manager always; employees handled per-ticket below
        pass
    if user.role not in ("ADMIN", "MANAGER", "EMPLOYEE"):
        raise HTTPException(403, "Not authorised")

    form = await request.form()
    ticket_ids = form.getlist("ticket_ids")
    next_stage_id = (form.get("next_stage_id") or "").strip()
    flow_id = (form.get("flow_id") or "").strip()
    current_stage_id = (form.get("current_stage_id") or "").strip()

    if not ticket_ids or not next_stage_id:
        return _redirect(f"/fms/dashboard?view=stage&flow_id={flow_id}&err=Invalid+bulk+transition+request")

    if len(ticket_ids) > 20:
        return _redirect(f"/fms/dashboard?view=stage&flow_id={flow_id}&err=Maximum+20+tickets+per+bulk+transition")

    tid = user.tenant_id
    next_stage = db.query(FMSStage).filter(FMSStage.id == next_stage_id).first()
    if not next_stage:
        return _redirect(f"/fms/dashboard?view=stage&flow_id={flow_id}&err=Invalid+target+stage")

    admins = _admin_ids(db, tid)
    moved = 0
    skipped = []
    now = datetime.utcnow()

    for t_id in ticket_ids:
        ticket = db.query(FMSTicket).filter(
            FMSTicket.id == t_id,
            FMSTicket.tenant_id == tid,
            FMSTicket.is_deleted == False,
        ).first()
        if not ticket:
            skipped.append(f"{t_id[:8]}: not found")
            continue
        if ticket.status == "CLOSED":
            skipped.append(f"{ticket.display_id or t_id[:8]}: already closed")
            continue
        # Phase 0: resolve the active split(s) sitting at this stage — the common
        # case is exactly one (mirrors ticket.current_stage_id), but a ticket can
        # have more than one leaf split parked at the same stage.
        target_splits = [
            s for s in _active_splits(db, t_id)
            if s.status not in ("COMPLETED", "CLOSED")
            and (not current_stage_id or s.current_stage_id == current_stage_id)
        ]
        if not target_splits:
            skipped.append(f"{ticket.display_id or t_id[:8]}: no longer at this stage")
            continue
        if user.role == "EMPLOYEE" and not any(_can_transition(user, ticket, s) for s in target_splits):
            skipped.append(f"{ticket.display_id or t_id[:8]}: not your ticket")
            continue

        completion_note = (form.get(f"completion_note_{t_id}") or "").strip()
        cur_stage = target_splits[0].current_stage
        if cur_stage and cur_stage.completion_note_required and not completion_note:
            skipped.append(f"{ticket.display_id or t_id[:8]}: completion note required")
            continue

        # Collect and validate per-ticket custom field values from bulk modal
        import json as _json
        custom_fields_data: dict = {}
        if cur_stage and cur_stage.custom_fields_json:
            try:
                field_defs = _json.loads(cur_stage.custom_fields_json)
            except Exception:
                field_defs = []

            missing_required = []
            for fdef in field_defs:
                fid = fdef.get("id", "")
                if fdef.get("field_type") == "formula":
                    continue
                # Values are submitted as bulk_cf__{fid}__{ticket_id}
                val = str(form.get(f"bulk_cf__{fid}__{t_id}", "") or "").strip()
                if fdef.get("required") and not val:
                    missing_required.append(fdef.get("label", fid))
                if val:
                    custom_fields_data[fid] = val

            if missing_required:
                skipped.append(
                    f"{ticket.display_id or t_id[:8]}: required field(s) not filled: "
                    f"{', '.join(missing_required)}"
                )
                continue

            # Second pass — evaluate formula columns. Merge in values captured
            # at earlier stages so cross-stage formula references resolve.
            all_flow_stages = db.query(FMSStage).filter(
                FMSStage.flow_id == ticket.flow_id, FMSStage.is_deleted == False
            ).all()
            _bulk_open_h = _open_history(db, t_id, split_id=target_splits[0].id)
            _bulk_tff = {}
            if ticket.ticket_custom_fields_json:
                try:
                    _bulk_tff = _json.loads(ticket.ticket_custom_fields_json)
                except Exception:
                    pass
            formula_lookup = {
                **_bulk_tff,
                **_cross_stage_cf(db, t_id, all_flow_stages, split_id=_split_lineage_ids(db, t_id, target_splits[0].id), exclude_history_id=_bulk_open_h.id if _bulk_open_h else None),
                **custom_fields_data,
            }

            def _eval_formula_bulk(steps: list) -> str | None:
                result = None
                for i, step in enumerate(steps):
                    raw = formula_lookup.get(step.get("col_id", ""), "")
                    try:
                        val = float(raw)
                    except (ValueError, TypeError):
                        return None
                    if i == 0:
                        result = val; continue
                    op = step.get("op", "+")
                    if op == "+":   result += val
                    elif op == "-": result -= val
                    elif op == "*": result *= val
                    elif op == "/":
                        if val == 0: return None
                        result /= val
                if result is None: return None
                return str(int(result)) if result == int(result) else f"{result:.4f}".rstrip("0")

            for fdef in field_defs:
                if fdef.get("field_type") != "formula": continue
                computed = _eval_formula_bulk(fdef.get("formula_steps") or [])
                if computed is not None:
                    custom_fields_data[fdef.get("id", "")] = computed

        # Enforce the flow's closing rule before letting a ticket land on the
        # terminal stage — aggregate across every active split (see
        # _ticket_closing_rule_check), not just target_splits alone.
        if next_stage.is_terminal and ticket.flow and ticket.flow.closing_rule_json:
            try:
                rule = _json.loads(ticket.flow.closing_rule_json)
            except Exception:
                rule = None
            if rule and rule.get("col_id"):
                bulk_lookup = {**(formula_lookup if cur_stage and cur_stage.custom_fields_json else {}), **custom_fields_data}
                stages_for_rule = db.query(FMSStage).filter(
                    FMSStage.flow_id == ticket.flow_id, FMSStage.is_deleted == False).all()
                ok, _err = _ticket_closing_rule_check(
                    db, ticket, stages_for_rule, rule,
                    in_progress_split_id=[s.id for s in target_splits],
                    in_progress_values=bulk_lookup)
                if not ok:
                    skipped.append(f"{ticket.display_id or t_id[:8]}: closing rule not met")
                    continue

        multi = len(target_splits) > 1 or len(_active_splits(db, t_id)) > 1
        moved_assignees = set()
        for tsplit in target_splits:
            open_h = _open_history(db, t_id, split_id=tsplit.id)
            if open_h:
                open_h.exited_at = now
                open_h.completion_note = completion_note or None
                open_h.custom_fields_data_json = _json.dumps(custom_fields_data) if custom_fields_data else None
                suffix = f" [{tsplit.split_label}]" if multi else ""
                _log(db, t_id, user.id, "STAGE_EXITED", f"From: {cur_stage.name if cur_stage else '?'}{suffix}")

            new_assignee_id = _stage_default_assignee(next_stage) or tsplit.current_assignee_id
            db.add(FMSStageHistory(
                ticket_id=t_id, split_id=tsplit.id, stage_id=next_stage_id,
                stage_name=next_stage.name, assignee_id=new_assignee_id,
                direction="FORWARD",
            ))
            tsplit.current_stage_id = next_stage_id
            tsplit.current_assignee_id = new_assignee_id
            tsplit.updated_at = now
            tsplit.status = "COMPLETED" if next_stage.is_terminal else "ACTIVE"
            suffix = f" [{tsplit.split_label}]" if multi else ""
            _log(db, t_id, user.id, "STAGE_ENTERED", f"To: {next_stage.name}{suffix}")
            if new_assignee_id:
                moved_assignees.add(new_assignee_id)

        ticket.updated_at = now
        _sync_ticket_cache(db, ticket)
        _mark_completed_by(ticket, user.id)
        _check_qty_discrepancy(db, ticket, user.id)
        _notify_linked_parent_if_ready(db, ticket)

        # Notify assignees of the moved split(s)
        for new_assignee_id in moved_assignees:
            mgrs = _manager_ids_for(db, new_assignee_id)
            notify_fms_stage_transition(
                db, tid, t_id, ticket.title, next_stage.name, user.id,
                admins, mgrs, new_assignee_id,
            )
        moved += 1

    db.commit()

    from urllib.parse import quote as _q
    if skipped:
        msg = _q(f"{moved} moved; {len(skipped)} skipped: {'; '.join(skipped[:3])}")
        return _redirect(f"/fms/dashboard?view=stage&flow_id={flow_id}&stage_id={next_stage_id}&msg={msg}")
    msg = _q(f"{moved} ticket{'s' if moved != 1 else ''} moved to {next_stage.name}")
    return _redirect(f"/fms/dashboard?view=stage&flow_id={flow_id}&stage_id={next_stage_id}&msg={msg}")


@router.get("/tickets/{ticket_id}", response_class=HTMLResponse)
def fms_ticket_detail(
    ticket_id: str, request: Request,
    user: User = Depends(get_current_user_or_redirect),
    db: Session = Depends(get_db),
):
    # Phase A2: detail page removed — redirect to Stage view
    ticket = _get_ticket(db, ticket_id, user.tenant_id)
    return _redirect(
        f"/fms/dashboard?view=stage"
        f"{'&flow_id=' + ticket.flow_id if ticket.flow_id else ''}"
        f"{'&stage_id=' + ticket.current_stage_id if ticket.current_stage_id else ''}"
    )
    # --- legacy detail page below (unreachable, kept for reference) ---
    ticket = _get_ticket(db, ticket_id, user.tenant_id)  # noqa: F841
    flow   = ticket.flow
    now    = datetime.utcnow()
    stages = sorted(
        [s for s in flow.stages if not s.is_deleted], key=lambda s: s.order
    ) if flow else []

    # All history rows for this ticket, oldest first
    all_histories = db.query(FMSStageHistory).filter(
        FMSStageHistory.ticket_id == ticket_id,
    ).order_by(FMSStageHistory.entered_at).all()

    # All events for this ticket, oldest first
    all_events = db.query(FMSEvent).filter(
        FMSEvent.ticket_id == ticket_id,
    ).order_by(FMSEvent.created_at).all()

    # Current open history row
    open_h  = next((h for h in all_histories if h.exited_at is None), None)
    tat_pct = _tat_pct(open_h, ticket.current_stage) if open_h and ticket.current_stage else None

    cur_order = ticket.current_stage.order if ticket.current_stage else -1

    # ── Build per-stage accordion data ───────────────────────────────────────
    def _events_in_window(entered_at, exited_at):
        """Events whose created_at falls within a history row's time window."""
        result = []
        for ev in all_events:
            if ev.created_at >= entered_at:
                if exited_at is None or ev.created_at <= exited_at:
                    result.append(ev)
        return result

    stage_panels = []
    for s in stages:
        is_current = (s.id == ticket.current_stage_id)
        is_done    = (s.order < cur_order)
        is_future  = not is_current and not is_done

        visits = [h for h in all_histories if h.stage_id == s.id]

        # Total time spent across all completed visits
        total_secs = sum(
            (h.exited_at - h.entered_at).total_seconds()
            for h in visits if h.exited_at
        )
        # Time on current open visit (if this is the active stage)
        current_secs = None
        if is_current and open_h and open_h.stage_id == s.id:
            current_secs = (now - open_h.entered_at).total_seconds()

        # TaT status for this stage
        if is_current and tat_pct is not None:
            s_tat_pct   = tat_pct
            s_tat_color = "green" if tat_pct < 50 else "amber" if tat_pct < 90 else "red"
        else:
            s_tat_pct, s_tat_color = None, "gray"

        # Collect events per visit window
        enriched_visits = []
        for h in visits:
            win_events = _events_in_window(h.entered_at, h.exited_at)
            enriched_visits.append({
                "history": h,
                "events":  win_events,
                "duration_h": round(
                    (h.exited_at - h.entered_at).total_seconds() / 3600, 1
                ) if h.exited_at else None,
                "is_open": h.exited_at is None,
            })

        # Unique assignees seen at this stage
        assignee_ids = list(dict.fromkeys(
            h.assignee_id for h in visits if h.assignee_id
        ))
        assignee_names = []
        for aid in assignee_ids:
            u = db.query(User).get(aid)
            if u:
                assignee_names.append(u.name)

        import json as _json
        try:
            stage_custom_fields = _json.loads(s.custom_fields_json or "[]")
        except Exception:
            stage_custom_fields = []

        stage_panels.append({
            "stage":          s,
            "is_current":     is_current,
            "is_done":        is_done,
            "is_future":      is_future,
            "total_visits":   len(visits),
            "total_hours":    round(total_secs / 3600, 1),
            "current_secs":   current_secs,
            "tat_pct":        s_tat_pct,
            "tat_color":      s_tat_color,
            "enriched_visits": enriched_visits,
            "assignee_names": assignee_names,
            "qty_done":       _stage_cumulative_qty(db, ticket_id, s.id),
            "sub_module_tag": getattr(s, "sub_module_tag", None),
            "custom_fields":  stage_custom_fields,
        })

    # Manager override window still open?
    can_override = False
    if user.role in ("ADMIN", "MANAGER"):
        last_exit = next(
            (h for h in reversed(all_histories) if h.exited_at), None
        )
        if last_exit and last_exit.exited_at:
            age_h = (now - last_exit.exited_at).total_seconds() / 3600
            can_override = age_h <= MANAGER_OVERRIDE_HOURS

    employees = db.query(User).filter(
        User.tenant_id == user.tenant_id, User.is_deleted == False,
        User.is_active == True).all()
    helper_ids = [h.user_id for h in ticket.helpers]

    can_transition = _can_transition(user, ticket)
    can_manage     = user.role in ("ADMIN", "MANAGER")

    # Custom field definitions for the current active stage (for transition form)
    import json as _json
    current_stage_custom_fields = []
    if ticket.current_stage:
        try:
            current_stage_custom_fields = _json.loads(ticket.current_stage.custom_fields_json or "[]")
        except Exception:
            current_stage_custom_fields = []

    # Parse stage pre-assignments for the transition form auto-fill
    stage_assignees: dict = {}
    try:
        stage_assignees = _json.loads(ticket.stage_assignees_json or "{}")
    except Exception:
        stage_assignees = {}

    # Parse stage schedule for display in stage panels
    stage_schedule: dict = {}
    try:
        stage_schedule = _json.loads(ticket.stage_schedule_json or "{}")
    except Exception:
        stage_schedule = {}

    from .linked_entities import get_linked_entity_options as _geo
    from .database import LinkedEntityReference as _LER
    entity_options = _geo(db, user.tenant_id)
    linked_refs = db.query(_LER).filter(
        _LER.tenant_id == user.tenant_id,
        _LER.parent_type == "FMS_TICKET",
        _LER.parent_id == ticket_id,
    ).order_by(_LER.created_at).all()

    # Entity data for custom field entity_link dropdowns — Admin only
    entity_data = {}
    if user.role == "ADMIN":
        entity_data = {
            "customer": [{"id": c.id, "name": c.name} for c in
                         db.query(Customer).filter(Customer.tenant_id == user.tenant_id,
                                                   Customer.is_deleted == False).order_by(Customer.name).all()],
            "vendor": [{"id": v.id, "name": v.name} for v in
                       db.query(Vendor).filter(Vendor.tenant_id == user.tenant_id,
                                               Vendor.is_deleted == False).order_by(Vendor.name).all()],
            "raw_material": [{"id": r.id, "name": r.name} for r in
                             db.query(RawMaterial).filter(RawMaterial.tenant_id == user.tenant_id,
                                                          RawMaterial.is_deleted == False).order_by(RawMaterial.name).all()],
            "employee": [{"id": e.id, "name": e.name} for e in employees],
        }

    knowledge_items = db.query(KnowledgeItem).filter(
        KnowledgeItem.tenant_id == user.tenant_id, KnowledgeItem.is_deleted == False,
    ).order_by(KnowledgeItem.title).all()
    closing_media = db.query(MediaUpload).filter(
        MediaUpload.tenant_id == user.tenant_id,
        MediaUpload.entity_type == "fms_ticket",
        MediaUpload.entity_id == ticket_id,
    ).order_by(MediaUpload.created_at).all()

    return templates.TemplateResponse(request, "fms/ticket_detail.html", _ctx(
        request, user, db,
        ticket=ticket, flow=flow, stages=stages,
        stage_panels=stage_panels,
        open_h=open_h, tat_pct=tat_pct,
        can_override=can_override,
        employees=employees, helper_ids=helper_ids,
        can_transition=can_transition, can_manage=can_manage,
        now=now,
        entity_options=entity_options,
        linked_refs=linked_refs,
        entity_data=entity_data,
        current_stage_custom_fields=current_stage_custom_fields,
        stage_assignees=stage_assignees,
        stage_schedule=stage_schedule,
        all_events=list(reversed(all_events)),
        knowledge_items=knowledge_items,
        closing_media=closing_media,
    ))


@router.post("/tickets/{ticket_id}/transition")
async def fms_transition(
    request: Request,
    ticket_id: str,
    next_stage_id: str = Form(""),
    new_assignee_id: str = Form(""),
    completion_note: str = Form(""),
    qty_completed: str = Form("0"),
    return_reason: str = Form(""),
    is_override: bool = Form(False),
    split_id: str = Form(""),
    evidence_file: UploadFile = File(None),
    excluded_from_perf: bool = Form(False),
    user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """
    2-C-2/3/4/5/6/7: Stage transition engine.
    Handles FORWARD, BACKWARD, non-linear revisits, and manager override.

    Phase 0 (split flows): operates on a *split*, not the ticket directly.
    split_id is optional — when omitted (every existing caller/template),
    resolves to the ticket's sole/primary split via _ensure_ticket_has_split,
    so single-split tickets (the overwhelming majority) behave identically
    to before.
    """
    ticket = _get_ticket(db, ticket_id, user.tenant_id)
    if split_id:
        split = db.query(FMSTicketSplit).filter(
            FMSTicketSplit.id == split_id,
            FMSTicketSplit.ticket_id == ticket_id,
            FMSTicketSplit.is_deleted == False,
        ).first()
        if not split:
            raise HTTPException(404, "Split not found")
    else:
        split = _ensure_ticket_has_split(db, ticket)

    _can_via_whitelist = (
        user.role == "EMPLOYEE" and ticket.flow and ticket.flow.restrict_to_assignee and
        _can_act_on_ticket(user, ticket, split)
    )
    if not (_can_transition(user, ticket, split) or _can_via_whitelist):
        raise HTTPException(403, "Not authorised to transition this ticket")
    # CLOSED is a deliberate, permanent administrative closure — never reopenable
    # here. COMPLETED (reached the flow's terminal stage) is different: a
    # manager/admin override can reopen it to correct a mistake or handle a
    # rejection discovered after close (brief: "moving the splits back" must
    # work even once a split has completed) — everyone else still gets the
    # old hard block.
    if ticket.status == "CLOSED" or split.status == "CLOSED":
        raise HTTPException(400, "This ticket/split is closed")
    if split.status == "COMPLETED" and not (is_override and user.role in ("ADMIN", "MANAGER")):
        raise HTTPException(400,
            "This ticket/split is already completed — a manager/admin override is required to reopen it")

    # Empty next_stage_id means "complete the current terminal stage"
    terminal_complete = not next_stage_id.strip()
    next_stage = None
    if not terminal_complete:
        next_stage = db.query(FMSStage).filter(
            FMSStage.id == next_stage_id,
            FMSStage.flow_id == ticket.flow_id).first()
        if not next_stage:
            raise HTTPException(400, "Invalid next stage")

        new_assignee_id = (new_assignee_id or "").strip()
        if not new_assignee_id:
            new_assignee_id = _stage_default_assignee(next_stage) or ""
        if not new_assignee_id:
            raise HTTPException(400, "Please select an assignee for the next stage")

    cur_stage  = split.current_stage
    open_h     = _open_history(db, ticket_id, split_id=split.id)
    split_label_suffix = ""
    if len(_active_splits(db, ticket_id)) > 1:
        split_label_suffix = f" [{split.split_label}]"

    if terminal_complete:
        direction = "FORWARD"
        next_order = (cur_stage.order if cur_stage else 0)
    else:
        # Determine direction (2-C-3/4)
        cur_order  = cur_stage.order  if cur_stage  else 0
        next_order = next_stage.order
        direction  = "BACKWARD" if next_order < cur_order else "FORWARD"

    # A1-5: Enforce linear stage movement (skip for terminal-complete).
    # BACKWARD moves are allowed to any earlier stage in the flow (not just the
    # adjacent one) — gated below by a mandatory return_reason instead of the
    # one-step-only restriction, which only applies to FORWARD movement.
    if not is_override and not terminal_complete:
        if direction == "FORWARD" and next_order != cur_order + 1:
            raise HTTPException(400, "Tickets can only move to the next stage in sequence")

    if is_override:
        # 2-C-7: manager override
        if user.role not in ("ADMIN", "MANAGER"):
            raise HTTPException(403, "Only managers/admins can override")
        direction = "MANAGER_OVERRIDE"

    # 2-C-4: backward requires a valid (non-trivial) reason — required regardless
    # of how many stages the ticket is being moved back.
    if direction == "BACKWARD" and len(return_reason.strip()) < 5:
        raise HTTPException(400, "A valid return reason (at least 5 characters) is required to move a ticket back")

    # Stage requires completion note — only enforced when actually completing
    # the stage's work (FORWARD/terminal). A BACKWARD return or manager
    # override isn't "finishing" the stage, so it's exempt (same rule already
    # applied to required custom fields below).
    if (cur_stage and cur_stage.completion_note_required and not completion_note.strip()
            and direction not in ("BACKWARD", "MANAGER_OVERRIDE")):
        raise HTTPException(400, f"Stage '{cur_stage.name}' requires a completion note")

    # Stage requires evidence upload — same BACKWARD/override exemption as above.
    evidence_url = None
    evidence_filename = None
    if (cur_stage and getattr(cur_stage, "evidence_required", False)
            and direction not in ("BACKWARD", "MANAGER_OVERRIDE")):
        has_file = (evidence_file is not None
                    and evidence_file.filename
                    and evidence_file.filename.strip())
        if not has_file:
            raise HTTPException(
                400,
                f"Stage '{cur_stage.name}' requires an evidence file upload before moving on"
            )
        from .uploads import save_upload as _save_upload
        result = await _save_upload(evidence_file, user.tenant_id)
        evidence_url = result["file_path"]
        evidence_filename = result["file_name"]

    qty = int(qty_completed) if qty_completed.strip().isdigit() else 0

    # Collect custom field values for current stage (A4 + Phase B)
    import json as _json
    custom_fields_data = {}
    if cur_stage:
        try:
            field_defs = _json.loads(cur_stage.custom_fields_json or "[]")
        except Exception:
            field_defs = []
        form_data = await request.form()

        # First pass — collect all non-formula values keyed by id
        missing_required = []
        for fdef in field_defs:
            fid = fdef.get("id", "")
            if fdef.get("field_type") == "formula":
                continue  # evaluated in second pass
            key = f"cf__{fid}"
            val = str(form_data.get(key, "") or "").strip()
            # Required-field enforcement only applies on FORWARD moves — a
            # BACKWARD return or manager override isn't "finishing" the stage.
            if fdef.get("required") and not val and direction not in ("BACKWARD", "MANAGER_OVERRIDE"):
                missing_required.append(fdef.get("label", fid))
            if val:
                custom_fields_data[fid] = val
        if missing_required:
            raise HTTPException(400, f"Required column(s) not filled: {', '.join(missing_required)}")

        # Second pass — evaluate formula columns server-side.
        # Formulas may reference columns captured in earlier stages (the
        # formula builder UI allows this), so merge in cross-stage values —
        # current-stage values take precedence on id collisions.
        all_flow_stages = db.query(FMSStage).filter(
            FMSStage.flow_id == ticket.flow_id, FMSStage.is_deleted == False
        ).all()
        _ticket_tff = {}
        if ticket.ticket_custom_fields_json:
            try:
                _ticket_tff = _json.loads(ticket.ticket_custom_fields_json)
            except Exception:
                pass
        formula_lookup = {
            **_ticket_tff,
            **_cross_stage_cf(db, ticket_id, all_flow_stages, split_id=_split_lineage_ids(db, ticket_id, split.id), exclude_history_id=open_h.id if open_h else None),
            **custom_fields_data,
        }

        def _eval_formula(steps: list) -> str | None:
            result = None
            for i, step in enumerate(steps):
                col_id = step.get("col_id", "")
                raw = formula_lookup.get(col_id, "")
                try:
                    val = float(raw)
                except (ValueError, TypeError):
                    return None  # referenced column missing or non-numeric
                if i == 0:
                    result = val
                    continue
                op = step.get("op", "+")
                if op == "+":   result += val
                elif op == "-": result -= val
                elif op == "*": result *= val
                elif op == "/":
                    if val == 0:
                        return None
                    result /= val
            if result is None:
                return None
            # Return as integer string if it's a whole number, else up to 4 dp
            return str(int(result)) if result == int(result) else f"{result:.4f}".rstrip("0")

        for fdef in field_defs:
            if fdef.get("field_type") != "formula":
                continue
            fid = fdef.get("id", "")
            steps = fdef.get("formula_steps") or []
            computed = _eval_formula(steps)
            if computed is not None:
                custom_fields_data[fid] = computed

    # Close current stage history row
    if open_h:
        open_h.exited_at              = datetime.utcnow()
        open_h.completion_note        = completion_note.strip() or None
        open_h.qty_completed          = qty
        open_h.evidence_url           = evidence_url
        open_h.evidence_filename      = evidence_filename
        open_h.custom_fields_data_json = _json.dumps(custom_fields_data) if custom_fields_data else None
        # Whoever performs the return can flag it — employees return their own
        # tickets just as often as managers/admins do (see _can_transition),
        # so restricting this to ADMIN/MANAGER would mean the common case
        # (an employee returning their own ticket for an external reason)
        # could never be flagged at the time it happens. The mandatory
        # return_reason + audit log (below) keeps this reviewable.
        if direction in ("BACKWARD", "MANAGER_OVERRIDE") and excluded_from_perf:
            open_h.excluded_from_perf = True
        _log(db, ticket_id, user.id, "STAGE_EXITED",
             f"Stage: {cur_stage.name if cur_stage else '?'}{split_label_suffix} | "
             f"Entered: {open_h.entered_at.strftime('%d %b %H:%M') if open_h.entered_at else '—'} | "
             f"Qty completed: {qty} | Note: {completion_note[:80] or '—'} | "
             f"Evidence: {evidence_filename or '—'} | "
             f"Custom fields: {_fmt_cf(custom_fields_data)}"
             + (" | Excluded from performance scoring" if getattr(open_h, "excluded_from_perf", False) else ""),
             meta={
                 "stage_name": cur_stage.name if cur_stage else None,
                 "qty": qty, "note": completion_note.strip() or None,
                 "evidence_filename": evidence_filename,
                 "custom_fields": _cf_by_label(custom_fields_data, field_defs if cur_stage else []),
                 "excluded_from_perf": bool(getattr(open_h, "excluded_from_perf", False)),
             })

    ticket.updated_at = datetime.utcnow()
    split.updated_at = datetime.utcnow()

    if terminal_complete:
        # Enforce the flow's closing rule (e.g. "excess quantity = 0") before
        # allowing the ticket to close — evaluated in aggregate across every
        # active split (brief follow-up: the rule gates the ticket as a
        # whole, not just whichever split is completing right now).
        flow_for_rule = ticket.flow
        if flow_for_rule and flow_for_rule.closing_rule_json:
            try:
                rule = _json.loads(flow_for_rule.closing_rule_json)
            except Exception:
                rule = None
            if rule and rule.get("col_id"):
                in_progress = {**formula_lookup, **custom_fields_data} if cur_stage else custom_fields_data
                stages_for_rule = db.query(FMSStage).filter(
                    FMSStage.flow_id == ticket.flow_id, FMSStage.is_deleted == False).all()
                ok, err = _ticket_closing_rule_check(
                    db, ticket, stages_for_rule, rule,
                    in_progress_split_id=split.id, in_progress_values=in_progress)
                if not ok:
                    raise HTTPException(400, err)

        # Completing the current terminal stage — no new history row needed
        split.status = "COMPLETED"
        _sync_ticket_cache(db, ticket)
        _mark_completed_by(ticket, user.id)
        _check_qty_discrepancy(db, ticket, user.id)
        _log(db, ticket_id, user.id, "COMPLETED",
             f"Completed terminal stage: {cur_stage.name if cur_stage else '?'}{split_label_suffix} | "
             f"Qty: {qty} | Custom fields: {_fmt_cf(custom_fields_data)}", meta={
                 "stage_name": cur_stage.name if cur_stage else None, "qty": qty,
                 "custom_fields": _cf_by_label(custom_fields_data, field_defs if cur_stage else []),
             })
        db.commit()
        admins = _admin_ids(db, user.tenant_id)
        broadcast_sync(user.tenant_id, admins, FMS_STAGE_TRANSITION, {
            "ticket_id": ticket_id, "display_id": ticket.display_id,
            "title": ticket.title, "stage": cur_stage.name if cur_stage else "",
            "status": ticket.status,
        })
        return _redirect(
            f"/fms/dashboard?view=stage"
            f"{'&flow_id=' + ticket.flow_id if ticket.flow_id else ''}"
        )

    # Look up planned dates for next stage from ticket schedule.
    # BACKWARD moves reset the TAT clock: the target stage and every stage
    # after it (in flow order) get a fresh planned_start/planned_end chain
    # starting now, based on each stage's own target_tat_hours — the old
    # schedule (computed under the original timeline) no longer applies once
    # the ticket has been sent back.
    import json as _json2
    _sched: dict = {}
    try:
        _sched = _json2.loads(ticket.stage_schedule_json or "{}")
    except Exception:
        _sched = {}

    if direction == "BACKWARD":
        from .notifications import add_business_hours
        _tenant_sched = db.query(Tenant).get(user.tenant_id)
        _flow_stages_sched = sorted(
            [s for s in ticket.flow.stages if not s.is_deleted], key=lambda s: s.order
        )
        _cursor = datetime.utcnow()
        _reached = False
        for _fs in _flow_stages_sched:
            if _fs.id == next_stage_id:
                _reached = True
            if not _reached:
                continue
            _tat_h = _fs.target_tat_hours or 24
            _p_end = add_business_hours(_tenant_sched, _cursor, _tat_h)
            _sched[_fs.id] = {"planned_start": _cursor.isoformat(), "planned_end": _p_end.isoformat()}
            _cursor = _p_end
        ticket.stage_schedule_json = _json2.dumps(_sched)

    _ns = _sched.get(next_stage_id, {})
    _nps = datetime.fromisoformat(_ns["planned_start"]) if _ns.get("planned_start") else None
    _npe = datetime.fromisoformat(_ns["planned_end"])   if _ns.get("planned_end")   else None

    # FMS Auto-Split Engine (R1-R6) — evaluated inline, only for a genuine
    # FORWARD stage completion (not BACKWARD/MANAGER_OVERRIDE, which aren't
    # "finishing" the stage's work). Reassigns `split` to the moved-forward
    # split when a shortfall triggers a split; otherwise `split` is unchanged
    # and everything below behaves exactly as it did before this engine.
    if cur_stage is not None and direction == "FORWARD":
        _split_lookup = dict(formula_lookup) if formula_lookup else {}
        if ticket.ticket_custom_fields_json:
            try:
                _split_lookup.update(_json2.loads(ticket.ticket_custom_fields_json))
            except Exception:
                pass
        _split_lookup["__target_qty__"] = ticket.target_qty
        split = _evaluate_auto_split(
            db, ticket, split, cur_stage, qty,
            custom_fields_data, _split_lookup,
            next_stage_id, new_assignee_id, user)
        split_label_suffix = f" [{split.split_label}]" if len(_active_splits(db, ticket_id)) > 1 else ""

    # Create new stage history row (2-C-5: non-linear — always new row)
    new_h = FMSStageHistory(
        ticket_id=ticket_id, split_id=split.id, stage_id=next_stage_id,
        stage_name=next_stage.name, assignee_id=new_assignee_id,
        direction=direction,
        return_reason=return_reason.strip() or None,
        from_stage_id=cur_stage.id if cur_stage else None,
        from_stage_name=cur_stage.name if cur_stage else None,
        planned_start=_nps,
        planned_end=_npe,
    )
    db.add(new_h)

    # Compute THIS split's formula columns for the stage it's ENTERING, not
    # just the one it's exiting. Formula computation used to be exit-time
    # only (or triggered by a manual Enter Data save) — a split arriving at
    # a new stage with formula columns showed them blank until someone
    # visited the stage and saved something, or the split moved on again.
    # Persist the computed values onto the new history row immediately so
    # they're correct the moment the split lands here, split-scoped like
    # every other formula evaluation in this route.
    if next_stage.custom_fields_json:
        try:
            _next_field_defs = _json.loads(next_stage.custom_fields_json)
        except Exception:
            _next_field_defs = []
        if any(fd.get("field_type") == "formula" for fd in _next_field_defs):
            db.flush()  # new_h needs an id to exclude itself from _cross_stage_cf
            _all_flow_stages_entry = db.query(FMSStage).filter(
                FMSStage.flow_id == ticket.flow_id, FMSStage.is_deleted == False
            ).all()
            _ticket_tff_entry = {}
            if ticket.ticket_custom_fields_json:
                try:
                    _ticket_tff_entry = _json.loads(ticket.ticket_custom_fields_json)
                except Exception:
                    pass
            _entry_formula_lookup = {
                **_ticket_tff_entry,
                **_cross_stage_cf(db, ticket_id, _all_flow_stages_entry, split_id=_split_lineage_ids(db, ticket_id, split.id), exclude_history_id=new_h.id),
            }

            def _eval_entry_formula(steps: list) -> str | None:
                result = None
                for i, step in enumerate(steps):
                    raw = _entry_formula_lookup.get(step.get("col_id", ""), "")
                    try:
                        val = float(raw)
                    except (ValueError, TypeError):
                        return None
                    if i == 0:
                        result = val
                        continue
                    op = step.get("op", "+")
                    if op == "+":   result += val
                    elif op == "-": result -= val
                    elif op == "*": result *= val
                    elif op == "/":
                        if val == 0:
                            return None
                        result /= val
                if result is None:
                    return None
                return str(int(result)) if result == int(result) else f"{result:.4f}".rstrip("0")

            _entry_computed = {}
            for fdef in _next_field_defs:
                if fdef.get("field_type") != "formula":
                    continue
                computed = _eval_entry_formula(fdef.get("formula_steps") or [])
                if computed is not None:
                    _entry_computed[fdef.get("id", "")] = computed
            if _entry_computed:
                new_h.custom_fields_data_json = _json.dumps(_entry_computed)

    # Update the split (the ticket-level cache is refreshed below via _sync_ticket_cache)
    split.current_stage_id    = next_stage_id
    split.current_assignee_id = new_assignee_id

    if next_stage.is_terminal:
        flow_for_rule = ticket.flow
        if flow_for_rule and flow_for_rule.closing_rule_json:
            try:
                rule = _json.loads(flow_for_rule.closing_rule_json)
            except Exception:
                rule = None
            if rule and rule.get("col_id"):
                in_progress = {**formula_lookup, **custom_fields_data} if cur_stage else custom_fields_data
                stages_for_rule = db.query(FMSStage).filter(
                    FMSStage.flow_id == ticket.flow_id, FMSStage.is_deleted == False).all()
                ok, err = _ticket_closing_rule_check(
                    db, ticket, stages_for_rule, rule,
                    in_progress_split_id=split.id, in_progress_values=in_progress)
                if not ok:
                    raise HTTPException(400, err)
        split.status = "COMPLETED"
        _log(db, ticket_id, user.id, "COMPLETED",
             f"Reached terminal stage: {next_stage.name}{split_label_suffix}",
             meta={"stage_name": next_stage.name})
    else:
        split.status = "ACTIVE"

    _sync_ticket_cache(db, ticket)
    _mark_completed_by(ticket, user.id)
    _check_qty_discrepancy(db, ticket, user.id)
    _notify_linked_parent_if_ready(db, ticket)

    event_type = "RETURNED" if direction == "BACKWARD" else (
        "MANAGER_OVERRIDE" if direction == "MANAGER_OVERRIDE" else "STAGE_ENTERED")
    new_assignee_obj = db.query(User).filter(User.id == new_assignee_id).first()
    detail_parts = [
        f"From: {cur_stage.name if cur_stage else '—'} → To: {next_stage.name}{split_label_suffix}",
        f"Assignee: {new_assignee_obj.name if new_assignee_obj else '—'}",
        f"TAT window: {_fmt_window(_nps, _npe)} (target {next_stage.target_tat_hours or '—'}h)",
    ]
    if direction == "BACKWARD":
        detail_parts.append(f"Stages skipped back: {cur_order - next_order}")
    if return_reason: detail_parts.append(f"Reason: {return_reason}")
    _log(db, ticket_id, user.id, event_type, " | ".join(detail_parts), meta={
        "stage_name": next_stage.name,
        "from_stage_name": cur_stage.name if cur_stage else None,
        "assignee_name": new_assignee_obj.name if new_assignee_obj else None,
        "tat_window": _fmt_window(_nps, _npe), "target_tat_hours": next_stage.target_tat_hours,
        "reason": return_reason.strip() or None,
    })

    db.commit()

    # WS broadcast + WhatsApp
    admins   = _admin_ids(db, user.tenant_id)
    managers = _manager_ids_for(db, new_assignee_id)
    notify_fms_stage_transition(
        db, user.tenant_id, ticket_id, f"{ticket.title}{split_label_suffix}",
        next_stage.name, user.id, admins, managers, new_assignee_id,
        backward=(direction == "BACKWARD"))
    # WhatsApp excluded from both forward and backward stage transitions per
    # client rules — no send_whatsapp_for_fms_stage_transition call here.
    if next_stage.is_terminal:
        send_whatsapp_for_fms_ticket_closed(db, user.tenant_id, ticket, admins, managers, user.name)
    audience = list(set(admins + managers + [new_assignee_id]))
    broadcast_sync(user.tenant_id, audience, FMS_STAGE_TRANSITION, {
        "ticket_id": ticket_id, "display_id": ticket.display_id,
        "title": ticket.title, "stage": next_stage.name,
        "status": ticket.status,
    })

    # Backward move: notify managers + admin + the new assignee (previously
    # excluded — a real gap, now fixed) with a 2-hour override window message
    if direction == "BACKWARD":
        mgr_ids = _manager_ids_for(db, new_assignee_id)
        for mid in set(mgr_ids + admins + [new_assignee_id]):
            if not mid:
                continue
            create_notification(
                db, user.tenant_id,
                user_id=mid,
                notif_type="FMS_BACKWARD_MOVE",
                title=f"Ticket returned: {ticket.title}",
                body=(f"{ticket.display_id or ticket_id} was returned to "
                      f"'{next_stage.name}'. Reason: {return_reason[:120]}. "
                      f"You can reverse this within 2 hours."),
                link=f"/fms/dashboard?view=stage&flow_id={ticket.flow_id}&stage_id={next_stage_id}",
            )
        db.commit()

    return _redirect(
        f"/fms/dashboard?view=stage"
        f"{'&flow_id=' + ticket.flow_id if ticket.flow_id else ''}"
        f"{'&stage_id=' + next_stage_id}"
    )


# ── Phase 0: Split Flows — the split action ─────────────────────────────────

@router.post("/tickets/{ticket_id}/splits/{split_id}/split")
async def fms_split_ticket(
    request: Request,
    ticket_id: str,
    split_id: str,
    qty_to_move: str = Form(""),
    target_stage_id: str = Form(...),
    new_assignee_id: str = Form(""),
    completion_note: str = Form(""),
    return_reason: str = Form(""),
    is_override: bool = Form(False),
    user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """
    Brief §5: carve part of a split's qty off into a new, independently
    trackable split at a (usually different) stage. Reuses the same
    linear-movement / backward-reason / manager-override / completion-note
    rules as fms_transition(), applied to the source split rather than the
    whole ticket.

    Follow-up (splitting "at the custom-column level"): a split isn't
    purely quantity-only — the source stage's own custom fields (cf__{id})
    can optionally be submitted alongside qty_to_move, using the same
    required/formula rules as a normal transition. Where they land depends
    on what happens to the source split: if it's fully consumed by this
    split (qty hits 0), the values describe what happened at its exit, so
    they're recorded there — exactly like a normal transition. If the
    source continues (the common case — only part of its qty moved), there
    is no exit event to attach to, so the values describe the new branch
    instead and are recorded on the new split's opening history row.
    """
    ticket = _get_ticket(db, ticket_id, user.tenant_id)
    source = db.query(FMSTicketSplit).filter(
        FMSTicketSplit.id == split_id,
        FMSTicketSplit.ticket_id == ticket_id,
        FMSTicketSplit.is_deleted == False,
    ).first()
    if not source:
        raise HTTPException(404, "Split not found")
    if not _can_transition(user, ticket, source):
        raise HTTPException(403, "Not authorised to split this ticket")
    # Same COMPLETED-is-reopenable-by-override / CLOSED-is-permanent rule as
    # fms_transition — a manual split off a completed split needs the same
    # escape hatch (e.g. a manager decides more of it needs rework after all).
    if ticket.status == "CLOSED" or source.status == "CLOSED":
        raise HTTPException(400, "This split is closed")
    if source.status == "COMPLETED" and not (is_override and user.role in ("ADMIN", "MANAGER")):
        raise HTTPException(400,
            "This split is already completed — a manager/admin override is required to reopen it")

    # Splitting isn't always a quantity concept — plenty of flows never set
    # target_qty at all (it's nullable on FMSTicket). When this ticket
    # doesn't track quantity, qty_to_move is ignored entirely: the split is
    # purely "carve a copy of this portion off to another stage/assignee",
    # qty stays null on both sides, and the source is never auto-retired
    # (there's no quantity signal to say it's "used up" — it keeps existing
    # independently until acted on directly).
    qty_tracked = source.qty is not None
    orig_qty = source.qty
    qty = None
    if qty_tracked:
        try:
            qty = int(qty_to_move)
        except (TypeError, ValueError):
            raise HTTPException(400, "qty_to_move must be a whole number")
        if qty <= 0 or qty > orig_qty:
            raise HTTPException(400, f"qty_to_move must be > 0 and no more than the split's remaining qty ({orig_qty})")

    target_stage = db.query(FMSStage).filter(
        FMSStage.id == target_stage_id, FMSStage.flow_id == ticket.flow_id).first()
    if not target_stage:
        raise HTTPException(400, "Invalid target stage")

    cur_stage = source.current_stage
    cur_order = cur_stage.order if cur_stage else 0
    next_order = target_stage.order
    direction = "BACKWARD" if next_order < cur_order else "FORWARD"

    if is_override:
        if user.role not in ("ADMIN", "MANAGER"):
            raise HTTPException(403, "Only managers/admins can override")
        direction = "MANAGER_OVERRIDE"
    else:
        if direction == "FORWARD" and next_order != cur_order + 1:
            raise HTTPException(400, "Splits can only move to the next stage in sequence")
        if direction == "BACKWARD" and next_order != cur_order - 1:
            raise HTTPException(400, "Splits can only move to the previous stage in sequence")
    if direction == "BACKWARD" and len(return_reason.strip()) < 5:
        raise HTTPException(400, "A valid return reason (at least 5 characters) is required to move a split back")
    if cur_stage and cur_stage.completion_note_required and not completion_note.strip():
        raise HTTPException(400, f"Stage '{cur_stage.name}' requires a completion note")

    open_h = _open_history(db, ticket_id, split_id=source.id)
    now = datetime.utcnow()

    # Optional custom-field capture for the source stage (same rules as
    # fms_transition: required fields only enforced on FORWARD moves,
    # formula columns evaluated server-side against this split's own
    # lineage so they can reference values captured before this split).
    custom_fields_data: dict = {}
    all_flow_stages = db.query(FMSStage).filter(
        FMSStage.flow_id == ticket.flow_id, FMSStage.is_deleted == False).all()
    if cur_stage:
        try:
            field_defs = _json.loads(cur_stage.custom_fields_json or "[]")
        except Exception:
            field_defs = []
        if field_defs:
            form_data = await request.form()
            missing_required = []
            for fdef in field_defs:
                fid = fdef.get("id", "")
                if fdef.get("field_type") == "formula":
                    continue
                val = str(form_data.get(f"cf__{fid}", "") or "").strip()
                if fdef.get("required") and not val and direction != "BACKWARD":
                    missing_required.append(fdef.get("label", fid))
                if val:
                    custom_fields_data[fid] = val
            if missing_required:
                raise HTTPException(400, f"Required column(s) not filled: {', '.join(missing_required)}")

            _ticket_tff = {}
            if ticket.ticket_custom_fields_json:
                try:
                    _ticket_tff = _json.loads(ticket.ticket_custom_fields_json)
                except Exception:
                    pass
            formula_lookup = {
                **_ticket_tff,
                **_cross_stage_cf(db, ticket_id, all_flow_stages, split_id=source.id,
                                   exclude_history_id=open_h.id if open_h else None),
                **custom_fields_data,
            }

            def _eval_formula_split(steps: list):
                result = None
                for i, step in enumerate(steps):
                    raw = formula_lookup.get(step.get("col_id", ""), "")
                    try:
                        val = float(raw)
                    except (ValueError, TypeError):
                        return None
                    if i == 0:
                        result = val
                        continue
                    op = step.get("op", "+")
                    if op == "+":   result += val
                    elif op == "-": result -= val
                    elif op == "*": result *= val
                    elif op == "/":
                        if val == 0: return None
                        result /= val
                if result is None:
                    return None
                return str(int(result)) if result == int(result) else f"{result:.4f}".rstrip("0")

            for fdef in field_defs:
                if fdef.get("field_type") != "formula":
                    continue
                computed = _eval_formula_split(fdef.get("formula_steps") or [])
                if computed is not None:
                    custom_fields_data[fdef.get("id", "")] = computed

    # 1. Decrement the source split; if fully consumed, its story ends here.
    # (Qty-less tickets: nothing to decrement, source never auto-retires.)
    if qty_tracked:
        remaining = orig_qty - qty
        source.qty = remaining
        source_retired = remaining <= 0
    else:
        remaining = None
        source_retired = False
    source.updated_at = now
    if source_retired:
        if open_h:
            open_h.exited_at = now
            open_h.completion_note = completion_note.strip() or open_h.completion_note
            # Custom-field values belong to the source's exit record when it
            # has one (it fully left this stage) — same as a normal transition.
            if custom_fields_data:
                open_h.custom_fields_data_json = _json.dumps(custom_fields_data)
        source.is_deleted = True
        source.status = "CLOSED"

    # 2. Create the new split + open its stage-history row.
    # Snapshot everything the source split had accumulated up to (and
    # including) this split action — this is what lets the new branch's
    # formula columns / closing-rule contribution keep working immediately,
    # without waiting for it to revisit every earlier stage itself. Taken
    # now (source's own exit/update above has already happened), and static
    # from here on — the source's future is its own, independent history.
    inherited_snapshot = _cross_stage_cf(db, ticket_id, all_flow_stages, split_id=source.id)
    new_label = _next_split_label(db, ticket_id)
    new_assignee = new_assignee_id or source.current_assignee_id
    # Lineage metadata (root_ticket_id/split_display_id/split_sequence/
    # split_stage_id) — a manual split is a sibling branch of `source` exactly
    # like an auto-split is, so it gets the same hierarchical id fields
    # (e.g. F-0042-1-2) rather than just a bare "S3" label. Without this a
    # manually-split branch's origin is invisible next to auto-split siblings
    # in the same tree (brief §6: "clearly defined - from which branch it
    # originated").
    base_display = source.split_display_id or ticket.display_id or ticket.id[:8]
    seq = _next_auto_split_sequence(db, ticket_id, source.id)
    new_split = FMSTicketSplit(
        tenant_id=ticket.tenant_id, ticket_id=ticket_id, parent_split_id=source.id,
        root_ticket_id=source.root_ticket_id or ticket_id,
        split_label=new_label, split_display_id=f"{base_display}-{seq}", split_sequence=seq,
        split_stage_id=cur_stage.id if cur_stage else None,
        qty=qty,
        current_stage_id=target_stage_id, current_assignee_id=new_assignee,
        status="COMPLETED" if target_stage.is_terminal else "ACTIVE",
    )
    db.add(new_split)
    db.flush()
    # Source continues (no exit event) — the newly captured values describe
    # this new branch, not a source update, so they're merged into the
    # inherited snapshot rather than left on the source.
    new_row_cf = {**inherited_snapshot, **(custom_fields_data if not source_retired else {})}

    # TAT window for the new split's opening history row — same schedule
    # logic as fms_transition (a manual split is still a stage move for the
    # portion that lands on target_stage). BACKWARD resets the clock for
    # target_stage and everything after it in flow order, mirroring the
    # ticket-level schedule fms_transition maintains; this split shares that
    # same ticket.stage_schedule_json cache.
    _sched: dict = {}
    try:
        _sched = _json.loads(ticket.stage_schedule_json or "{}")
    except Exception:
        _sched = {}
    if direction == "BACKWARD":
        from .notifications import add_business_hours
        _tenant_sched = db.query(Tenant).get(user.tenant_id)
        _flow_stages_sched = sorted(
            [fs for fs in ticket.flow.stages if not fs.is_deleted], key=lambda fs: fs.order
        )
        _cursor = now
        _reached = False
        for _fs in _flow_stages_sched:
            if _fs.id == target_stage_id:
                _reached = True
            if not _reached:
                continue
            _tat_h = _fs.target_tat_hours or 24
            _p_end = add_business_hours(_tenant_sched, _cursor, _tat_h)
            _sched[_fs.id] = {"planned_start": _cursor.isoformat(), "planned_end": _p_end.isoformat()}
            _cursor = _p_end
        ticket.stage_schedule_json = _json.dumps(_sched)
    _ns = _sched.get(target_stage_id, {})
    _nps = datetime.fromisoformat(_ns["planned_start"]) if _ns.get("planned_start") else None
    _npe = datetime.fromisoformat(_ns["planned_end"]) if _ns.get("planned_end") else None

    db.add(FMSStageHistory(
        ticket_id=ticket_id, split_id=new_split.id, stage_id=target_stage_id,
        stage_name=target_stage.name, assignee_id=new_assignee,
        direction=direction,
        return_reason=return_reason.strip() or None,
        from_stage_id=cur_stage.id if cur_stage else None,
        from_stage_name=cur_stage.name if cur_stage else None,
        planned_start=_nps,
        planned_end=_npe,
        custom_fields_data_json=_json.dumps(new_row_cf) if new_row_cf else None,
    ))

    # 2b. Closing rule (aggregate across the ticket's active splits — see
    # _ticket_closing_rule_check) if this split just landed on a terminal
    # stage. Checked here — after the new split exists so it's part of the
    # aggregate — rather than up front, since only NOW do we know its
    # contribution. Validation failure rolls back everything above.
    if target_stage.is_terminal and ticket.flow and ticket.flow.closing_rule_json:
        try:
            rule = _json.loads(ticket.flow.closing_rule_json)
        except Exception:
            rule = None
        if rule and rule.get("col_id"):
            # new_split's history row (and, if source_retired, source's exit
            # row) are already flushed above — _cross_stage_cf reads them via
            # the same session, and lineage-walks back through source, so no
            # separate in-progress override is needed here.
            ok, err = _ticket_closing_rule_check(db, ticket, all_flow_stages, rule)
            if not ok:
                db.rollback()
                raise HTTPException(400, err)

    # 3. Audit trail (brief §5: reuses the existing append-only event log).
    if qty_tracked:
        detail = (f"{source.split_label} ({orig_qty}) -> {new_label} ({qty}) @ {target_stage.name}. "
                  + (f"{source.split_label} fully consumed (retired)." if source_retired
                     else f"{source.split_label} continues at {cur_stage.name if cur_stage else '?'} with {remaining} remaining."))
    else:
        detail = (f"{source.split_label} -> {new_label} @ {target_stage.name} "
                  f"(no quantity tracked for this ticket). "
                  f"{source.split_label} continues at {cur_stage.name if cur_stage else '?'}.")
    _log(db, ticket_id, user.id, "SPLIT_CREATED", detail)

    _sync_ticket_cache(db, ticket)
    _check_qty_discrepancy(db, ticket, user.id)
    ticket.updated_at = now
    db.commit()

    admins = _admin_ids(db, user.tenant_id)
    managers = _manager_ids_for(db, new_assignee)
    notify_fms_stage_transition(
        db, user.tenant_id, ticket_id, f"{ticket.title} [{new_label}]",
        target_stage.name, user.id, admins, managers, new_assignee)
    new_assignee_obj = db.query(User).filter(User.id == new_assignee).first() if new_assignee else None
    # WhatsApp excluded from stage transitions per client rules.
    if target_stage.is_terminal:
        send_whatsapp_for_fms_ticket_closed(db, user.tenant_id, ticket, admins, managers, user.name)
    audience = list(set(admins + managers + ([new_assignee] if new_assignee else [])))
    broadcast_sync(user.tenant_id, audience, FMS_STAGE_TRANSITION, {
        "ticket_id": ticket_id, "display_id": ticket.display_id,
        "title": ticket.title, "stage": target_stage.name,
        "status": ticket.status,
    })

    return _redirect(
        f"/fms/dashboard?view=stage&flow_id={ticket.flow_id}&stage_id={target_stage_id}"
        f"&msg={new_label}+created"
    )


@router.post("/tickets/{ticket_id}/splits/merge")
def fms_merge_splits(
    ticket_id: str,
    split_ids: List[str] = Form(...),
    reason: str = Form(""),
    user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """Merge two or more of a ticket's active splits back into one — the
    inverse of 'Split This Ticket'. The split furthest along the flow
    (highest current_stage.order, ties broken by most recent updated_at)
    survives and absorbs the others' qty; the rest are retired
    (is_deleted=True, status=CLOSED) with their open history rows closed,
    mirroring how fms_split_ticket already retires a fully-carved-off split."""
    ticket = _get_ticket(db, ticket_id, user.tenant_id)
    if user.role not in ("ADMIN", "MANAGER"):
        raise HTTPException(403, "Managers only")
    if not reason.strip():
        raise HTTPException(400, "A reason is required to merge splits")

    splits = db.query(FMSTicketSplit).filter(
        FMSTicketSplit.id.in_(split_ids),
        FMSTicketSplit.ticket_id == ticket_id,
        FMSTicketSplit.is_deleted == False,
    ).all()
    if len(splits) < 2:
        raise HTTPException(400, "Select at least 2 active splits to merge")

    splits_sorted = sorted(
        splits,
        key=lambda s: ((s.current_stage.order if s.current_stage else -1), s.updated_at),
        reverse=True,
    )
    primary = splits_sorted[0]
    others = splits_sorted[1:]

    total_qty = sum((s.qty or 0) for s in splits)
    primary.qty = total_qty if any(s.qty is not None for s in splits) else None
    primary.updated_at = datetime.utcnow()

    merged_labels = [s.split_label for s in others]
    for s in others:
        open_h = _open_history(db, ticket_id, split_id=s.id)
        if open_h:
            open_h.exited_at = datetime.utcnow()
            open_h.completion_note = f"Merged into {primary.split_label}. {reason.strip()}"
        s.status = "CLOSED"
        s.is_deleted = True
        s.updated_at = datetime.utcnow()

    _sync_ticket_cache(db, ticket)
    _check_qty_discrepancy(db, ticket, user.id)
    ticket.updated_at = datetime.utcnow()
    stage_name = primary.current_stage.name if primary.current_stage else None
    _log(db, ticket_id, user.id, "SPLITS_MERGED",
         f"Merged {', '.join(merged_labels)} into {primary.split_label} at {stage_name or '?'} | "
         f"Combined qty: {total_qty} | Reason: {reason.strip()}",
         meta={"stage_name": stage_name, "qty": total_qty, "reason": reason.strip()})
    db.commit()

    return _redirect(
        f"/fms/dashboard?view=stage"
        f"{'&flow_id=' + ticket.flow_id if ticket.flow_id else ''}"
        f"{'&stage_id=' + ticket.current_stage_id if ticket.current_stage_id else ''}"
    )


@router.post("/tickets/{ticket_id}/discrepancy/acknowledge")
def fms_discrepancy_acknowledge(
    ticket_id: str,
    user: User = Depends(require_manager),
    db: Session = Depends(get_db),
):
    """Brief §5: manager/admin acknowledges a system-detected qty discrepancy —
    no requirement to 'fix' it, just to consciously note the drift is expected."""
    ticket = _get_ticket(db, ticket_id, user.tenant_id)
    ticket.has_qty_discrepancy = False
    ticket.updated_at = datetime.utcnow()
    _log(db, ticket_id, user.id, "QTY_DISCREPANCY_ACKNOWLEDGED",
         "Manager acknowledged quantity drift as expected/explained")
    db.commit()
    return _redirect(
        f"/fms/dashboard?view=stage"
        f"{'&flow_id=' + ticket.flow_id if ticket.flow_id else ''}"
        f"{'&stage_id=' + ticket.current_stage_id if ticket.current_stage_id else ''}"
    )


@router.post("/tickets/bulk-action")
async def fms_bulk_action(
    request: Request,
    user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """Bulk send-back, bulk close (optionally + continue), or bulk send-to-linked-flow
    for FMS tickets from list view."""
    if user.role not in ("ADMIN", "MANAGER"):
        raise HTTPException(403)
    form = await request.form()
    action = form.get("action", "")
    ids = form.getlist("ticket_ids")
    also_continue = form.get("also_continue", "") == "1"
    if not ids or action not in ("send_back", "close", "send_to_linked_flow"):
        return _redirect("/fms/dashboard?view=stage")

    tid = user.tenant_id
    tickets = db.query(FMSTicket).filter(
        FMSTicket.id.in_(ids), FMSTicket.tenant_id == tid,
        FMSTicket.is_deleted == False).all()

    skipped = []
    for t in tickets:
        if action == "close":
            if t.status not in ("COMPLETED", "CLOSED"):
                t.status = "CLOSED"
                t.updated_at = datetime.utcnow()
                _log(db, t.id, user.id, "CLOSED", "Bulk closed from list view")
                _notify_linked_parent_if_ready(db, t)
                if also_continue:
                    target_flow = _resolve_linked_flow(db, tid, t.flow.next_library_flow_id if t.flow else None)
                    if target_flow:
                        continuation = _spawn_linked_ticket(db, t, target_flow, user)
                        continuation.continued_from_ticket_id = t.id
                        t.continued_to_ticket_id = continuation.id
                        _log(db, t.id, user.id, "CONTINUED", f"Continued as {continuation.display_id} on '{target_flow.name}' (bulk)")
                    else:
                        skipped.append(f"{t.display_id or t.id[:8]}: no continuation flow configured")
        elif action == "send_to_linked_flow":
            if t.status in ("COMPLETED", "CLOSED", "ON_HOLD"):
                skipped.append(f"{t.display_id or t.id[:8]}: not active")
                continue
            stage = t.current_stage
            target_flow = _resolve_linked_flow(db, tid, stage.linked_library_flow_id if stage else None)
            if not target_flow:
                skipped.append(f"{t.display_id or t.id[:8]}: no linked flow configured on current stage")
                continue
            linked = _spawn_linked_ticket(db, t, target_flow, user,
                                           title=f"{t.title} — linked from {t.display_id}")
            linked.linked_parent_ticket_id = t.id
            t.status = "ON_HOLD"
            t.linked_child_ticket_id = linked.id
            t.pause_reason = f"Waiting on {linked.display_id} ({target_flow.name})"
            _log(db, t.id, user.id, "SENT_TO_LINKED_FLOW", f"Spawned {linked.display_id} on '{target_flow.name}' (bulk)")
        elif action == "send_back":
            # Find the last exited stage and send back to it
            prev_h = db.query(FMSStageHistory).filter(
                FMSStageHistory.ticket_id == t.id,
                FMSStageHistory.exited_at != None,
            ).order_by(FMSStageHistory.exited_at.desc()).first()
            if prev_h and prev_h.stage_id and t.status not in ("COMPLETED", "CLOSED"):
                # Close current open history row
                open_h = _open_history(db, t.id)
                if open_h:
                    open_h.exited_at = datetime.utcnow()
                # Open new history row for prev stage
                db.add(FMSStageHistory(
                    id=new_id(), ticket_id=t.id,
                    stage_id=prev_h.stage_id,
                    assignee_id=t.current_assignee_id,
                    entered_at=datetime.utcnow(),
                    direction="BACKWARD",
                    return_reason="Bulk send-back from list view",
                ))
                t.current_stage_id = prev_h.stage_id
                t.status = "ACTIVE"
                t.updated_at = datetime.utcnow()
                _log(db, t.id, user.id, "RETURNED",
                     f"Bulk send-back to {prev_h.stage.name if prev_h.stage else '?'}")

    db.commit()
    if skipped:
        from urllib.parse import quote as _q
        msg = _q(f"{len(tickets) - len(skipped)} done; {len(skipped)} skipped: {'; '.join(skipped[:3])}")
        return _redirect(f"/fms/dashboard?view=stage&msg={msg}")
    return _redirect("/fms/dashboard?view=stage")


@router.get("/tickets/{ticket_id}/knowledge-links")
def fms_ticket_knowledge_links(ticket_id: str,
                                user: User = Depends(get_current_user),
                                db: Session = Depends(get_db)):
    """Backs the Knowledge/Training modal — current links + all available items."""
    from fastapi.responses import JSONResponse
    ticket = _get_ticket(db, ticket_id, user.tenant_id)
    links = db.query(FMSTicketKnowledgeLink).filter(
        FMSTicketKnowledgeLink.ticket_id == ticket.id).all()
    available = db.query(KnowledgeItem).filter(
        KnowledgeItem.tenant_id == user.tenant_id, KnowledgeItem.is_deleted == False,
    ).order_by(KnowledgeItem.title).all()
    return JSONResponse({
        "links": [{"id": l.id, "knowledge_item_id": l.knowledge_item_id,
                   "title": l.knowledge_item.title if l.knowledge_item else "—"} for l in links],
        "available": [{"id": k.id, "title": k.title} for k in available],
    })

@router.post("/tickets/{ticket_id}/link-knowledge")
def fms_link_knowledge(ticket_id: str, knowledge_item_id: str = Form(...),
                        user: User = Depends(get_current_user),
                        db: Session = Depends(get_db)):
    """Link an FMS ticket to a Knowledge/Training item for quick reference."""
    from fastapi.responses import JSONResponse
    ticket = _get_ticket(db, ticket_id, user.tenant_id)
    item = db.query(KnowledgeItem).filter(
        KnowledgeItem.id == knowledge_item_id, KnowledgeItem.tenant_id == user.tenant_id,
        KnowledgeItem.is_deleted == False).first()
    if not item:
        raise HTTPException(404, "Knowledge item not found")
    already = db.query(FMSTicketKnowledgeLink).filter(
        FMSTicketKnowledgeLink.ticket_id == ticket_id,
        FMSTicketKnowledgeLink.knowledge_item_id == knowledge_item_id).first()
    if not already:
        db.add(FMSTicketKnowledgeLink(
            tenant_id=user.tenant_id, ticket_id=ticket_id,
            knowledge_item_id=knowledge_item_id, linked_by_id=user.id,
        ))
        _log(db, ticket_id, user.id, "KNOWLEDGE_LINKED", item.title)
        db.commit()
    return JSONResponse({"ok": True})

@router.post("/tickets/{ticket_id}/unlink-knowledge")
def fms_unlink_knowledge(ticket_id: str, link_id: str = Form(...),
                          user: User = Depends(require_manager),
                          db: Session = Depends(get_db)):
    from fastapi.responses import JSONResponse
    db.query(FMSTicketKnowledgeLink).filter(
        FMSTicketKnowledgeLink.id == link_id, FMSTicketKnowledgeLink.tenant_id == user.tenant_id,
        FMSTicketKnowledgeLink.ticket_id == ticket_id).delete()
    db.commit()
    return JSONResponse({"ok": True})


@router.post("/tickets/{ticket_id}/action")
async def fms_action(
    ticket_id: str,
    action: str = Form(...),
    comment: str = Form(""),
    reason: str = Form(""),
    new_assignee_id: str = Form(""),
    helper_id: str = Form(""),
    flag_reason: str = Form(""),
    split_id: str = Form(""),
    closing_file: UploadFile = File(None),
    user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """2-D: Reassign, help request, flag, comment, on-hold, close.

    Phase 0 (split flows): split_id is optional. Omitted (every existing
    caller) → today's exact ticket-level behavior, unchanged. Present (only
    reachable from the Splits modal on a multi-split ticket) → reassign/
    help-request/flag/mark-stage-complete apply to that split instead, then
    roll up to the ticket. comment/on_hold/resume/close/helpers stay
    ticket-wide administrative actions either way (brief §6 only calls out
    advance/return/flag/help/complete as per-split)."""
    ticket = _get_ticket(db, ticket_id, user.tenant_id)
    split = None
    if split_id:
        split = db.query(FMSTicketSplit).filter(
            FMSTicketSplit.id == split_id,
            FMSTicketSplit.ticket_id == ticket_id,
            FMSTicketSplit.is_deleted == False,
        ).first()
        if not split:
            raise HTTPException(404, "Split not found")

    if action not in ("add_helper", "remove_helper") and not _can_act_on_ticket(user, ticket, split):
        raise HTTPException(403, "Only the assigned employee for this stage can act on this ticket")

    if action == "comment" and comment.strip():
        _log(db, ticket_id, user.id, "COMMENT", comment.strip())

    elif action == "reassign" and new_assignee_id and reason.strip():
        # 2-D-1/2: reassign — mandatory handoff form
        if split:
            if user.role == "EMPLOYEE" and split.current_assignee_id != user.id:
                raise HTTPException(403, "Only the current assignee can reassign")
            old_assignee = split.current_assignee_id
            split.current_assignee_id = new_assignee_id
            split.updated_at = datetime.utcnow()
            open_h = _open_history(db, ticket_id, split_id=split.id)
            if open_h:
                open_h.assignee_id = new_assignee_id
            _sync_ticket_cache(db, ticket)
            _log(db, ticket_id, user.id, "REASSIGNED",
                 f"[{split.split_label}] From: {old_assignee} → To: {new_assignee_id} | {reason}")
        else:
            if user.role == "EMPLOYEE" and ticket.current_assignee_id != user.id:
                raise HTTPException(403, "Only the current assignee can reassign")
            old_assignee = ticket.current_assignee_id
            ticket.current_assignee_id = new_assignee_id
            ticket.updated_at = datetime.utcnow()
            # Update open stage history assignee
            open_h = _open_history(db, ticket_id)
            if open_h:
                open_h.assignee_id = new_assignee_id
            _log(db, ticket_id, user.id, "REASSIGNED",
                 f"From: {old_assignee} → To: {new_assignee_id} | {reason}")

    elif action == "help_request" and comment.strip():
        # 2-D-3
        if split:
            split.status = "HELP_REQUESTED"
            split.updated_at = datetime.utcnow()
            _sync_ticket_cache(db, ticket)
            _log(db, ticket_id, user.id, "HELP_REQUESTED", f"[{split.split_label}] {comment.strip()}")
        else:
            ticket.status = "HELP_REQUESTED"
            _log(db, ticket_id, user.id, "HELP_REQUESTED", comment.strip())

    elif action == "add_helper" and helper_id:
        # 2-D-3
        if user.role not in ("ADMIN", "MANAGER"):
            raise HTTPException(403, "Managers only")
        existing = db.query(FMSTicketHelper).filter(
            FMSTicketHelper.ticket_id == ticket_id,
            FMSTicketHelper.user_id == helper_id).first()
        if not existing:
            db.add(FMSTicketHelper(
                ticket_id=ticket_id, user_id=helper_id,
                added_by_id=user.id, reason=reason.strip() or None))
            _log(db, ticket_id, user.id, "HELPER_ADDED",
                 f"Helper: {helper_id} | {reason}")

    elif action == "remove_helper" and helper_id:
        if user.role not in ("ADMIN", "MANAGER"):
            raise HTTPException(403, "Managers only")
        db.query(FMSTicketHelper).filter(
            FMSTicketHelper.ticket_id == ticket_id,
            FMSTicketHelper.user_id == helper_id).delete()
        _log(db, ticket_id, user.id, "HELPER_REMOVED", f"Helper: {helper_id}")

    elif action == "flag" and flag_reason.strip():
        # 2-D-4. FMSTicketSplit has no separate flagged_reason column (brief §3
        # schema) — a split flag sets its status and mirrors is_flagged onto
        # the ticket so the existing 🚩 badge keeps working unchanged.
        # Any user may flag; an EMPLOYEE may only flag their own assignment
        # (mirrors regular Tickets/Checklists).
        if user.role == "EMPLOYEE":
            current_assignee = (split.assignee_id if split else ticket.current_assignee_id)
            if current_assignee != user.id:
                raise HTTPException(status_code=403, detail="Only the current assignee can flag this ticket")
        if split:
            split.status = "FLAGGED"
            split.updated_at = datetime.utcnow()
            ticket.is_flagged = True
            ticket.flagged_reason = flag_reason.strip()
            _sync_ticket_cache(db, ticket)
            _log(db, ticket_id, user.id, "FLAGGED", f"[{split.split_label}] {flag_reason.strip()}")
        else:
            ticket.is_flagged    = True
            ticket.flagged_reason = flag_reason.strip()
            _log(db, ticket_id, user.id, "FLAGGED", flag_reason.strip())
        _flag_admins = _admin_ids(db, user.tenant_id)
        _flag_managers = _manager_ids_for(db, ticket.current_assignee_id)
        notify_fms_flagged(
            db, user.tenant_id, ticket, _flag_admins, _flag_managers,
            flag_reason.strip(), user.name, actor_id=user.id)

    elif action == "unflag":
        if user.role == "EMPLOYEE":
            current_assignee = (split.assignee_id if split else ticket.current_assignee_id)
            if current_assignee != user.id:
                raise HTTPException(status_code=403, detail="Only the current assignee can unflag this ticket")
        if split and split.status == "FLAGGED":
            split.status = "ACTIVE"
            split.updated_at = datetime.utcnow()
            _sync_ticket_cache(db, ticket)
            other_splits = _active_splits(db, ticket_id)
            ticket.is_flagged = any(s.status == "FLAGGED" for s in other_splits)
            if not ticket.is_flagged:
                ticket.flagged_reason = None
            _log(db, ticket_id, user.id, "UNFLAGGED", f"[{split.split_label}]")
        else:
            ticket.is_flagged     = False
            ticket.flagged_reason = None
            _log(db, ticket_id, user.id, "UNFLAGGED")

    elif action == "on_hold":
        if user.role not in ("ADMIN", "MANAGER"):
            raise HTTPException(403, "Managers only")
        ticket.status = "ON_HOLD"
        _log(db, ticket_id, user.id, "ON_HOLD", reason)

    elif action == "resume":
        if user.role not in ("ADMIN", "MANAGER"):
            raise HTTPException(403, "Managers only")
        ticket.status = "ACTIVE"
        ticket.pause_reason = None
        ticket.linked_child_ticket_id = None
        ticket.is_flagged = False
        ticket.flagged_reason = None
        _log(db, ticket_id, user.id, "RESUMED")

    elif action == "send_to_linked_flow":
        if user.role not in ("ADMIN", "MANAGER"):
            raise HTTPException(403, "Managers only")
        stage = ticket.current_stage
        target_flow = _resolve_linked_flow(db, user.tenant_id, stage.linked_library_flow_id if stage else None)
        if not target_flow:
            raise HTTPException(400, "This stage has no linked flow configured (or it isn't deployed to this tenant)")
        linked = _spawn_linked_ticket(db, ticket, target_flow, user,
                                       title=f"{ticket.title} — linked from {ticket.display_id}")
        linked.linked_parent_ticket_id = ticket.id
        ticket.status = "ON_HOLD"
        ticket.linked_child_ticket_id = linked.id
        ticket.pause_reason = f"Waiting on {linked.display_id} ({target_flow.name})"
        _log(db, ticket_id, user.id, "SENT_TO_LINKED_FLOW",
             f"Spawned {linked.display_id} on '{target_flow.name}'" + (f" | Reason: {reason.strip()}" if reason.strip() else ""))

    elif action == "close_and_continue":
        if user.role not in ("ADMIN", "MANAGER"):
            raise HTTPException(403, "Managers only")
        target_flow = _resolve_linked_flow(db, user.tenant_id, ticket.flow.next_library_flow_id)
        if not target_flow:
            raise HTTPException(400, "This flow has no continuation flow configured (or it isn't deployed to this tenant)")
        ticket.status    = "CLOSED"
        ticket.closed_at = datetime.utcnow()
        _log(db, ticket_id, user.id, "CLOSED", reason)
        continuation = _spawn_linked_ticket(db, ticket, target_flow, user)
        continuation.continued_from_ticket_id = ticket.id
        ticket.continued_to_ticket_id = continuation.id
        _log(db, ticket_id, user.id, "CONTINUED", f"Continued as {continuation.display_id} on '{target_flow.name}'")
        _fms_admins = _admin_ids(db, user.tenant_id)
        _fms_managers = _manager_ids_for(db, ticket.current_assignee_id)
        send_whatsapp_for_fms_ticket_closed(db, user.tenant_id, ticket, _fms_admins, _fms_managers, user.name)

    elif action == "close":
        if user.role not in ("ADMIN", "MANAGER"):
            raise HTTPException(403, "Managers only")
        ticket.status    = "CLOSED"
        ticket.closed_at = datetime.utcnow()
        _log(db, ticket_id, user.id, "CLOSED", reason)
        _notify_linked_parent_if_ready(db, ticket)
        _fms_admins = _admin_ids(db, user.tenant_id)
        _fms_managers = _manager_ids_for(db, ticket.current_assignee_id)
        send_whatsapp_for_fms_ticket_closed(db, user.tenant_id, ticket, _fms_admins, _fms_managers, user.name)
        if closing_file and closing_file.filename:
            from .uploads import save_upload as _save_upload_close
            info = await _save_upload_close(closing_file, user.tenant_id)
            db.add(MediaUpload(
                tenant_id=user.tenant_id, entity_type="fms_ticket", entity_id=ticket_id,
                uploaded_by_id=user.id, **info,
            ))
            _log(db, ticket_id, user.id, "PROOF_UPLOADED", info["file_name"])

    elif action == "close_split":
        # Close a single split without touching the others — the ticket's
        # target_qty vs. remaining-active-splits qty may no longer match once
        # this split's qty drops out of the active pool; that's surfaced via
        # the existing qty-discrepancy flag rather than blocked here.
        if not split:
            raise HTTPException(400, "split_id is required to close a split")
        if user.role not in ("ADMIN", "MANAGER") and split.current_assignee_id != user.id:
            raise HTTPException(403, "Not authorised to close this split")
        if not reason.strip():
            raise HTTPException(400, "A reason is required to close a split")
        open_h = _open_history(db, ticket_id, split_id=split.id)
        if open_h:
            open_h.exited_at = datetime.utcnow()
            open_h.completion_note = reason.strip()
        stage_name = split.current_stage.name if split.current_stage else None
        split.status = "CLOSED"
        split.is_deleted = True
        split.updated_at = datetime.utcnow()
        _sync_ticket_cache(db, ticket)
        _check_qty_discrepancy(db, ticket, user.id)
        _log(db, ticket_id, user.id, "SPLIT_CLOSED",
             f"[{split.split_label}] Closed at {stage_name or '?'} | Qty: {split.qty or 0} | Reason: {reason.strip()}",
             meta={"stage_name": stage_name, "qty": split.qty, "reason": reason.strip()})

    elif action == "mark_stage_complete":
        if not _can_transition(user, ticket, split):
            raise HTTPException(403)
        if split:
            split.status = "STAGE_COMPLETE"
            split.updated_at = datetime.utcnow()
            _sync_ticket_cache(db, ticket)
            _log(db, ticket_id, user.id, "STAGE_EXITED",
                 f"[{split.split_label}] Marked complete at {split.current_stage.name if split.current_stage else '?'}")
        else:
            ticket.status = "STAGE_COMPLETE"
            _log(db, ticket_id, user.id, "STAGE_EXITED",
                 f"Marked complete at {ticket.current_stage.name if ticket.current_stage else '?'}")

    ticket.updated_at = datetime.utcnow()
    db.commit()
    # Redirect back to stage view at the ticket's current stage
    return _redirect(
        f"/fms/dashboard?view=stage"
        f"{'&flow_id=' + ticket.flow_id if ticket.flow_id else ''}"
        f"{'&stage_id=' + ticket.current_stage_id if ticket.current_stage_id else ''}"
    )


@router.post("/tickets/{ticket_id}/help_request")
def fms_help_request(
    ticket_id: str,
    reason: str = Form(...),
    helper_id: str = Form(""),
    split_id: str = Form(""),
    user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """Phase A2: Help Needed popup — set HELP_REQUESTED, notify admin/manager.
    Phase 0: split_id optional — omitted keeps today's ticket-level behavior."""
    ticket = _get_ticket(db, ticket_id, user.tenant_id)
    if not reason.strip():
        raise HTTPException(400, "Reason is required")
    split = None
    if split_id:
        split = db.query(FMSTicketSplit).filter(
            FMSTicketSplit.id == split_id, FMSTicketSplit.ticket_id == ticket_id,
            FMSTicketSplit.is_deleted == False).first()
    if split:
        split.status = "HELP_REQUESTED"
        split.updated_at = datetime.utcnow()
        _sync_ticket_cache(db, ticket)
        _log(db, ticket_id, user.id, "HELP_REQUESTED", f"[{split.split_label}] {reason.strip()}")
    else:
        ticket.status = "HELP_REQUESTED"
        _log(db, ticket_id, user.id, "HELP_REQUESTED", reason.strip())
    # Notify configured recipients (default admin + manager) — in-app + push + WhatsApp
    from .notification_rules import filter_recipients
    admins = _admin_ids(db, user.tenant_id)
    mgrs   = _manager_ids_for(db, ticket.current_assignee_id)
    recipients = filter_recipients(
        db, user.tenant_id, "fms_help_needed",
        admin_ids=admins, manager_ids=mgrs,
        assignee_id=ticket.current_assignee_id, actor_id=user.id,
    )
    for uid in recipients:
        create_notification(
            db, user.tenant_id,
            user_id=uid,
            notif_type="FMS_HELP_NEEDED",
            title=f"Help needed: {ticket.title}",
            body=f"{user.name} needs help on {ticket.display_id or ticket_id}. Reason: {reason[:200]}",
            link=f"/fms/dashboard?view=stage&flow_id={ticket.flow_id}&stage_id={ticket.current_stage_id}",
            condition_key="fms_help_needed",
        )
    if channel_enabled(db, user.tenant_id, "fms_help_needed", "whatsapp"):
        from .notifications import _send_gupshup_wa
        try:
            for uid in recipients:
                recipient = db.query(User).filter(User.id == uid).first()
                if not recipient or not recipient.phone:
                    continue
                variables = [recipient.name, ticket.title, user.name, reason[:200]]
                _send_gupshup_wa(db, user.tenant_id, recipient, "omniflow_fms_help_needed", variables,
                                  related_entity_type="fms_ticket", related_entity_id=ticket_id,
                                  event_key="fms_help_needed")
        except Exception:
            pass
    if helper_id:
        existing = db.query(FMSTicketHelper).filter(
            FMSTicketHelper.ticket_id == ticket_id,
            FMSTicketHelper.user_id == helper_id).first()
        if not existing:
            db.add(FMSTicketHelper(
                ticket_id=ticket_id, user_id=helper_id,
                added_by_id=user.id, reason=reason.strip()))
    ticket.updated_at = datetime.utcnow()
    db.commit()
    return _redirect(
        f"/fms/dashboard?view=stage"
        f"{'&flow_id=' + ticket.flow_id if ticket.flow_id else ''}"
        f"{'&stage_id=' + ticket.current_stage_id if ticket.current_stage_id else ''}"
        f"&msg=Help+request+sent"
    )


# ── 2-F: FMS Analytics ───────────────────────────────────────────────────────

@router.get("/analytics", response_class=HTMLResponse)
def fms_analytics(
    request: Request,
    user: User = Depends(get_current_user_or_redirect),
    db: Session = Depends(get_db),
):
    """2-F-1/2: TaT breach rates and stage compliance per employee."""
    tid    = user.tenant_id
    tenant = db.query(Tenant).get(tid)
    is_pro = has_feature(tenant, "KPI_CHARTS_ADMIN", db)

    # 2-F-2: My own compliance (all roles)
    my_history = db.query(FMSStageHistory).join(
        FMSTicket, FMSStageHistory.ticket_id == FMSTicket.id
    ).join(FMSStage, FMSStageHistory.stage_id == FMSStage.id).filter(
        FMSTicket.tenant_id == tid,
        FMSStageHistory.assignee_id == user.id,
        FMSStageHistory.exited_at != None,
        FMSStage.target_tat_hours != None,
    ).all()
    my_total  = len(my_history)
    my_ontime = sum(
        1 for h in my_history
        if h.stage and h.stage.target_tat_hours and
           (h.exited_at - h.entered_at).total_seconds() / 3600 <= h.stage.target_tat_hours
    )
    my_compliance = int(my_ontime / my_total * 100) if my_total else 0

    # 2-F-1: Per-employee TaT analysis (Professional+ admin/manager only)
    emp_analytics = []
    if is_pro and user.role in ("ADMIN", "MANAGER"):
        employees = db.query(User).filter(
            User.tenant_id == tid, User.is_deleted == False).all()
        for emp in employees:
            emp_hist = db.query(FMSStageHistory).join(
                FMSTicket, FMSStageHistory.ticket_id == FMSTicket.id
            ).join(FMSStage, FMSStageHistory.stage_id == FMSStage.id).filter(
                FMSTicket.tenant_id == tid,
                FMSStageHistory.assignee_id == emp.id,
                FMSStageHistory.exited_at != None,
                FMSStage.target_tat_hours != None,
            ).all()
            if not emp_hist:
                continue
            breaches  = sum(
                1 for h in emp_hist
                if h.stage and h.stage.target_tat_hours and
                   (h.exited_at - h.entered_at).total_seconds() / 3600
                   > h.stage.target_tat_hours
            )
            ontime_pct = int((1 - breaches / len(emp_hist)) * 100) if emp_hist else 100
            emp_analytics.append({
                "user": emp,
                "total_stages": len(emp_hist),
                "breaches": breaches,
                "ontime_pct": ontime_pct,
            })
        emp_analytics.sort(key=lambda x: x["ontime_pct"])

    # Per-flow breach counts (for summary)
    flows = db.query(FMSFlow).filter(
        FMSFlow.tenant_id == tid, FMSFlow.is_deleted == False).all()
    flow_breaches = {}
    for f in flows:
        fh = db.query(FMSStageHistory).join(
            FMSTicket, FMSStageHistory.ticket_id == FMSTicket.id
        ).join(FMSStage, FMSStageHistory.stage_id == FMSStage.id).filter(
            FMSTicket.flow_id == f.id,
            FMSStageHistory.exited_at != None,
            FMSStage.target_tat_hours != None,
        ).all()
        flow_breaches[f.id] = sum(
            1 for h in fh
            if h.stage and h.stage.target_tat_hours and
               (h.exited_at - h.entered_at).total_seconds() / 3600
               > h.stage.target_tat_hours
        )

    return templates.TemplateResponse(request, "fms/analytics.html", _ctx(
        request, user, db,
        my_total=my_total, my_ontime=my_ontime, my_compliance=my_compliance,
        emp_analytics=emp_analytics, is_pro=is_pro,
        flows=flows, flow_breaches=flow_breaches,
    ))


# ── Internal helpers ──────────────────────────────────────────────────────────

def _get_deployed_submodules(db: Session, tenant_id: str) -> list:
    """Return list of LibrarySubmoduleDefinition deployed to this tenant.
    Used by the flow editor to populate the sub-module dropdown."""
    deployed_ids = [
        row.library_item_id for row in
        db.query(TenantDeployedItem).filter(
            TenantDeployedItem.tenant_id == tenant_id,
            TenantDeployedItem.item_type == "submodule",
        ).all()
    ]
    if not deployed_ids:
        return []
    return db.query(LibrarySubmoduleDefinition).filter(
        LibrarySubmoduleDefinition.id.in_(deployed_ids),
        LibrarySubmoduleDefinition.status == "PUBLISHED",
    ).order_by(
        LibrarySubmoduleDefinition.sub_module_type,
        LibrarySubmoduleDefinition.name,
    ).all()


def _get_flow(db: Session, flow_id: str, tenant_id: str) -> FMSFlow:
    f = db.query(FMSFlow).filter(
        FMSFlow.id == flow_id, FMSFlow.tenant_id == tenant_id,
        FMSFlow.is_deleted == False).first()
    if not f:
        raise HTTPException(404, "Flow not found")
    return f


def _save_stages(db: Session, flow_id: str, tenant_id: str, stages_json: str):
    """Parse stages JSON from the stage editor and insert FMSStage rows."""
    try:
        stages = _json.loads(stages_json) if stages_json else []
    except Exception:
        stages = []
    for i, s in enumerate(stages):
        name = (s.get("name") or "").strip()
        if not name:
            continue
        smt = (s.get("sub_module_tag") or "").strip().upper() or None
        assignee_ids = [a for a in (s.get("default_assignee_ids") or []) if a]
        if not assignee_ids and s.get("default_assignee_id"):
            assignee_ids = [s.get("default_assignee_id")]
        db.add(FMSStage(
            flow_id=flow_id, tenant_id=tenant_id,
            name=name, order=s.get("order", i),
            color=s.get("color", "#3b82f6"),
            target_tat_hours=s.get("target_tat_hours") or None,
            default_assignee_id=(assignee_ids[0] if assignee_ids else None),
            default_assignee_ids_json=(_json.dumps(assignee_ids) if assignee_ids else None),
            sub_module_tag=smt,
            is_mandatory=bool(s.get("is_mandatory", True)),
            completion_note_required=bool(s.get("completion_note_required", False)),
            is_terminal=bool(s.get("is_terminal", False)),
        ))

@router.post("/tickets/{ticket_id}/stage-data")
async def fms_save_stage_data(
    ticket_id: str,
    request: Request,
    user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """Save custom stage column data to a split's current open stage history row.

    Phase 0: body may include split_id — omitted (today's caller) resolves
    via _ensure_ticket_has_split, unchanged for single-split tickets. Without
    this, a multi-split ticket would silently write into whichever open
    history row the old ticket.current_stage_id cache happened to point at
    (or an arbitrary "any open row" fallback) — the wrong split entirely
    whenever the ticket has more than one active split."""
    from fastapi.responses import JSONResponse
    ticket = db.query(FMSTicket).filter(
        FMSTicket.id == ticket_id,
        FMSTicket.tenant_id == user.tenant_id,
        FMSTicket.is_deleted == False,
    ).first()
    if not ticket:
        return JSONResponse({"ok": False, "error": "Not found"}, status_code=404)
    import json as _json
    body = await request.json()
    incoming = body.get("data", {})
    req_split_id = (body.get("split_id") or "").strip()

    if req_split_id:
        split = db.query(FMSTicketSplit).filter(
            FMSTicketSplit.id == req_split_id,
            FMSTicketSplit.ticket_id == ticket_id,
            FMSTicketSplit.is_deleted == False,
        ).first()
        if not split:
            return JSONResponse({"ok": False, "error": "Split not found"}, status_code=404)
    else:
        split = _ensure_ticket_has_split(db, ticket)

    history = _open_history(db, ticket_id, split_id=split.id)
    if not history:
        # Create one if truly missing (edge case for legacy tickets)
        history = FMSStageHistory(
            id=new_id(), ticket_id=ticket_id, split_id=split.id,
            stage_id=split.current_stage_id, entered_at=datetime.utcnow(),
        )
        db.add(history)
    existing = _json.loads(history.custom_fields_data_json or "{}") if history.custom_fields_data_json else {}
    existing.update(incoming)

    # Evaluate formula columns using cross-stage values so references to prior
    # stage columns resolve (the client can only see current-stage inputs).
    cur_stage = split.current_stage
    if cur_stage and cur_stage.custom_fields_json:
        try:
            field_defs = _json.loads(cur_stage.custom_fields_json)
        except Exception:
            field_defs = []
        all_flow_stages = db.query(FMSStage).filter(
            FMSStage.flow_id == ticket.flow_id, FMSStage.is_deleted == False
        ).all()
        _tff = {}
        if ticket.ticket_custom_fields_json:
            try:
                _tff = _json.loads(ticket.ticket_custom_fields_json)
            except Exception:
                pass
        formula_lookup = {
            **_tff,
            **_cross_stage_cf(db, ticket.id, all_flow_stages, split_id=_split_lineage_ids(db, ticket.id, split.id), exclude_history_id=history.id),
            **existing,
        }
        # The split "actual" field is entered as THIS VISIT'S increment
        # (brief §5), but formulas like "Short Quantity" need the running
        # cumulative to compare against target — swap in the cumulative for
        # formula evaluation only; `existing` (what's actually persisted for
        # the field itself) keeps the raw incremental value the user typed.
        if getattr(cur_stage, "split_enabled", False) and cur_stage.split_actual_field:
            afield = cur_stage.split_actual_field
            if afield in existing:
                try:
                    delta_val = float(existing[afield])
                    formula_lookup[afield] = (split.last_cumulative_entered or 0) + delta_val
                except (TypeError, ValueError):
                    pass

        def _eval_sd_formula(steps):
            result = None
            for i, step in enumerate(steps):
                raw = formula_lookup.get(step.get("col_id", ""), "")
                try:
                    val = float(raw)
                except (ValueError, TypeError):
                    return None
                if i == 0:
                    result = val; continue
                op = step.get("op", "+")
                if op == "+":   result += val
                elif op == "-": result -= val
                elif op == "*": result *= val
                elif op == "/":
                    if val == 0: return None
                    result /= val
            if result is None: return None
            return str(int(result)) if result == int(result) else f"{result:.4f}".rstrip("0")

        for fdef in field_defs:
            if fdef.get("field_type") != "formula": continue
            computed = _eval_sd_formula(fdef.get("formula_steps") or [])
            if computed is not None:
                existing[fdef.get("id", "")] = computed

    history.custom_fields_data_json = _json.dumps(existing)
    db.commit()
    return JSONResponse({"ok": True, "computed": {k: v for k, v in existing.items()}})


@router.post("/splits/{split_id}/evidence")
async def fms_upload_split_evidence(
    split_id: str,
    file: UploadFile = File(...),
    user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """R7: optional per-split evidence upload — photo/pdf/audio/video, never
    mandatory, traceable to the specific split (not the parent ticket).
    Stored on local disk via app/uploads.py::save_upload — this repo has no
    Cloudinary integration anywhere (see deviation note in final report)."""
    split = db.query(FMSTicketSplit).filter(
        FMSTicketSplit.id == split_id,
        FMSTicketSplit.tenant_id == user.tenant_id,
        FMSTicketSplit.is_deleted == False,
    ).first()
    if not split:
        return JSONResponse({"ok": False, "error": "Split not found"}, status_code=404)

    from .uploads import save_upload as _save_upload
    result = await _save_upload(file, user.tenant_id)

    ctype = (file.content_type or "").lower()
    if ctype.startswith("image/"):
        ftype = "photo"
    elif ctype.startswith("audio/"):
        ftype = "audio"
    elif ctype.startswith("video/"):
        ftype = "video"
    else:
        ftype = "pdf"

    ev = FMSSplitEvidence(
        tenant_id=user.tenant_id, split_id=split.id,
        file_type=ftype, file_url=result["file_path"],
        file_name=result["file_name"], uploaded_by=user.id,
    )
    db.add(ev)
    db.commit()
    return JSONResponse({"ok": True, "evidence": {
        "id": ev.id, "file_type": ev.file_type,
        "file_url": ev.file_url, "file_name": ev.file_name,
    }})


@router.post("/tickets/{ticket_id}/cell-edit")
async def fms_table_cell_edit(
    ticket_id: str,
    request: Request,
    user: User = Depends(require_manager),
    db: Session = Depends(get_db),
):
    """Manual edit of a ticket/stage custom-column value directly from the
    Table view. Requires a reason (audit trail via FMSFieldEditLog). Any
    formula columns (any stage, or ticket-level) that reference the edited
    field are recalculated and persisted too, so dependent columns never go
    stale after a manual correction."""
    from fastapi.responses import JSONResponse
    import json as _json
    ticket = db.query(FMSTicket).filter(
        FMSTicket.id == ticket_id, FMSTicket.tenant_id == user.tenant_id,
        FMSTicket.is_deleted == False,
    ).first()
    if not ticket:
        raise HTTPException(404, "Ticket not found")

    body = await request.json()
    stage_id = (body.get("stage_id") or "").strip()
    field_id = (body.get("field_id") or "").strip()
    new_value = str(body.get("value", "")).strip()
    reason = (body.get("reason") or "").strip()
    if not field_id:
        raise HTTPException(400, "field_id is required")
    if not reason:
        raise HTTPException(400, "A reason is required for this edit")

    all_stages = db.query(FMSStage).filter(
        FMSStage.flow_id == ticket.flow_id, FMSStage.is_deleted == False,
    ).all()

    def _field_defs_for(sid):
        if not sid:
            return (_json.loads(ticket.flow.ticket_form_fields_json)
                    if ticket.flow and ticket.flow.ticket_form_fields_json else [])
        st = next((s for s in all_stages if s.id == sid), None)
        if not st or not st.custom_fields_json:
            return []
        try:
            return _json.loads(st.custom_fields_json)
        except Exception:
            return []

    fdef = next((f for f in _field_defs_for(stage_id) if f.get("id") == field_id), None)
    if not fdef:
        raise HTTPException(404, "Column not found on this ticket/stage")
    if fdef.get("field_type") == "formula":
        raise HTTPException(400, "Calculated columns can't be edited directly — edit one of the columns it's built from")

    def _latest_history_for_stage(sid):
        return (
            db.query(FMSStageHistory)
            .filter(FMSStageHistory.ticket_id == ticket_id, FMSStageHistory.stage_id == sid)
            .order_by(FMSStageHistory.entered_at.desc())
            .first()
        )

    # 1. Apply the direct edit.
    if not stage_id:
        try:
            tff = _json.loads(ticket.ticket_custom_fields_json or "{}")
        except Exception:
            tff = {}
        old_value = tff.get(field_id)
        tff[field_id] = new_value
        ticket.ticket_custom_fields_json = _json.dumps(tff)
    else:
        history = _latest_history_for_stage(stage_id)
        if not history:
            raise HTTPException(400, "This stage hasn't been visited yet — nothing to edit")
        try:
            hdata = _json.loads(history.custom_fields_data_json or "{}")
        except Exception:
            hdata = {}
        old_value = hdata.get(field_id)
        hdata[field_id] = new_value
        history.custom_fields_data_json = _json.dumps(hdata)

    db.add(FMSFieldEditLog(
        tenant_id=user.tenant_id, ticket_id=ticket_id, stage_id=stage_id or None,
        field_id=field_id, field_label=fdef.get("label", field_id),
        old_value=old_value, new_value=new_value, reason=reason,
        edited_by_id=user.id,
    ))
    _log(db, ticket_id, user.id, "FIELD_EDITED",
         f"{fdef.get('label', field_id)}: {old_value or '—'} → {new_value or '—'} ({reason})")

    # 2. Cascade: recompute every formula column (any stage, plus ticket-level
    # fields feed them too) against the fresh merged value set, fixed-point
    # iterating so multi-step chains (A -> B -> C) settle. Only persist+log
    # the ones whose value actually changed.
    def _merged_cf():
        merged = {}
        try:
            merged.update(_json.loads(ticket.ticket_custom_fields_json or "{}"))
        except Exception:
            pass
        for s in all_stages:
            h = _latest_history_for_stage(s.id)
            if h and h.custom_fields_data_json:
                try:
                    merged.update(_json.loads(h.custom_fields_data_json))
                except Exception:
                    pass
        return merged

    def _eval_formula(steps, lookup):
        result = None
        for i, step in enumerate(steps):
            raw = lookup.get(step.get("col_id", ""), "")
            try:
                val = float(raw)
            except (ValueError, TypeError):
                return None
            if i == 0:
                result = val
                continue
            op = step.get("op", "+")
            if op == "+":   result += val
            elif op == "-": result -= val
            elif op == "*": result *= val
            elif op == "/":
                if val == 0: return None
                result /= val
        if result is None:
            return None
        return str(int(result)) if result == int(result) else f"{result:.4f}".rstrip("0")

    try:
        ticket_field_defs = (_json.loads(ticket.flow.ticket_form_fields_json)
                             if ticket.flow and ticket.flow.ticket_form_fields_json else [])
    except Exception:
        ticket_field_defs = []

    cascaded = []
    for _pass in range(5):
        changed_this_pass = False
        lookup = _merged_cf()

        # Ticket-level formula columns.
        ticket_formula_defs = [f for f in ticket_field_defs if f.get("field_type") == "formula"]
        if ticket_formula_defs:
            try:
                tff = _json.loads(ticket.ticket_custom_fields_json or "{}")
            except Exception:
                tff = {}
            row_changed = False
            for f in ticket_formula_defs:
                fid = f.get("id", "")
                computed = _eval_formula(f.get("formula_steps") or [], lookup)
                if computed is None:
                    continue
                if tff.get(fid) != computed:
                    old = tff.get(fid)
                    tff[fid] = computed
                    row_changed = True
                    changed_this_pass = True
                    cascaded.append({"field_id": fid, "stage_id": None, "value": computed})
                    db.add(FMSFieldEditLog(
                        tenant_id=user.tenant_id, ticket_id=ticket_id, stage_id=None,
                        field_id=fid, field_label=f.get("label", fid),
                        old_value=old, new_value=computed,
                        reason=f"Auto-recalculated after '{fdef.get('label', field_id)}' was edited",
                        is_cascade=True, edited_by_id=user.id,
                    ))
            if row_changed:
                ticket.ticket_custom_fields_json = _json.dumps(tff)

        for s in all_stages:
            try:
                defs = _json.loads(s.custom_fields_json or "[]")
            except Exception:
                defs = []
            formula_defs = [f for f in defs if f.get("field_type") == "formula"]
            if not formula_defs:
                continue
            h = _latest_history_for_stage(s.id)
            if not h:
                continue
            try:
                hdata = _json.loads(h.custom_fields_data_json or "{}")
            except Exception:
                hdata = {}
            row_changed = False
            for f in formula_defs:
                fid = f.get("id", "")
                computed = _eval_formula(f.get("formula_steps") or [], lookup)
                if computed is None:
                    continue
                if hdata.get(fid) != computed:
                    old = hdata.get(fid)
                    hdata[fid] = computed
                    row_changed = True
                    changed_this_pass = True
                    cascaded.append({"field_id": fid, "stage_id": s.id, "value": computed})
                    db.add(FMSFieldEditLog(
                        tenant_id=user.tenant_id, ticket_id=ticket_id, stage_id=s.id,
                        field_id=fid, field_label=f.get("label", fid),
                        old_value=old, new_value=computed,
                        reason=f"Auto-recalculated after '{fdef.get('label', field_id)}' was edited",
                        is_cascade=True, edited_by_id=user.id,
                    ))
            if row_changed:
                h.custom_fields_data_json = _json.dumps(hdata)
        if not changed_this_pass:
            break

    db.commit()
    return JSONResponse({
        "ok": True,
        "field_id": field_id, "stage_id": stage_id or None, "value": new_value,
        "edited_by": user.name, "edited_at": datetime.utcnow().strftime("%d %b %Y, %H:%M"),
        "reason": reason,
        "cascaded": cascaded,
    })


@router.get("/tickets/{ticket_id}/cf-carry-forward")
def fms_cf_carry_forward(
    ticket_id: str,
    stage_id: str = Query(...),
    user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """Return carry-forward pre-fill values for a ticket entering a stage.
    Looks up stage custom_fields_json, finds columns with carry_forward=true,
    then searches stage history for the most recent value per column id.
    Returns {column_id: value} for all columns that have a prior value.
    """
    import json as _json
    ticket = _get_ticket(db, ticket_id, user.tenant_id)
    stage = db.query(FMSStage).filter(
        FMSStage.id == stage_id,
        FMSStage.tenant_id == user.tenant_id,
    ).first()
    if not stage:
        return {"values": {}}

    try:
        field_defs = _json.loads(stage.custom_fields_json or "[]")
    except Exception:
        field_defs = []

    cf_ids = [f["id"] for f in field_defs if f.get("carry_forward") and f.get("id")]
    if not cf_ids:
        return {"values": {}}

    histories = db.query(FMSStageHistory).filter(
        FMSStageHistory.ticket_id == ticket_id,
        FMSStageHistory.custom_fields_data_json.isnot(None),
    ).order_by(FMSStageHistory.entered_at.desc()).all()

    values = {}
    remaining = set(cf_ids)
    for h in histories:
        if not remaining:
            break
        try:
            data = _json.loads(h.custom_fields_data_json)
        except Exception:
            continue
        for cid in list(remaining):
            if cid in data:
                values[cid] = data[cid]
                remaining.discard(cid)

    # Fall back to values captured on the ticket creation form (e.g. columns
    # reused from the Ticket Creation Form have no stage history to draw from).
    if remaining and ticket.ticket_custom_fields_json:
        try:
            tff = _json.loads(ticket.ticket_custom_fields_json)
        except Exception:
            tff = {}
        for cid in list(remaining):
            if cid in tff:
                values[cid] = tff[cid]
                remaining.discard(cid)

    return {"values": values}
